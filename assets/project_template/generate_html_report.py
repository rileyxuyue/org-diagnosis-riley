#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
组织诊断报告生成器 - HTML版本
生成漂亮的HTML报告，可在浏览器中查看和打印为PDF
"""

import zipfile
import xml.etree.ElementTree as ET
import re
import json
import math
import requests
from typing import List, Dict, Tuple, Optional
from datetime import datetime
import os
import webbrowser
from collections import Counter

class ExcelReader:
    """Excel读取器"""
    
    def __init__(self, filepath, sheet_name=None):
        self.filepath = filepath
        self.sheet_name = sheet_name
        self.shared_strings = []
        self.rows = []
        self.headers = []
        self._load()
    
    def _find_sheet_file(self, zf):
        if self.sheet_name is None:
            return 'xl/worksheets/sheet1.xml'
        
        with zf.open('xl/workbook.xml') as f:
            tree = ET.parse(f)
            root = tree.getroot()
            ns = '{http://schemas.openxmlformats.org/spreadsheetml/2006/main}'
            r_ns = '{http://schemas.openxmlformats.org/officeDocument/2006/relationships}'
            
            for sheet in root.findall(f'.//{ns}sheet'):
                if sheet.get('name') == self.sheet_name:
                    rid = sheet.get(f'{r_ns}id')
                    with zf.open('xl/_rels/workbook.xml.rels') as rf:
                        rel_tree = ET.parse(rf)
                        rel_root = rel_tree.getroot()
                        rel_ns = '{http://schemas.openxmlformats.org/package/2006/relationships}'
                        for rel in rel_root.findall(f'{rel_ns}Relationship'):
                            if rel.get('Id') == rid:
                                return f'xl/{rel.get("Target")}'
        return 'xl/worksheets/sheet1.xml'
    
    def _load(self):
        with zipfile.ZipFile(self.filepath) as zf:
            try:
                with zf.open('xl/sharedStrings.xml') as f:
                    tree = ET.parse(f)
                    root = tree.getroot()
                    ns = '{http://schemas.openxmlformats.org/spreadsheetml/2006/main}'
                    for si in root.findall(f'{ns}si'):
                        texts = []
                        for t in si.iter(f'{ns}t'):
                            if t.text:
                                texts.append(t.text)
                        self.shared_strings.append(''.join(texts))
            except:
                pass
            
            sheet_file = self._find_sheet_file(zf)
            with zf.open(sheet_file) as f:
                tree = ET.parse(f)
                root = tree.getroot()
                ns = '{http://schemas.openxmlformats.org/spreadsheetml/2006/main}'
                
                for row_elem in root.findall(f'.//{ns}row'):
                    cells_dict = {}
                    for cell in row_elem.findall(f'{ns}c'):
                        cell_ref = cell.get('r')
                        col_letter = ''.join(c for c in cell_ref if c.isalpha())
                        col_idx = sum((ord(c) - ord('A') + 1) * (26 ** i) 
                                    for i, c in enumerate(reversed(col_letter))) - 1
                        
                        cell_type = cell.get('t')
                        v_elem = cell.find(f'{ns}v')
                        value = None
                        
                        if cell_type == 'inlineStr':
                            is_elem = cell.find(f'{ns}is')
                            if is_elem is not None:
                                texts = []
                                for t in is_elem.iter(f'{ns}t'):
                                    if t.text:
                                        texts.append(t.text)
                                value = ''.join(texts)
                        elif v_elem is not None:
                            value = v_elem.text or ''
                            if cell_type == 's' and value and self.shared_strings:
                                try:
                                    value = self.shared_strings[int(value)]
                                except:
                                    pass
                        
                        if value is not None:
                            cells_dict[col_idx] = value
                    
                    if cells_dict:
                        max_col = max(cells_dict.keys())
                        row_data = [cells_dict.get(i, '') for i in range(max_col + 1)]
                        self.rows.append(row_data)
                
                if self.rows:
                    self.headers = self.rows[0]
    
    def find_rows(self, col_idx: int, keyword: str) -> List[Tuple[int, List]]:
        results = []
        for i, row in enumerate(self.rows):
            if len(row) > col_idx:
                cell_value = str(row[col_idx])
                if keyword.lower() in cell_value.lower():
                    results.append((i, row))
        return results
    
    def get_row(self, row_idx: int) -> List:
        if 0 <= row_idx < len(self.rows):
            return self.rows[row_idx]
        return []
    
    def get_cell(self, row_idx: int, col_idx: int) -> str:
        row = self.get_row(row_idx)
        if len(row) > col_idx:
            return str(row[col_idx])
        return ''


def simplify_org_name(full_name: str) -> str:
    if not full_name:
        return ""
    parts = full_name.split('/')
    if len(parts) > 0:
        first_part = parts[0]
        
        # 保留 S1/S2/S3 等职能系统前缀
        if re.match(r'^S\d+', first_part):
            # 保持原样，不简化
            pass
        else:
            # 其他情况，只保留开头的大写字母
            match = re.search(r'^([A-Z]+)', first_part)
            if match:
                parts[0] = match.group(1)
    return '/'.join(parts)


def format_percentage(value, is_rank=False) -> str:
    if not value or value == '' or value == 'NA':
        return "N/A"
    
    # 处理Excel错误值
    value_str = str(value).strip()
    if value_str.startswith('#') or value_str in ['#NUM!', '#DIV/0!', '#VALUE!', '#REF!', '#NAME?', '#N/A', '#NULL!']:
        return "N/A"
    
    try:
        val = float(value)
        if is_rank and val > 1/3:
            back_pct = round((1 - val) * 100)
            if back_pct == 0:
                back_pct = 1
            return f"后{back_pct}%"
        else:
            pct = round(val * 100)
            return f"前{pct}%" if is_rank else f"{pct}%"
    except (ValueError, TypeError):
        return "N/A"


def color_to_text(color: str) -> str:
    color = str(color).replace('灯', '')
    color_map = {'红': "严重预警", '黄': "预警", '绿': "正常"}
    return color_map.get(color, "未知")


def safe_float(value, default=0.0) -> float:
    """安全地将值转换为float，处理Excel错误值和特殊情况"""
    if not value or value == '' or value == 'NA' or value == 'N/A':
        return default
    
    # 处理Excel错误值
    value_str = str(value).strip()
    if value_str.startswith('#') or value_str in ['#NUM!', '#DIV/0!', '#VALUE!', '#REF!', '#NAME?', '#N/A', '#NULL!']:
        return default
    
    try:
        return float(value)
    except (ValueError, TypeError):
        return default


def safe_int(value, default=0) -> int:
    """安全地将值转换为int，处理Excel错误值和特殊情况"""
    if not value or value == '' or value == 'NA' or value == 'N/A':
        return default
    
    # 处理Excel错误值
    value_str = str(value).strip()
    if value_str.startswith('#') or value_str in ['#NUM!', '#DIV/0!', '#VALUE!', '#REF!', '#NAME?', '#N/A', '#NULL!']:
        return default
    
    try:
        return int(float(value))
    except (ValueError, TypeError):
        return default


# ============================================================
# 离职原因分析模块
# ============================================================

class ResignationAnalyzer:
    """离职原因分析器：读取离职明细Excel，按部门匹配，调用大模型分析"""
    
    LLM_URL = os.environ.get("ORG_DIAG_LLM_URL", "http://127.0.0.1:1234/v1/chat/completions")
    LLM_MODELS = [m.strip() for m in os.environ.get("ORG_DIAG_LLM_MODELS", os.environ.get("ORG_DIAG_LLM_MODEL", "qwen2.5-7b-instruct-mlx")).split(",") if m.strip()]  # 支持本地或线上 OpenAI 兼容模型
    BATCH_THRESHOLD = 5  # 超过此人数就分批调用大模型
    
    def __init__(self, filepath: str):
        """
        Args:
            filepath: 离职明细Excel路径，如 【异动】/WXG25年离职明细.xlsx
        """
        import pandas as pd
        self.pd = pd
        self.filepath = filepath
        self.df = pd.read_excel(filepath, sheet_name=0)
        print(f"✓ 离职明细数据已加载 ({len(self.df)} 行): {os.path.basename(filepath)}")
    
    @staticmethod
    def find_file_for_bg(org_full_path: str) -> Optional[str]:
        """根据org_full_path自动定位离职明细文件，优先按BG找，找不到就用通用模板"""
        if not org_full_path:
            return None
        bg_match = re.match(r'^[A-Za-z]+\d*', org_full_path.split('/')[0])
        if bg_match:
            bg_abbr = bg_match.group()
            filepath = f'【异动】/{bg_abbr}25年离职明细.xlsx'
            if os.path.exists(filepath):
                return filepath
        # fallback to generic template
        fallback = '【异动】/模板25年离职明细.xlsx'
        if os.path.exists(fallback):
            return fallback
        return None
    
    def get_dept_resignations(self, org_full_path: str):
        """根据org_full_path前缀匹配获取该部门所有离职人员"""
        pd = self.pd
        mask = self.df['离职组织全路径'].str.startswith(org_full_path, na=False)
        return self.df[mask].copy()
    
    def classify_type(self, raw_type: str) -> str:
        """离职类型分类：被动离职 vs 主动离职"""
        if raw_type == '被动离职':
            return '被动离职'
        return '主动离职'
    
    def build_pie_data(self, dept_df) -> Dict:
        """构建离职类型分布饼图数据"""
        classified = dept_df['离职类型'].apply(self.classify_type)
        counts = classified.value_counts()
        total = len(dept_df)
        passive = int(counts.get('被动离职', 0))
        active = int(counts.get('主动离职', 0))
        return {
            'total': total,
            'passive': passive,
            'active': active,
            'passive_pct': round(passive / total * 100, 1) if total > 0 else 0,
            'active_pct': round(active / total * 100, 1) if total > 0 else 0,
        }
    
    def _build_person_info(self, row) -> Dict:
        """从一行数据中提取人员信息"""
        pd = self.pd
        def safe_str(val):
            if pd.isna(val) or str(val).strip() in ('', 'nan', '-'):
                return ''
            return str(val).strip()
        
        name = safe_str(row.get('员工中英文名', ''))
        center = safe_str(row.get('中心', ''))
        dept = safe_str(row.get('部门', ''))
        grade = safe_str(row.get('离职时专业职级', ''))
        grade_interval = safe_str(row.get('专业职级区间', ''))
        is_echelon = safe_str(row.get('离职最后是否梯队', ''))
        perf = safe_str(row.get('离职最后绩效等级', ''))
        resign_type = self.classify_type(safe_str(row.get('离职类型', '')))
        
        # Good+ 及以上算高绩效
        high_perf_levels = {'Good+', 'Outstanding', 'Exceptional'}
        is_high_perf = perf in high_perf_levels
        
        return {
            'name': name,
            'center': center if center else dept,
            'grade': grade if grade else grade_interval,
            'is_echelon': is_echelon == '是',
            'is_high_perf': is_high_perf,
            'perf': perf,
            'resign_type': resign_type,
        }
    
    def _build_llm_prompt(self, persons_with_text: List[Dict]) -> str:
        """构建发给大模型的prompt，精简以适应4096 context"""
        prompt = """分析以下离职员工面谈记录。
任务：1.内容去重 2.严格过滤纯个人原因(退休/异地配偶/回高校/家庭照顾等必须过滤) 3.只保留与组织管理相关的原因(团队定位/晋升/压力/管理/协作等) 4.归类

重要：只输出与组织问题相关的类别，纯个人原因不要输出。

严格按JSON返回，不要有其他文字：
```json
[{"category":"归类名","summary":"一句话总结","quotes":["典型原文"],"people":["英文名"]}]
```
无组织相关原因返回[]

员工记录：
"""
        for p in persons_with_text:
            # 截断文本，每人最多250字
            text = p['text'][:250] if len(p['text']) > 250 else p['text']
            prompt += f"\n[{p['eng_name']}|{p['resign_type']}] {text}\n"
        
        return prompt
    
    def _call_llm(self, prompt: str) -> List[Dict]:
        """调用本地大模型（OpenAI 兼容接口）"""
        for model in self.LLM_MODELS:
            try:
                # Qwen2.5-32B-4bit 推理时间较长，给予足够超时
                timeout = 300
                print(f"      尝试模型: {model}...")
                print(f"      尝试模型: {model}...")
                # API key：优先读环境变量 LLM_API_KEY，未设置则使用默认值
                import os
                api_key = os.environ.get("ORG_DIAG_LLM_API_KEY", os.environ.get("LLM_API_KEY", "lm-studio"))
                resp = requests.post(
                    self.LLM_URL,
                    headers={"Authorization": f"Bearer {api_key}"},
                    json={
                        "model": model,
                        "messages": [{"role": "user", "content": prompt}],
                        "temperature": 0.3,
                        "max_tokens": 3500,
                    },
                    timeout=timeout
                )
                resp.raise_for_status()
                content = resp.json()['choices'][0]['message']['content']
                
                # 提取JSON部分（可能被```json包裹）
                json_match = re.search(r'```json\s*(.*?)\s*```', content, re.DOTALL)
                if json_match:
                    json_str = json_match.group(1)
                else:
                    # 尝试直接解析
                    json_str = content.strip()
                    start = json_str.find('[')
                    end = json_str.rfind(']')
                    if start != -1 and end != -1:
                        json_str = json_str[start:end+1]
                
                result = json.loads(json_str)
                if isinstance(result, list):
                    # 过滤掉纯个人原因类别
                    personal_keywords = ['个人原因', '家庭', '退休', '异地', '配偶', '回乡', '回高校']
                    result = [item for item in result 
                              if not any(kw in item.get('category', '') for kw in personal_keywords)]
                    print(f"      ✓ {model} 返回 {len(result)} 个类别")
                    return result
            except requests.exceptions.Timeout:
                print(f"      ⚠ {model} 超时，尝试下一个模型...")
                continue
            except json.JSONDecodeError as e:
                # 尝试修复常见JSON问题
                try:
                    # 移除控制字符
                    cleaned = re.sub(r'[\x00-\x1f\x7f]', ' ', json_str)
                    # 移除尾逗号
                    cleaned = re.sub(r',\s*([}\]])', r'\1', cleaned)
                    result = json.loads(cleaned)
                    if isinstance(result, list):
                        personal_keywords = ['个人原因', '家庭', '退休', '异地', '配偶', '回乡', '回高校']
                        result = [item for item in result 
                                  if not any(kw in item.get('category', '') for kw in personal_keywords)]
                        print(f"      ✓ {model} 返回 {len(result)} 个类别 (修复后)")
                        return result
                except:
                    pass
                print(f"      ⚠ {model} JSON解析失败: {e}，尝试下一个模型...")
                continue
            except Exception as e:
                print(f"      ⚠ {model} 调用失败: {e}，尝试下一个模型...")
                continue
        
        print(f"    ⚠ 所有模型均调用失败")
        return []
    
    @staticmethod
    def _simple_merge(batch_results: List[Dict]) -> List[Dict]:
        """简单合并：按关键词相似度合并类别"""
        # 定义合并关键词组
        merge_groups = {
            '绩效': ['绩效', '能力不足', '不达标'],
            '晋升/职业发展': ['晋升', '职业发展', '发展受限', '发展空间'],
            '团队定位': ['团队定位', '定位模糊', '组织定位', '业务方向'],
            '管理/协作': ['管理', '协作', '沟通', 'leader', '团队管理'],
            '工作压力': ['压力', '工作强度', '身心', '时长'],
        }
        
        merged = {}
        for item in batch_results:
            cat = item.get('category', '未分类')
            # 找匹配的组
            matched_group = None
            for group_name, keywords in merge_groups.items():
                if any(kw in cat for kw in keywords):
                    matched_group = group_name
                    break
            
            key = matched_group or cat
            if key not in merged:
                merged[key] = {
                    'category': key,
                    'summary': item.get('summary', ''),
                    'quotes': [],
                    'people': [],
                }
            merged[key]['quotes'].extend(item.get('quotes', []))
            merged[key]['people'].extend(item.get('people', []))
        
        results = []
        for cat_data in merged.values():
            cat_data['quotes'] = list(dict.fromkeys(cat_data['quotes']))[:3]
            cat_data['people'] = list(dict.fromkeys(cat_data['people']))
            results.append(cat_data)
        return results
    
    def analyze_reasons(self, dept_df) -> List[Dict]:
        """分析离职原因：拼接文本 -> 调用大模型 -> 结构化输出
        
        Returns:
            List of {category, summary, quotes, people_names, people_stats}
        """
        pd = self.pd
        
        # 构建每个人的文本和信息
        persons_with_text = []
        person_info_map = {}  # name -> person_info
        
        text_fields = ['离职分析-面谈', '离职分析-离职原因挖掘', '离职分析-员工对组织的建议/反馈', '离职分析-HR建议']
        
        for _, row in dept_df.iterrows():
            info = self._build_person_info(row)
            if not info['name']:
                continue
            
            # 合并4个字段的文本，去掉空值
            texts = []
            for field in text_fields:
                val = row.get(field, '')
                if pd.notna(val) and str(val).strip() not in ('', 'nan', '无', '-'):
                    texts.append(str(val).strip())
            
            combined = '\n'.join(texts) if texts else ''
            
            # 提取英文名（括号前的部分）
            eng_match = re.match(r'^([a-zA-Z]+)', info['name'])
            eng_name = eng_match.group(1) if eng_match else info['name']
            
            person_info_map[eng_name.lower()] = info
            
            if combined:
                persons_with_text.append({
                    'name': info['name'],
                    'eng_name': eng_name,
                    'resign_type': info['resign_type'],
                    'text': combined,
                })
        
        if not persons_with_text:
            return []
        
        # 根据人数决定一次性还是分批
        all_llm_results = []
        if len(persons_with_text) <= self.BATCH_THRESHOLD:
            prompt = self._build_llm_prompt(persons_with_text)
            print(f"  📡 调用大模型分析 {len(persons_with_text)} 人的离职原因...")
            all_llm_results = self._call_llm(prompt)
        else:
            # 分批，每批 BATCH_THRESHOLD 人
            batches = []
            for i in range(0, len(persons_with_text), self.BATCH_THRESHOLD):
                batches.append(persons_with_text[i:i+self.BATCH_THRESHOLD])
            
            print(f"  📡 分 {len(batches)} 批调用大模型分析 {len(persons_with_text)} 人的离职原因...")
            batch_results = []
            for bi, batch in enumerate(batches):
                print(f"    批次 {bi+1}/{len(batches)}...")
                prompt = self._build_llm_prompt(batch)
                result = self._call_llm(prompt)
                batch_results.extend(result)
            
            # 用大模型合并相似类别
            if batch_results:
                print(f"    📡 合并 {len(batch_results)} 个类别...")
                merge_prompt = """以下是分批分析的离职原因类别，请合并相似类别（如"绩效问题"和"绩效与晋升"应合并）。
合并时：合并quotes(每类最多3条)，合并people列表。

严格按JSON返回：
```json
[{"category":"归类名","summary":"总结","quotes":["原文"],"people":["英文名"]}]
```

待合并类别：
"""
                for item in batch_results:
                    merge_prompt += f"\n类别:{item.get('category')} | 人员:{item.get('people',[])} | 原文:{item.get('quotes',[])[0][:80] if item.get('quotes') else ''}"
                
                merged_results = self._call_llm(merge_prompt)
                if merged_results:
                    # 对LLM返回结果再做一次按category去重
                    deduped = {}
                    for item in merged_results:
                        cat = item.get('category', '未分类')
                        if cat not in deduped:
                            deduped[cat] = item
                        else:
                            # 合并quotes和people
                            deduped[cat]['quotes'] = list(dict.fromkeys(
                                deduped[cat].get('quotes', []) + item.get('quotes', [])
                            ))[:3]
                            deduped[cat]['people'] = list(dict.fromkeys(
                                deduped[cat].get('people', []) + item.get('people', [])
                            ))
                    all_llm_results = list(deduped.values())
                else:
                    # 如果合并调用失败，用简单的字符串相似度合并
                    all_llm_results = self._simple_merge(batch_results)
        
        if not all_llm_results:
            return []
        
        # 为每个分类计算人员统计
        final_results = []
        for item in all_llm_results:
            people_names = item.get('people', [])
            
            # 匹配回person_info_map
            matched_infos = []
            matched_keys = set()
            for name in people_names:
                name_lower = name.lower().strip()
                # 精确匹配
                if name_lower in person_info_map and name_lower not in matched_keys:
                    matched_infos.append(person_info_map[name_lower])
                    matched_keys.add(name_lower)
                    continue
                # 模糊匹配：前缀或子串
                for key, info in person_info_map.items():
                    if key not in matched_keys:
                        if (key.startswith(name_lower) or name_lower.startswith(key) 
                            or name_lower in key or key in name_lower):
                            matched_infos.append(info)
                            matched_keys.add(key)
                            break
            
            # 跳过没有匹配到任何人的类别
            if not matched_infos:
                continue
            
            # 统计
            n_people = len(matched_infos)
            n_active = sum(1 for p in matched_infos if p['resign_type'] == '主动离职')
            n_passive = sum(1 for p in matched_infos if p['resign_type'] == '被动离职')
            
            # 中心分布
            center_counts = {}
            for p in matched_infos:
                c = p['center'] if p['center'] else '未知'
                center_counts[c] = center_counts.get(c, 0) + 1
            
            # 职级分布
            grade_counts = {}
            for p in matched_infos:
                g = p['grade'] if p['grade'] else '未知'
                grade_counts[g] = grade_counts.get(g, 0) + 1
            
            n_high_perf = sum(1 for p in matched_infos if p['is_high_perf'])
            n_echelon = sum(1 for p in matched_infos if p['is_echelon'])
            
            # 构建人员分布文字
            stats_parts = [f"{n_people}人提及"]
            if n_active > 0:
                stats_parts.append(f"主动离职{n_active}人")
            if n_passive > 0:
                stats_parts.append(f"被动离职{n_passive}人")
            
            # 中心分布（最多显示3个）
            sorted_centers = sorted(center_counts.items(), key=lambda x: -x[1])
            for center_name, cnt in sorted_centers[:3]:
                stats_parts.append(f"{center_name} {cnt}人")
            
            # 职级分布（最多显示3个）
            sorted_grades = sorted(grade_counts.items(), key=lambda x: -x[1])
            for grade_name, cnt in sorted_grades[:3]:
                stats_parts.append(f"{grade_name} {cnt}人")
            
            if n_high_perf > 0:
                stats_parts.append(f"离职前高绩效{n_high_perf}人")
            if n_echelon > 0:
                stats_parts.append(f"梯队{n_echelon}人")
            
            final_results.append({
                'category': item.get('category', '未分类'),
                'summary': item.get('summary', ''),
                'quotes': item.get('quotes', [])[:3],
                'people_stats': '\n'.join(stats_parts),
                'n_people': n_people,
            })
        
        # 按提及人数降序排列
        final_results.sort(key=lambda x: -x['n_people'])
        
        return final_results


def build_resignation_data(org_full_path: str, resignation_analyzer) -> Optional[Dict]:
    """构建离职原因分析数据
    
    Args:
        org_full_path: 组织全路径
        resignation_analyzer: ResignationAnalyzer 实例，可以为 None
    
    Returns:
        包含饼图数据和离职原因分析的字典，或 None
    """
    if resignation_analyzer is None:
        return None
    
    dept_df = resignation_analyzer.get_dept_resignations(org_full_path)
    if len(dept_df) == 0:
        return None
    
    # 饼图数据
    pie_data = resignation_analyzer.build_pie_data(dept_df)
    
    # 大模型分析
    print(f"  🔍 正在分析 {len(dept_df)} 名离职人员的离职原因...")
    reason_analysis = resignation_analyzer.analyze_reasons(dept_df)
    
    return {
        'available': True,
        'pie_data': pie_data,
        'reason_analysis': reason_analysis,
        'has_org_reasons': len(reason_analysis) > 0,
    }


# ─── BP观察数据加载器 ──────────────────────────────────
class BPObservationLoader:
    """加载BP观察数据，提取每个人的BP点赞/BP提醒关注标签"""
    
    def __init__(self, bp_dir='【全面反馈】'):
        self.bp_dir = bp_dir
        self._cache = {}  # bg_short -> {eng_name_lower: {'tags': [...], 'content': str}}
        self._scanned_files = {}  # bg_short -> filepath
        self._scan_files()
    
    def _scan_files(self):
        """扫描BP观察文件，兼容 {BG}部门负责人-1BP观察.xlsx 和 模板部门负责人-1BP观察.xlsx"""
        import glob
        pattern = os.path.join(self.bp_dir, '*BP观察*.xlsx')
        self._template_file = None
        for f in glob.glob(pattern):
            basename = os.path.basename(f)
            if basename.startswith('~$'):
                continue
            # 检查是否为通用模板
            if basename.startswith('模板'):
                self._template_file = f
                continue
            # 提取BG缩写
            m = re.match(r'([A-Za-z]+)', basename)
            if m:
                self._scanned_files[m.group(1).upper()] = f
    
    def _load_bg(self, bg_short):
        """懒加载某个BG的BP观察数据，优先特定文件，找不到则回退通用模板"""
        if bg_short in self._cache:
            return
        filepath = self._scanned_files.get(bg_short)
        if not filepath and hasattr(self, '_template_file') and self._template_file:
            filepath = self._template_file
        if not filepath:
            self._cache[bg_short] = {}
            return
        
        try:
            reader = ExcelReader(filepath)
            # 在headers中查找列
            col_name = -1
            col_bp = -1
            for i, h in enumerate(reader.headers):
                if h.strip() == '中英文名':
                    col_name = i
                elif h.strip() == 'BP观察':
                    col_bp = i
            if col_name == -1 or col_bp == -1:
                self._cache[bg_short] = {}
                return
            
            bp_map = {}
            for row in reader.rows[1:]:  # 跳过表头
                name = row[col_name].strip() if col_name < len(row) else ''
                bp_text = row[col_bp] if col_bp < len(row) else ''
                if not name:
                    continue
                # 提取英文名
                m = re.match(r'(\w+)\(', name)
                en = m.group(1).lower() if m else name.lower().strip()
                
                tags = []
                if 'BP点赞' in bp_text or 'BP 点赞' in bp_text:
                    tags.append('BP点赞')
                if 'BP提醒关注' in bp_text or 'BP 提醒关注' in bp_text:
                    tags.append('BP提醒关注')
                
                bp_map[en] = {'tags': tags, 'content': bp_text}
            
            self._cache[bg_short] = bp_map
        except Exception as e:
            print(f"  ⚠ BP观察数据加载失败({bg_short}): {e}")
            self._cache[bg_short] = {}
    
    def get_tags(self, org_full_path, person_name):
        """获取某人的BP标签列表，如 ['BP点赞'] 或 ['BP提醒关注'] 或 []"""
        bg_short = self._extract_bg_short(org_full_path)
        if not bg_short:
            return []
        self._load_bg(bg_short)
        m = re.match(r'(\w+)\(', person_name)
        en = m.group(1).lower() if m else person_name.lower().strip()
        info = self._cache.get(bg_short, {}).get(en)
        return info['tags'] if info else []
    
    def _extract_bg_short(self, org_full_path):
        """从组织全路径提取BG缩写"""
        bg_part = org_full_path.split('/')[0] if '/' in org_full_path else org_full_path
        m = re.match(r'([A-Za-z]+\d*)', bg_part)
        return m.group(1).upper() if m else ''
    
    def has_data_for_bg(self, org_full_path):
        """检查该BG是否有BP观察文件"""
        bg_short = self._extract_bg_short(org_full_path)
        return bg_short in self._scanned_files


class JianGangLoader:
    """加载岗位信息表，判断某人在某组织下是否为横向兼岗"""
    
    def __init__(self, filepath='【组织架构信息】/岗位信息表.xlsx'):
        self._records = []  # [(org_full_path, eng_name_lower), ...]
        self._load(filepath)
    
    def _load(self, filepath):
        """加载岗位信息表，构建横向兼岗索引"""
        if not os.path.exists(filepath):
            print(f"  ⚠ 岗位信息表不存在: {filepath}")
            return
        
        try:
            reader = ExcelReader(filepath)
            col_org = -1
            col_name = -1
            col_jg_type = -1  # 兼岗类型列
            for i, h in enumerate(reader.headers):
                hs = h.strip()
                if hs == '组织全路径':
                    col_org = i
                elif hs == '员工中英文名':
                    col_name = i
                elif hs == '兼岗类型':
                    col_jg_type = i
            
            if any(c == -1 for c in [col_org, col_name, col_jg_type]):
                print(f"  ⚠ 岗位信息表列名不匹配（需要：组织全路径、员工中英文名、兼岗类型）")
                return
            
            for row in reader.rows[1:]:
                org = row[col_org].strip() if col_org < len(row) else ''
                name = row[col_name].strip() if col_name < len(row) else ''
                jg_type = row[col_jg_type].strip() if col_jg_type < len(row) else ''
                
                # 横向兼岗_跨部门 或 横向兼岗_跨BG
                if '横向兼岗' in jg_type and name:
                    m = re.match(r'(\w+)\(', name)
                    en = m.group(1).lower() if m else name.lower().strip()
                    self._records.append((org, en))
            
            print(f"✓ 岗位信息已加载 (横向兼岗记录 {len(self._records)} 条)")
        except Exception as e:
            print(f"  ⚠ 岗位信息加载失败: {e}")
    
    def is_jiangan(self, org_full_path, person_name):
        """判断某人在某部门下是否为横向兼岗。
        
        匹配逻辑：岗位表的组织全路径以报告部门路径开头即算。
        例如岗位表记录 BG/业务线/产品部/产品中心 包含 BG/业务线/产品部。
        """
        m = re.match(r'(\w+)\(', person_name)
        en = m.group(1).lower() if m else person_name.lower().strip()
        for (org, name) in self._records:
            if name == en and org.startswith(org_full_path):
                return True
        return False


class OpenFeedbackLoader:
    """加载全面反馈开放题总结数据（干部侧 + 组织侧）"""
    
    def __init__(self, feedback_dir='【全面反馈】'):
        self._cadre_data = {}   # org_full_path -> F列内容
        self._org_data = {}     # org_full_path -> F列内容
        self._feedback_dir = feedback_dir
        self._scanned_bgs = set()
    
    def _load_bg(self, bg_short):
        """按需加载某个BG的干部侧和组织侧文件，优先特定文件，找不到则回退通用模板"""
        if bg_short in self._scanned_bgs:
            return
        self._scanned_bgs.add(bg_short)
        
        import glob
        # 搜索干部侧文件
        cadre_pattern = os.path.join(self._feedback_dir, f'{bg_short}-干部侧问题总结.xlsx')
        cadre_files = glob.glob(cadre_pattern)
        if not cadre_files:
            cadre_files = glob.glob(os.path.join(self._feedback_dir, '模板-干部侧问题总结.xlsx'))
        for fp in cadre_files:
            if os.path.basename(fp).startswith('~$'):
                continue
            self._load_file(fp, self._cadre_data)
        
        # 搜索组织侧文件
        org_pattern = os.path.join(self._feedback_dir, f'{bg_short}-组织侧问题.xlsx')
        org_files = glob.glob(org_pattern)
        if not org_files:
            org_files = glob.glob(os.path.join(self._feedback_dir, '模板-组织侧问题.xlsx'))
        for fp in org_files:
            if os.path.basename(fp).startswith('~$'):
                continue
            self._load_file(fp, self._org_data)
    
    def _load_file(self, filepath, target_dict):
        """读取Excel文件，提取A列(org_full_path)和F列(分析结果)"""
        try:
            reader = ExcelReader(filepath)
            # 找列索引
            col_org = -1
            col_result = -1
            for i, h in enumerate(reader.headers):
                hs = h.strip()
                if hs == '组织全路径' and col_org == -1:
                    col_org = i  # 只取第一个匹配（新Excel有两列同名"组织全路径"）
                elif hs == '分析结果' and col_result == -1:
                    col_result = i
            
            if col_org == -1 or col_result == -1:
                print(f"  ⚠ 开放题文件列名不匹配: {os.path.basename(filepath)}")
                return
            
            count = 0
            for row in reader.rows[1:]:
                org = row[col_org].strip() if col_org < len(row) else ''
                result = row[col_result].strip() if col_result < len(row) else ''
                if org and result:
                    target_dict[org] = result
                    count += 1
            
            print(f"  ✓ 开放题已加载: {os.path.basename(filepath)} ({count} 个部门)")
        except Exception as e:
            print(f"  ⚠ 开放题加载失败 {os.path.basename(filepath)}: {e}")
    
    def get_data(self, org_full_path):
        """获取指定部门的干部侧和组织侧开放题数据"""
        # 根据org_full_path提取BG缩写
        bg_short = ''
        m = re.match(r'([A-Za-z0-9]+)', org_full_path)
        if m:
            bg_short = m.group(1).upper()
            if bg_short == 'OVERSEAS':
                bg_short = 'OFS'
        if not bg_short:
            parts = org_full_path.split('/')
            if parts:
                m2 = re.match(r'([A-Za-z0-9]+)', parts[0])
                if m2:
                    bg_short = m2.group(1).upper()
                    if bg_short == 'OVERSEAS':
                        bg_short = 'OFS'
        
        if bg_short:
            self._load_bg(bg_short)
        
        cadre = self._cadre_data.get(org_full_path, '')
        org = self._org_data.get(org_full_path, '')
        
        # fallback: 末段部门名匹配（处理OFS等路径格式不一致的情况）
        if not cadre and not org and '/' in org_full_path:
            dept_name = org_full_path.split('/')[-1]
            for key in self._cadre_data:
                if key.endswith('/' + dept_name):
                    cadre = self._cadre_data[key]
                    break
            for key in self._org_data:
                if key.endswith('/' + dept_name):
                    org = self._org_data[key]
                    break
        
        return cadre, org
    
    # ── 颜色常量 ──
    CLR_POS = '#389e0d'   # 正向 - 绿
    CLR_NEU = '#d48806'   # 中性/中立/待关注 - 黄橙
    CLR_NEG = '#cf1322'   # 负向 - 红
    CLR_POS_BG = '#f6ffed'
    CLR_NEU_BG = '#fffbe6'
    CLR_NEG_BG = '#fff1f0'
    CLR_POS_BD = '#b7eb8f'
    CLR_NEU_BD = '#ffe58f'
    CLR_NEG_BD = '#ffa39e'

    @classmethod
    def _sentiment_color(cls, text):
        """根据文本中的情感关键词返回 (文字色, 背景色, 边框色)"""
        t = text if text else ''
        if re.search(r'正向|共性优势|局部优势|优秀|良好|突出|表现优', t):
            return cls.CLR_POS, cls.CLR_POS_BG, cls.CLR_POS_BD
        elif re.search(r'共性不足|共性短板', t):
            return cls.CLR_NEG, cls.CLR_NEG_BG, cls.CLR_NEG_BD
        elif re.search(r'待关注|负向|不足|风险|短板|挑战', t):
            return cls.CLR_NEU, cls.CLR_NEU_BG, cls.CLR_NEU_BD
        elif re.search(r'中立|中性|局部问题|局部不足', t):
            return cls.CLR_NEU, cls.CLR_NEU_BG, cls.CLR_NEU_BD
        return '#444', '#f8f9fa', '#e8e8e8'

    @classmethod
    def _judgment_theme(cls, text):
        """根据判断标签返回 (kind, 文字色, 背景色, 边框色)。优先按共性优势/不足/局部问题精确判断。"""
        t = text if text else ''
        if re.search(r'共性优势|局部优势', t):
            return 'positive', cls.CLR_POS, cls.CLR_POS_BG, cls.CLR_POS_BD
        if re.search(r'共性不足|共性短板', t):
            return 'negative', cls.CLR_NEG, cls.CLR_NEG_BG, cls.CLR_NEG_BD
        if re.search(r'局部待关注|局部问题|局部不足|中立|中性', t):
            return 'neutral', cls.CLR_NEU, cls.CLR_NEU_BG, cls.CLR_NEU_BD
        if re.search(r'待确认|缺失|数据缺失|信息不足', t):
            return 'default', '#8c8c8c', '#fafafa', '#d9d9d9'
        fc, bg, bd = cls._sentiment_color(t)
        return 'default', fc, bg, bd

    @classmethod
    def _fmt(cls, text, keyword_highlight=True):
        """通用文本格式化：转义HTML + 加粗 + 引号美化 + 可选情感色标"""
        import html as html_mod
        if not text:
            return ''
        escaped = html_mod.escape(text)
        escaped = re.sub(r'\*\*([^*]+)\*\*', r'<b>\1</b>', escaped)
        escaped = re.sub(r'["""]([^"""]+)["""]', r'<span style="font-style:italic;color:#666;">"\1"</span>', escaped)
        if keyword_highlight:
            # 正向/负向/中立/待关注 着色
            escaped = re.sub(r'(正向)', f'<span style="color:{cls.CLR_POS};font-weight:600;">\\1</span>', escaped)
            escaped = re.sub(r'(待关注)', f'<span style="color:{cls.CLR_NEU};font-weight:600;">\\1</span>', escaped)
            escaped = re.sub(r'(负向)', f'<span style="color:{cls.CLR_NEG};font-weight:600;">\\1</span>', escaped)
            escaped = re.sub(r'(中立|中性)', f'<span style="color:{cls.CLR_NEU};font-weight:600;">\\1</span>', escaped)
            # 共性优势/不足/局部问题 着色
            escaped = re.sub(r'(共性优势|局部优势)', f'<span style="color:{cls.CLR_POS};font-weight:600;">\\1</span>', escaped)
            escaped = re.sub(r'(共性不足|共性短板)', f'<span style="color:{cls.CLR_NEG};font-weight:600;">\\1</span>', escaped)
            escaped = re.sub(r'(局部问题|局部不足|局部待关注)', f'<span style="color:{cls.CLR_NEU};font-weight:600;">\\1</span>', escaped)
        # 还原被 escape 的 <br> / <br/> 标签
        escaped = escaped.replace('&lt;br&gt;', '<br>').replace('&lt;br/&gt;', '<br>')
        return escaped

    @classmethod
    def _parse_md_table(cls, lines, start_idx):
        """从 lines[start_idx] 开始解析 markdown 表格，返回 (headers, rows, end_idx)"""
        headers = []
        rows = []
        i = start_idx
        if i < len(lines) and '|' in lines[i]:
            cells = [c.strip() for c in lines[i].strip().strip('|').split('|')]
            headers = cells
            i += 1
        # skip separator
        if i < len(lines) and re.match(r'^\|?[\s\-:|]+\|', lines[i]):
            i += 1
        while i < len(lines):
            s = lines[i].strip()
            if not s or '|' not in s:
                break
            cells = [c.strip() for c in s.strip('|').split('|')]
            rows.append(cells)
            i += 1
        return headers, rows, i

    # ── 档位 → 纯文字色映射（用于"整体表现"列，无底色）──
    GRADE_TEXT_COLOR = {
        '优秀': '#237804',    # 深绿 — 最高档
        '突出': '#237804',
        '表现优': '#237804',
        '良好': '#52c41a',    # 亮绿 — 次高档，与优秀有区分
        '一般': '#d48806',    # 黄色
        '中等': '#d48806',
        '待改进': '#d48806',
        '较差': '#cf1322',    # 红色
        '不足': '#cf1322',
        '风险': '#cf1322',
        '差': '#cf1322',
    }

    @classmethod
    def _grade_text_color(cls, grade_text):
        """根据档位文本返回纯文字颜色（无背景），用于"整体表现"列"""
        t = (grade_text or '').strip()
        for keyword, color in cls.GRADE_TEXT_COLOR.items():
            if keyword in t:
                return color
        return '#444'

    @classmethod
    def _column_role(cls, header_text):
        """根据表头文字判断列的角色: 'grade'=整体表现, 'positive'=主要优势, 'negative'=主要不足, 'neutral'=普通"""
        h = (header_text or '').strip()
        if re.search(r'整体表现|整体判断|总体表现|档位|评级', h):
            return 'grade'
        if re.search(r'主要优势|优势|亮点|优点', h):
            return 'positive'
        if re.search(r'主要不足|不足|主要问题|短板|风险', h):
            return 'negative'
        # "问题"列但排除"问题类别""问题类型""问题维度"等分类表头
        if '问题' in h and not re.search(r'问题[类型别维]', h):
            return 'negative'
        return 'neutral'

    @classmethod
    def _render_table_html(cls, headers, rows, highlight_sentiment=True, force_left_align=False, force_header_style=None):
        """渲染HTML表格，根据列语义着色。对"最终整体判断（跨维度）"表格做整行判断色与特殊对齐。
        force_header_style: 若传入 'negative'/'positive'，则所有列表头强制使用对应颜色风格。"""
        # 预判每列角色
        col_roles = [cls._column_role(h) for h in headers]
        header_texts = [re.sub(r'\*+', '', (h or '')).strip() for h in headers]
        is_cross_dimension_summary = (
            any('维度' in h for h in header_texts) and
            any('判断' in h for h in header_texts) and
            any('说明' in h or '总结' in h for h in header_texts)
        )
        judgment_col_idx = next((i for i, h in enumerate(header_texts) if '判断' in h), -1)
        desc_col_idx = next((i for i, h in enumerate(header_texts) if '说明' in h or '总结' in h), -1)

        h = '<table style="width:100%;border-collapse:collapse;font-size:13px;line-height:1.6;margin:12px 0;">'
        # thead
        h += '<thead><tr>'
        for ci, hdr in enumerate(headers):
            role = col_roles[ci] if ci < len(col_roles) else 'neutral'
            if is_cross_dimension_summary:
                align = 'left'
                th_style = f'padding:8px 12px;border:1px solid #d9d9d9;font-weight:700;text-align:{align};background:#f5f5f5;color:#333;'
            else:
                th_style = 'padding:8px 12px;border:1px solid #d9d9d9;font-weight:700;text-align:left;'
                effective_role = force_header_style if force_header_style else role
                if effective_role == 'positive':
                    th_style += 'background:#237804;color:#fff;'
                elif effective_role == 'negative':
                    th_style += 'background:#cf1322;color:#fff;'
                else:
                    th_style += 'background:#f0f5ff;color:#333;'
            h += f'<th style="{th_style}">{cls._fmt(hdr, keyword_highlight=False)}</th>'
        h += '</tr></thead><tbody>'
        # tbody
        for row in rows:
            row_theme = None
            if is_cross_dimension_summary and 0 <= judgment_col_idx < len(row):
                row_theme = cls._judgment_theme(row[judgment_col_idx])
            h += '<tr>'
            for ci, cell in enumerate(row):
                role = col_roles[ci] if ci < len(col_roles) else 'neutral'
                if is_cross_dimension_summary:
                    _, row_fc, _row_bg, row_bd = row_theme if row_theme else ('default', '#333', '#fff', '#e8e8e8')
                    align = 'left'
                    td_style = f'padding:8px 12px;border:1px solid #e8e8e8;vertical-align:top;text-align:{align};'
                    if ci == judgment_col_idx:
                        td_style += f'color:{row_fc};font-weight:700;'
                    else:
                        td_style += 'color:#333;'
                else:
                    td_style = 'padding:8px 12px;border:1px solid #e8e8e8;vertical-align:top;'
                    if force_left_align:
                        td_style += 'text-align:left;'
                    if highlight_sentiment:
                        if role == 'grade':
                            fc = cls._grade_text_color(cell)
                            align = 'left' if force_left_align else 'center'
                            td_style += f'color:{fc};font-weight:600;text-align:{align};'
                        elif role == 'positive':
                            td_style += f'background:{cls.CLR_POS_BG};color:{cls.CLR_POS};'
                        elif role == 'negative':
                            td_style += f'background:{cls.CLR_NEG_BG};color:{cls.CLR_NEG};'
                h += f'<td style="{td_style}">{cls._fmt(cell, keyword_highlight=False)}</td>'
            h += '</tr>'
        h += '</tbody></table>'
        return h

    @classmethod
    def _render_body_lines(cls, lines, max_quotes=2, org_number_heading_style=None, section_has_problem=None):
        """渲染正文行（维度分析等），返回HTML片段列表"""
        import html as html_mod
        parts = []
        quote_count = 0
        
        for line in lines:
            stripped = line.strip()
            if not stripped:
                parts.append('<div style="margin-top:6px;"></div>')
                continue
            
            # 清理行首/行中的 <br> / <br/> 文本（每行已是独立div，不需要br换行）
            stripped = re.sub(r'<br\s*/?>',  '', stripped).strip()
            if not stripped:
                parts.append('<div style="margin-top:6px;"></div>')
                continue
            
            # --- 分隔线 ---
            if stripped == '---':
                parts.append('<hr style="border:none;border-top:1px solid #eee;margin:16px 0;">')
                continue
            
            # ## 二级标题
            if stripped.startswith('## '):
                title = stripped.lstrip('#').strip()
                parts.append(f'<div style="margin:20px 0 10px;font-size:15px;font-weight:700;color:#1a1a1a;border-left:4px solid #1890ff;padding-left:10px;">{cls._fmt(title)}</div>')
                continue
            
            # 组织侧编号小标题：按组织模块章节分别使用蓝色/红色样式
            if org_number_heading_style in {'blue', 'red'} and re.match(r'^###\s*\d+[\.、]\s*', stripped):
                title = stripped.lstrip('#').strip()
                if org_number_heading_style == 'blue':
                    fc = '#1677ff'
                    bg = '#ffffff'
                    bd = '#91caff'
                else:
                    fc = cls.CLR_NEG
                    bg = cls.CLR_NEG_BG
                    bd = cls.CLR_NEG_BD
                parts.append(f'<div style="margin:18px 0 8px;padding:8px 14px;font-size:14px;font-weight:700;color:{fc};background:{bg};border-left:4px solid {bd};border-radius:0 4px 4px 0;">{cls._fmt(title)}</div>')
                quote_count = 0
                continue
            
            # ### 三级标题 / 维度标题
            if stripped.startswith('### ') or re.match(r'^维度[一二三四五六七八九十]+[：:]', stripped):
                title = stripped.lstrip('#').strip()
                fc, bg, bd = cls._sentiment_color(title)
                parts.append(f'<div style="margin:18px 0 8px;padding:8px 14px;font-size:14px;font-weight:700;color:#222;background:{bg};border-left:4px solid {bd};border-radius:0 4px 4px 0;">{cls._fmt(title)}</div>')
                quote_count = 0
                continue
            
            # #### 四级标题
            if stripped.startswith('#### '):
                title = stripped.lstrip('#').strip()
                parts.append(f'<div style="margin:12px 0 6px;font-size:13px;font-weight:700;color:#444;">{cls._fmt(title)}</div>')
                quote_count = 0
                continue
            
            # 总结行 — 不显示标题文字（如"1）总结：共性优势"），颜色由维度容器控制
            if re.match(r'^(\d+[）)]|[一二三四五六七八九十]+[）)])\s*\**总结[：:]', stripped):
                # 跳过此行，不渲染
                continue
            
            # 整体判断行 / 圆括号整体判断
            if stripped.startswith('（整体判断') or (stripped.startswith('(整体判断')):
                fc, bg, bd = cls._sentiment_color(stripped)
                parts.append(f'<div style="margin:8px 0;padding:8px 12px;background:{bg};border:1px solid {bd};border-radius:6px;font-size:13px;line-height:1.7;color:{fc};">{cls._fmt(stripped)}</div>')
                continue
            
            # ✅ / ⚠️ 判断行
            if stripped.startswith('✅') or stripped.startswith('⚠️') or stripped.startswith('⚠'):
                fc, bg, bd = cls._sentiment_color(stripped)
                icon = '✅' if stripped.startswith('✅') else '⚠️'
                rest = stripped.lstrip('✅⚠️⚠ ').strip()
                parts.append(f'<div style="margin:8px 0;padding:8px 12px;background:{bg};border:1px solid {bd};border-radius:6px;font-size:13px;line-height:1.7;">{icon} {cls._fmt(rest)}</div>')
                continue
            
            # 单行表格小结（| xxx |）— 组织侧维度概要行
            if stripped.startswith('|') and stripped.endswith('|') and stripped.count('|') == 2:
                inner = stripped.strip('|').strip()
                if inner:
                    # 直接以概览表格的"是否存在问题"字段(section_has_problem)为准
                    # 不做文本关键词猜测，数据源头说了算
                    if section_has_problem is True:
                        # 概览表标记"是" → 红色
                        parts.append(f'<div style="margin:4px 0 4px 0;padding:8px 14px;background:#fff1f0;border-left:3px solid #ff4d4f;color:#333;font-size:13px;line-height:1.6;border-radius:0 4px 4px 0;">{cls._fmt(inner)}</div>')
                    elif section_has_problem is False:
                        # 概览表标记"否" → 绿色
                        parts.append(f'<div style="margin:4px 0 4px 0;padding:8px 14px;background:#f6ffed;border-left:3px solid #52c41a;color:#333;font-size:13px;line-height:1.6;border-radius:0 4px 4px 0;">{cls._fmt(inner)}</div>')
                    else:
                        # 没有概览表数据（section_has_problem=None）→ 蓝色中性
                        parts.append(f'<div style="margin:4px 0 4px 0;padding:8px 14px;background:#f0f5ff;border-left:3px solid #1890ff;color:#333;font-size:13px;line-height:1.6;border-radius:0 4px 4px 0;">{cls._fmt(inner)}</div>')
                    continue
            
            # > 引用块 — 限制最多max_quotes句
            if stripped.startswith('>'):
                quote_text = stripped.lstrip('> ').strip()
                if not quote_text:
                    continue
                if quote_text.startswith('聚焦管理者'):
                    continue
                quote_count += 1
                if quote_count > max_quotes:
                    continue
                # "存在问题"行用淡红色底色突出
                if '存在问题' in quote_text and quote_count <= 1:
                    parts.append(f'<div style="margin:4px 0 4px 0;padding:8px 14px;background:#fff1f0;border-left:3px solid #ff4d4f;color:#333;font-size:13px;line-height:1.6;border-radius:0 4px 4px 0;">{cls._fmt(quote_text)}</div>')
                else:
                    parts.append(f'<div style="margin:4px 0 4px 20px;padding:6px 12px;border-left:3px solid #d9d9d9;background:#fafafa;color:#666;font-size:12px;line-height:1.6;border-radius:0 4px 4px 0;">{cls._fmt(quote_text)}</div>')
                continue
            
            # - / * 列表项
            if stripped.startswith('- ') or stripped.startswith('* '):
                content = stripped[2:].strip()
                # 人名行（正向/待关注/负向/中立）
                if re.search(r'(正向|待关注|负向|中立)[，,]', content):
                    fc, bg, bd = cls._sentiment_color(content)
                    parts.append(f'<div style="margin:6px 0 3px 12px;padding:4px 10px;font-size:13px;line-height:1.7;color:{fc};background:{bg};border-radius:4px;border:1px solid {bd};">● {cls._fmt(content)}</div>')
                    quote_count = 0
                else:
                    parts.append(f'<div style="margin:3px 0 3px 16px;font-size:13px;line-height:1.6;color:#444;">• {cls._fmt(content)}</div>')
                    if re.search(r'(正向|负向|中立)', content):
                        quote_count = 0
                continue
            
            # 缩进列表项
            if re.match(r'^\s+[-•]\s', line):
                content = re.sub(r'^\s+[-•]\s*', '', line).strip()
                if re.search(r'(正向|待关注|负向|中立)[，,]', content):
                    fc, bg, bd = cls._sentiment_color(content)
                    parts.append(f'<div style="margin:4px 0 3px 28px;padding:3px 8px;font-size:13px;line-height:1.7;color:{fc};background:{bg};border-radius:4px;border:1px solid {bd};">◦ {cls._fmt(content)}</div>')
                    quote_count = 0
                else:
                    parts.append(f'<div style="margin:3px 0 3px 32px;font-size:13px;line-height:1.6;color:#555;">◦ {cls._fmt(content)}</div>')
                continue
            
            # 管理者名字行（name(中文)：...）
            if re.match(r'^(\w+\([^)]+\)|[\w]+)[：:]', stripped):
                fc, bg, bd = cls._sentiment_color(stripped)
                if bg != '#f8f9fa':
                    parts.append(f'<div style="margin:6px 0 3px 12px;padding:4px 10px;font-size:13px;line-height:1.7;color:{fc};background:{bg};border-radius:4px;border:1px solid {bd};">{cls._fmt(stripped)}</div>')
                else:
                    parts.append(f'<div style="margin:6px 0 3px 8px;font-size:13px;line-height:1.7;color:#444;">{cls._fmt(stripped)}</div>')
                quote_count = 0
                continue
            
            # 普通文本
            parts.append(f'<div style="margin:3px 0;font-size:13px;line-height:1.7;color:#444;">{cls._fmt(stripped)}</div>')
        
        return parts

    # ── 情感分类标签归一化 ──
    # 新数据使用"正向/待关注"两类；旧数据可能含"负向/中立/中性"
    # _normalize_sentiment 将所有非正向标签统一归入"待关注"
    _SENTIMENT_LABELS = ('正向', '待关注')
    _SENTIMENT_POSITIVE = frozenset(['正向'])
    _SENTIMENT_ATTENTION = frozenset(['待关注', '负向', '中立', '中性'])

    @classmethod
    def _normalize_sentiment(cls, label):
        """将旧三分类标签映射到新二分类：正向 / 待关注"""
        label = (label or '').strip()
        if label in cls._SENTIMENT_POSITIVE:
            return '正向'
        if label in cls._SENTIMENT_ATTENTION:
            return '待关注'
        return label  # 未知标签原样返回

    @classmethod
    def _parse_manager_section(cls, lines):
        """
        解析"各管理者情况"区域的行，返回管理者列表。
        每个管理者: {'name': str, '正向': str, '待关注': str}
        
        支持三种格式：
        格式A (DISC单行): name(中文名)：正向，N人，"引用"；"引用"
        格式B (电商产品部多行):
          - **name (中文名)**：
            - 正向，N人：
              > "引用"
            - 待关注，N人：
              > "引用"
        格式C (读书产品部变体):
          - **name (中文名)**：
            正向，N人
            > "引用"
        
        兼容旧标签（负向/中立/中性）：统一映射到"待关注"。
        """
        import html as html_mod
        managers = []
        
        # 情感标签匹配正则（兼容新旧标签）
        SENT_RE = r'(?:正向|待关注|负向|中立|中性)'
        
        # 先检测是格式A还是格式B/C
        # 格式A的特征：一行内包含 name(中文)：正向/待关注/负向/中立 + 引号
        is_single_line_format = False
        for line in lines:
            s = line.strip()
            if re.match(r'^[\w]+\([^)]+\)[：:]', s) and re.search(SENT_RE, s):
                is_single_line_format = True
                break
        
        if is_single_line_format:
            # 格式A：单行格式
            for line in lines:
                s = line.strip()
                if not s:
                    continue
                m = re.match(r'^([\w]+\([^)]+\))[：:]\s*(.*)', s)
                if not m:
                    continue
                name = m.group(1)
                rest = m.group(2)
                
                entry = {'name': name, '正向': '', '待关注': ''}
                
                # 解析 "正向，N人，引用；引用" / "待关注，N人，引用" 等
                # 拆分为 段落：每段以情感标签开头
                segments = re.split(r'\s*(?=(?:' + SENT_RE + r')[，,])', rest)
                for seg in segments:
                    seg = seg.strip()
                    if not seg:
                        continue
                    # 提取开头的标签
                    label_m = re.match(r'^(' + SENT_RE + r')[，,]', seg)
                    if label_m:
                        norm = cls._normalize_sentiment(label_m.group(1))
                        # 对旧标签，替换文本中的标签名
                        if label_m.group(1) != norm:
                            seg = norm + seg[len(label_m.group(1)):]
                        if entry.get(norm):
                            entry[norm] += '；' + seg
                        else:
                            entry[norm] = seg
                    # else: 无前缀的内容忽略
                
                managers.append(entry)
        else:
            # 格式B/C：多行格式
            current_manager = None
            current_sentiment = None  # '正向' / '待关注'
            current_sentiment_count = 0
            current_quotes = []
            
            def flush_sentiment():
                nonlocal current_sentiment, current_sentiment_count, current_quotes
                if current_manager and current_sentiment:
                    norm = cls._normalize_sentiment(current_sentiment)
                    existing = current_manager.get(norm, '')
                    count_prefix = f'{norm}，{current_sentiment_count}人' if current_sentiment_count else norm
                    new_text = '；'.join(current_quotes)
                    if existing:
                        if new_text:
                            current_manager[norm] = existing + '；' + new_text
                    else:
                        current_manager[norm] = count_prefix + (f'，{new_text}' if new_text else '')
                current_sentiment = None
                current_sentiment_count = 0
                current_quotes = []
            
            for line in lines:
                s = line.strip()
                if not s:
                    continue
                
                # 检测管理者名字行：
                # 格式A: - **name (中文名)**： 或 name(中文)：
                # 格式B: - **name**： （纯英文ID，无中文名）
                name_match = re.match(r'^[-*\s]*\**\s*([\w]+\s*[\(（][^)）]+[\)）])\s*\**\s*[：:]', s)
                if not name_match:
                    # 尝试匹配纯英文ID格式: - **username**：
                    name_match = re.match(r'^[-*\s]*\**\s*([\w]+)\s*\**\s*[：:]', s)
                    # 排除误匹配：不匹配纯数字开头（如 "1）总结"）和常见标题词
                    if name_match:
                        candidate = name_match.group(1).strip()
                        if (re.match(r'^\d+$', candidate) or 
                            candidate in ('正向', '待关注', '负向', '中立', '总结', '整体判断', 'BP观察') or
                            re.match(r'^(维度|各管理者|整体|总结)', candidate)):
                            name_match = None
                if name_match:
                    # 先保存上一个
                    flush_sentiment()
                    if current_manager:
                        managers.append(current_manager)
                    
                    name = name_match.group(1).strip()
                    current_manager = {'name': name, '正向': '', '待关注': ''}
                    current_sentiment = None
                    current_sentiment_count = 0
                    current_quotes = []
                    continue
                
                if not current_manager:
                    continue
                
                # 检测情感行：正向，N人 / 待关注，N人 / 负向，N人 / 中立，N人
                sent_match = re.match(r'^[-*\s]*\**\s*(' + SENT_RE + r')[，,]\s*(\d+)\s*人[：:]*\s*(.*)', s)
                if sent_match:
                    flush_sentiment()
                    for sent, count in re.findall(r'(' + SENT_RE + r')[，,]\s*(\d+)\s*人', s):
                        norm = cls._normalize_sentiment(sent)
                        if not current_manager.get(norm):
                            current_manager[norm] = f'{norm}，{count}人'
                    current_sentiment = sent_match.group(1)
                    current_sentiment_count = int(sent_match.group(2))
                    rest = sent_match.group(3).strip()
                    if rest and re.match(r'^[；;]\s*(' + SENT_RE + r')[，,]', rest):
                        pass
                    elif rest:
                        current_quotes.append(rest)
                    continue
                
                # 备选：处理"正向，多人" / "待关注，若干人"等非数字人数格式
                sent_match_alt = re.match(r'^[-*\s]*\**\s*(' + SENT_RE + r')[，,]\s*([多若干数几]+人)\s*[：:]*\s*(.*)', s)
                if sent_match_alt:
                    flush_sentiment()
                    current_sentiment = sent_match_alt.group(1)
                    norm = cls._normalize_sentiment(current_sentiment)
                    alt_count_text = sent_match_alt.group(2)  # "多人" 等
                    if not current_manager.get(norm):
                        current_manager[norm] = f'{norm}，{alt_count_text}'
                    current_sentiment_count = 0  # 非精确数字
                    rest = sent_match_alt.group(3).strip()
                    if rest:
                        current_quotes.append(rest)
                    continue
                
                # 检测"无" 或 "无待关注"等（兼容旧"无负向"）
                if re.match(r'^[-*\s]*无(正向|待关注|负向|中立|该维度)?', s):
                    continue
                
                # BP观察行
                if s.startswith('（BP观察') or s.startswith('(BP观察'):
                    bp_content = re.sub(r'^[（(]BP观察[）)][：:]\s*', '', s).strip()
                    if bp_content and bp_content != '无':
                        if current_sentiment:
                            current_quotes.append(f'【BP观察】{bp_content}')
                    continue
                
                # 引用行 > "xxx" — 可能带内联情感标签
                if s.startswith('>'):
                    quote = s.lstrip('> ').strip()
                    if not quote:
                        continue
                    # 检查是否有内联情感切换标签
                    inline_sent = re.match(r'^(' + SENT_RE + r')[：:]\s*(.*)', quote)
                    if inline_sent:
                        new_sent = inline_sent.group(1)
                        quote_text = inline_sent.group(2).strip()
                        if new_sent != current_sentiment:
                            flush_sentiment()
                            current_sentiment = new_sent
                            current_sentiment_count = 0
                        if quote_text:
                            current_quotes.append(quote_text)
                    elif current_sentiment:
                        current_quotes.append(quote)
                    continue
                
                # 普通文本行（可能是引用的继续）
                if current_sentiment and (s.startswith('"') or s.startswith('"') or s.startswith('「')):
                    current_quotes.append(s)
                    continue
            
            # 保存最后一个
            flush_sentiment()
            if current_manager:
                managers.append(current_manager)
        
        return managers

    @classmethod
    def _render_manager_table_html(cls, managers, sentiment_color=None):
        """将管理者列表渲染为三列表格HTML：干部 | 正向 | 待关注
        sentiment_color: 可选的 (fc, bg, bd) 元组，来自"总结：共性优势/短板"的sentiment色
        """
        import html as html_mod
        
        if not managers:
            return ''
        
        # 如果有sentiment色，用它设置表格边框色
        border_color = sentiment_color[2] if sentiment_color else '#d9d9d9'
        
        html = [f'<table style="width:100%;border-collapse:collapse;font-size:12px;line-height:1.6;margin:8px 0;border:2px solid {border_color};border-radius:4px;">']
        # 表头
        html.append('<thead><tr>')
        headers = [
            ('干部', '#f0f5ff', '#333'),
            ('正向', '#f6ffed', cls.CLR_POS),
            ('待关注', '#fffbe6', cls.CLR_NEU),
        ]
        for label, bg, color in headers:
            html.append(f'<th style="padding:8px 10px;border:1px solid #d9d9d9;font-weight:700;text-align:center;background:{bg};color:{color};white-space:nowrap;">{label}</th>')
        html.append('</tr></thead>')
        
        # 数据行
        html.append('<tbody>')
        for mgr in managers:
            html.append('<tr>')
            # 名字列
            name = html_mod.escape(mgr['name'])
            html.append(f'<td style="padding:8px 10px;border:1px solid #e8e8e8;vertical-align:top;font-weight:600;white-space:nowrap;text-align:center;">{name}</td>')
            
            # 正向/待关注列
            for key, color in [('正向', cls.CLR_POS), ('待关注', cls.CLR_NEU)]:
                content = mgr.get(key, '').strip()
                if not content:
                    html.append(f'<td style="padding:8px 10px;border:1px solid #e8e8e8;vertical-align:top;text-align:left;color:#ccc;">—</td>')
                else:
                    formatted = cls._format_manager_cell(content, color)
                    html.append(f'<td style="padding:8px 10px;border:1px solid #e8e8e8;vertical-align:top;text-align:left;color:{color};">{formatted}</td>')
            
            html.append('</tr>')
        html.append('</tbody></table>')
        
        return '\n'.join(html)

    @classmethod
    def _format_manager_cell(cls, text, color, max_quotes=4):
        """格式化管理者表格单元格内容，保留人数统计，最多展示 max_quotes 条原文。"""
        import html as html_mod
        if not text:
            return ''
        
        # 提取人数标签（如"正向，3人"、"待关注，2人"）
        people_match = re.match(r'^(正向|待关注|负向|中立)[，,]\s*(\d+)\s*人', text)
        people_count = int(people_match.group(2)) if people_match else 0
        
        # 移除开头的 "正向，N人，" / "待关注，N人，" 等标签
        cleaned = re.sub(r'^(正向|待关注|负向|中立)[，,]\s*\d+\s*人[，,:：]*\s*', '', text).strip()
        
        prefix = ''
        if people_count:
            prefix = f'<div style="margin-bottom:6px;font-weight:700;font-size:11px;color:{color};">{people_count}人</div>'
        
        if not cleaned:
            return prefix if prefix else ''
        
        # 按 "" 分割为多条引用（每对引号内为一条），最多展示4条
        quotes = re.findall(r'["""]([^"""]+)["""]', cleaned)
        if not quotes:
            # 兜底：如果没有引号包裹，按；分割
            quotes = [q.strip() for q in re.split(r'[；;]\s*', cleaned) if q.strip()]
        # 过滤无效短引用
        _noise = {'bp', 'bp：', 'bp:', '；', ';', '无', ''}
        shown_parts = []
        for q in quotes:
            q = q.strip().strip('；; ')
            if not q or q.lower() in _noise or len(q) <= 3:
                continue
            escaped = html_mod.escape(q)
            shown_parts.append(f'"{escaped}"')
            if len(shown_parts) >= max_quotes:
                break
        
        if not shown_parts:
            return prefix if prefix else ''
        
        quote_blocks = []
        for p in shown_parts:
            quote_blocks.append(f'<div style="margin:4px 0;padding-left:8px;border-left:2px solid #d9d9d9;color:#555;">{p}</div>')
        
        return prefix + ''.join(quote_blocks)

    @classmethod
    def _render_body_lines_with_manager_table(cls, lines, max_quotes=2, dim_sentiment_color=None):
        """
        渲染正文行，与 _render_body_lines 相同，但当检测到"各管理者情况"
        区域时，将其解析为四列表格（名字/正向/中立/负向）。
        "1）总结：共性优势" 等标题行不显示，但其sentiment色会传递给管理者表格。
        dim_sentiment_color: 可选的 (fc, bg, bd) 元组，来自维度级别的sentiment判断（作为后备）。
        """
        import html as html_mod
        parts = []
        quote_count = 0
        
        # 检测"总结"行的sentiment色（如"1）总结：共性优势"→绿，"1）总结：共性短板"→红）
        summary_sentiment_color = None
        for line in lines:
            s = line.strip()
            if re.match(r'^(\d+[）)]|[一二三四五六七八九十]+[）)])\s*\**总结[：:]', s):
                fc, bg, bd = cls._sentiment_color(s)
                summary_sentiment_color = (fc, bg, bd)
                break
        
        # 先检测是否有"各管理者情况"区域
        manager_section_start = -1
        manager_section_end = -1
        
        for i, line in enumerate(lines):
            s = line.strip()
            s_plain = re.sub(r'\*+', '', s)
            if re.search(r'各管理者情况', s_plain) and (
                s_plain.startswith('#') or s_plain.startswith('-') or re.search(r'\d+[）)]', s_plain) or s_plain.endswith('：') or s_plain.endswith(':')
            ):
                manager_section_start = i + 1  # 标题行之后
                continue
            # 管理者区域结束标志：遇到下一个 #### 标题（非管理者名字）
            if manager_section_start >= 0 and manager_section_end < 0:
                if i > manager_section_start and (
                    (s.startswith('####') and not re.search(r'各管理者', s)) or
                    (s.startswith('###') and not re.search(r'各管理者', s))
                ):
                    manager_section_end = i
        
        # 如果管理者区域一直到末尾
        if manager_section_start >= 0 and manager_section_end < 0:
            manager_section_end = len(lines)
        
        # 有管理者区域则解析为表格
        if manager_section_start >= 0:
            manager_lines = lines[manager_section_start:manager_section_end]
            managers = cls._parse_manager_section(manager_lines)
            
            # 渲染三部分：管理者区域之前 + 表格 + 管理者区域之后
            before_lines = lines[:manager_section_start]
            after_lines = lines[manager_section_end:]
            
            # 渲染前部分（但去掉"各管理者情况"标题行本身——它已在 before_lines 最后一行）
            if before_lines:
                # 最后一行是 "各管理者情况" 标题，保留它作为表格标题
                title_line = before_lines[-1].strip()
                content_before = before_lines[:-1]
                before_parts = cls._render_body_lines(content_before, max_quotes=max_quotes)
                parts.extend(before_parts)
                
                # 添加标题
                title_text = re.sub(r'^#{1,4}\s*', '', title_line).strip()
                parts.append(f'<div style="margin:12px 0 6px;font-size:13px;font-weight:700;color:#444;">{cls._fmt(title_text)}</div>')
            
            # 渲染管理者表格（传入sentiment色，如果未检测到则回退到维度级别的配色）
            effective_color = summary_sentiment_color or dim_sentiment_color
            if managers:
                parts.append(cls._render_manager_table_html(managers, sentiment_color=effective_color))
            else:
                # 解析失败，降级回普通渲染
                fallback_parts = cls._render_body_lines(manager_lines, max_quotes=max_quotes)
                parts.extend(fallback_parts)
            
            # 渲染后部分
            if after_lines:
                after_parts = cls._render_body_lines(after_lines, max_quotes=max_quotes)
                parts.extend(after_parts)
        else:
            # 没有管理者区域，直接普通渲染
            parts = cls._render_body_lines(lines, max_quotes=max_quotes)
        
        return parts

    @classmethod
    def cadre_to_html(cls, text):
        """干部侧：提取跨维度整体判断表格、独立总体结论，并按判断标签渲染各维度卡片。"""
        if not text:
            return ''
        
        lines = text.split('\n')
        
        # ── 第一步：定位关键区域 ──
        first_dim_idx = -1
        summary_table_start = -1
        summary_section_start = -1
        conclusion_start = -1
        summary_text_after_table = []
        conclusion_lines = []
        
        for i, line in enumerate(lines):
            s = line.strip()
            if first_dim_idx == -1 and (
                re.match(r'^维度一[：:]', s) or
                re.match(r'^#{1,4}\s*\**维度一', s)
            ):
                first_dim_idx = i
            if summary_section_start == -1 and ('最终整体判断' in s or '最终维度总结' in s):
                summary_section_start = i
            if conclusion_start == -1 and '总体结论' in s:
                conclusion_start = i
        
        summary_headers = []
        summary_rows = []
        summary_table_end = -1
        if summary_section_start >= 0:
            for i in range(summary_section_start, len(lines)):
                s = lines[i].strip()
                if '|' in s and re.match(r'^\|?\s*\*?\*?维度', s.replace('|', ' ').strip()):
                    summary_headers, summary_rows, summary_table_end = cls._parse_md_table(lines, i)
                    summary_table_start = i
                    break
                if '|' in s and re.search(r'^\|.*维度.*\|.*判断', s):
                    summary_headers, summary_rows, summary_table_end = cls._parse_md_table(lines, i)
                    summary_table_start = i
                    break
        
        if first_dim_idx == -1:
            first_dim_idx = 0
        
        if summary_table_end > 0:
            # 表格后、维度正文前的文字才是补充说明
            # 取 first_dim_idx 和 conclusion_start 中在 summary_table_end 之后且最小的那个
            candidates = []
            if first_dim_idx > summary_table_end:
                candidates.append(first_dim_idx)
            if conclusion_start > summary_table_end:
                candidates.append(conclusion_start)
            summary_text_end = min(candidates) if candidates else len(lines)
            for i in range(summary_table_end, summary_text_end):
                s = lines[i].strip()
                if s and s != '---' and '总体结论' not in s:
                    summary_text_after_table.append(s)
        
        if conclusion_start >= 0:
            for i in range(conclusion_start + 1, len(lines)):
                s = lines[i].strip()
                if i >= first_dim_idx and first_dim_idx > conclusion_start:
                    break
                if s == '---' and conclusion_lines:
                    break
                if (s.startswith('### ') or s.startswith('## ')) and i > conclusion_start + 1 and conclusion_lines:
                    break
                if s and s != '---':
                    conclusion_lines.append(s)
        
        # (C) 维度分析正文 — 从维度一开始，到总体结论之前
        if conclusion_start > first_dim_idx:
            body_lines = lines[first_dim_idx:conclusion_start]
        else:
            body_lines = lines[first_dim_idx:]
        
        # ── 将 body_lines 按维度切分（提前到兜底预计算之前） ──
        dim_groups = []
        current_title = None
        current_lines = []
        
        for line in body_lines:
            s = line.strip()
            is_dim_title = (
                re.match(r'^维度[一二三四五六七八九十]+[：:]', s) or
                re.match(r'^#{1,4}\s*\**维度[一二三四五六七八九十]+', s)
            )
            if is_dim_title:
                if current_title is not None:
                    dim_groups.append((current_title, current_lines))
                current_title = s
                current_lines = []
            else:
                current_lines.append(line)
        if current_title is not None:
            dim_groups.append((current_title, current_lines))
        
        # ── 兜底预计算：遍历所有维度，收集判断修正映射 ──
        # 用于修正跨维度整体判断表格中的标签
        _fallback_dim_map = {}  # { 维度名关键词: 修正后的标签 }
        if dim_groups:
            for _fb_title, _fb_lines in dim_groups:
                _fb_other = []
                for _fbl in _fb_lines:
                    _fbs = _fbl.strip()
                    if _fbs.startswith('（整体判断') or _fbs.startswith('(整体判断'):
                        continue
                    if re.match(r'^#{0,4}\s*\d+[）)]\s*\**整体判断', _fbs):
                        continue
                    if re.match(r'^\*{2}整体判断[：:]', _fbs):
                        continue
                    _fb_other.append(_fbl)
                
                # 找管理者区域
                _fb_mgr_start = -1
                for _fci, _fcl in enumerate(_fb_other):
                    _fcs = _fcl.strip()
                    _fcs_plain = re.sub(r'\*+', '', _fcs)
                    if re.search(r'各管理者情况', _fcs_plain) and (
                        _fcs_plain.startswith('#') or _fcs_plain.startswith('-') or
                        re.search(r'\d+[）)]', _fcs_plain) or _fcs_plain.endswith('：') or _fcs_plain.endswith(':')
                    ):
                        _fb_mgr_start = _fci + 1
                        break
                
                if _fb_mgr_start >= 0:
                    _fb_mgr_lines = _fb_other[_fb_mgr_start:]
                    _fb_mgr_preview = cls._parse_manager_section(_fb_mgr_lines)
                    if _fb_mgr_preview:
                        def _fb_has_real(fv):
                            v = (fv or '').strip()
                            if not v:
                                return False
                            if re.match(r'^(正向|待关注|负向|中立)[，,]\s*0\s*人[，,：:]*\s*$', v):
                                return False
                            # 过滤"(无数据)""无数据""无相关数据"等缺失标记
                            if re.match(r'^[（(]?无.{0,4}数据[）)]?$', v):
                                return False
                            if re.match(r'^[（(]?缺失[）)]?$', v):
                                return False
                            return True
                        
                        _fb_has_pos = any(_fb_has_real(m.get('正向', '')) for m in _fb_mgr_preview)
                        _fb_has_att = any(_fb_has_real(m.get('待关注', '')) for m in _fb_mgr_preview)
                        
                        # 提取维度名关键词（去掉"维度N："前缀和markdown标记）
                        _fb_dim_name = re.sub(r'^#{1,4}\s*\**', '', _fb_title).strip().rstrip('*')
                        _fb_dim_name = re.sub(r'^维度[一二三四五六七八九十]+[：:]\s*', '', _fb_dim_name).strip()
                        
                        if not _fb_has_pos and _fb_has_att:
                            _fallback_dim_map[_fb_dim_name] = '共性不足'
                        elif _fb_has_pos and not _fb_has_att:
                            _fallback_dim_map[_fb_dim_name] = '共性优势'
                        elif not _fb_has_pos and not _fb_has_att:
                            # 完全无数据（既无正向也无待关注）→ 标记为缺失
                            _fallback_dim_map[_fb_dim_name] = '缺失'
                        # 既有正向又有待关注 → 不覆盖，保留原始判断
        
        # ── 第二步：构建HTML ──
        html_parts = []
        
        # 修正跨维度表格中的判断列
        if _fallback_dim_map and summary_headers and summary_rows:
            # 找到判断列和维度列的索引
            _s_judgment_idx = -1
            _s_dim_idx = -1
            for _si, _sh in enumerate(summary_headers):
                _sh_clean = re.sub(r'\*+', '', _sh).strip()
                if '判断' in _sh_clean:
                    _s_judgment_idx = _si
                if '维度' in _sh_clean:
                    _s_dim_idx = _si
            
            if _s_judgment_idx >= 0 and _s_dim_idx >= 0:
                for _sr in summary_rows:
                    if _s_dim_idx < len(_sr) and _s_judgment_idx < len(_sr):
                        _sr_dim_text = re.sub(r'\*+', '', _sr[_s_dim_idx]).strip()
                        # 提取维度名关键词（去掉"维度N："前缀）
                        _sr_dim_key = re.sub(r'^维度[一二三四五六七八九十]+[：:]\s*', '', _sr_dim_text).strip()
                        for _fb_key, _fb_label in _fallback_dim_map.items():
                            if _fb_key in _sr_dim_key or _sr_dim_key in _fb_key:
                                # 修正判断列文本
                                _old_judgment = _sr[_s_judgment_idx]
                                _new_judgment = re.sub(
                                    r'局部待关注|局部问题|局部不足|局部优势|共性优势|共性不足|共性短板',
                                    _fb_label, _old_judgment
                                )
                                # 如果原文没匹配到任何关键词但需要修正，直接替换
                                if _new_judgment == _old_judgment and _fb_label not in _old_judgment:
                                    _new_judgment = re.sub(r'\*+', '', _old_judgment).strip()
                                    _new_judgment = f'**{_fb_label}**'
                                _sr[_s_judgment_idx] = _new_judgment
                                break
        
        # (A) 跨维度整体判断表格
        if summary_headers and summary_rows:
            html_parts.append('<div style="margin-bottom:16px;">')
            html_parts.append('<div style="font-size:14px;font-weight:700;color:#1a1a1a;margin-bottom:8px;">📊 最终整体判断（跨维度）</div>')
            html_parts.append(cls._render_table_html(summary_headers, summary_rows, highlight_sentiment=True))
            for t in summary_text_after_table:
                html_parts.append(f'<div style="margin:6px 0;font-size:13px;line-height:1.7;color:#555;">{cls._fmt(t)}</div>')
            html_parts.append('</div>')
        
        # (B) 总体结论 — 样式参考组织部分"结语"
        if conclusion_lines:
            html_parts.append('<div style="margin:0 0 16px;padding:12px 16px;background:linear-gradient(135deg,#f0f5ff,#e6f7ff);border:1px solid #91d5ff;border-radius:8px;">')
            html_parts.append('<div style="font-size:13px;font-weight:700;color:#1890ff;margin-bottom:6px;">💡 总体结论</div>')
            for line in conclusion_lines:
                clean_line = line.strip('*').strip()
                clean_line = re.sub(r'^总体结论[：:]\s*', '', clean_line).strip()
                if not clean_line:
                    continue
                html_parts.append(f'<div style="font-size:13px;line-height:1.8;color:#333;">{cls._fmt(clean_line)}</div>')
            html_parts.append('</div>')
        
        
        if not dim_groups:
            body_parts = cls._render_body_lines(body_lines, max_quotes=2)
            html_parts.extend(body_parts)
        else:
            for dim_title, dim_lines in dim_groups:
                judgment_lines = []
                other_lines = []
                in_judgment_section = False
                
                for dl in dim_lines:
                    ds = dl.strip()
                    if ds.startswith('（整体判断') or ds.startswith('(整体判断'):
                        judgment_lines.append(ds)
                        continue
                    judgment_title_match = re.match(r'^#{0,4}\s*\d+[）)]\s*\**整体判断[：:]\s*(.*)', ds)
                    if judgment_title_match or re.match(r'^#{0,4}\s*\d+[）)]\s*\**整体判断\**\s*$', ds):
                        in_judgment_section = True
                        if judgment_title_match:
                            inline_text = judgment_title_match.group(1).strip().rstrip('*').strip()
                            if inline_text:
                                judgment_lines.append(inline_text)
                        continue
                    # 匹配 **整体判断：xxx** 格式（无数字前缀）
                    bold_judgment_match = re.match(r'^\*{2}整体判断[：:]\s*(.*?)\*{2}\s*$', ds)
                    if bold_judgment_match:
                        inline_text = bold_judgment_match.group(1).strip()
                        if inline_text:
                            judgment_lines.append(inline_text)
                        continue
                    if in_judgment_section:
                        if (
                            ds.startswith('####') or ds == '---' or
                            re.match(r'^#{0,4}\s*\d+[）)]\s*\**各管理者情况', ds) or
                            re.match(r'^#{0,4}\s*\d+[）)]\s*\**管理者情况', ds)
                        ):
                            in_judgment_section = False
                            other_lines.append(dl)
                            continue
                        if ds:
                            judgment_lines.append(ds)
                        continue
                    other_lines.append(dl)
                
                color_label = ' '.join(judgment_lines) if judgment_lines else ''
                
                # ── 兜底判断：根据管理者实际数据修正整体判断标签 ──
                # 提前解析管理者区域，检查正向/待关注的分布情况
                _mgr_lines_for_check = []
                _mgr_check_start = -1
                for _ci, _cl in enumerate(other_lines):
                    _cs = _cl.strip()
                    _cs_plain = re.sub(r'\*+', '', _cs)
                    if re.search(r'各管理者情况', _cs_plain) and (
                        _cs_plain.startswith('#') or _cs_plain.startswith('-') or
                        re.search(r'\d+[）)]', _cs_plain) or _cs_plain.endswith('：') or _cs_plain.endswith(':')
                    ):
                        _mgr_check_start = _ci + 1
                        break
                if _mgr_check_start >= 0:
                    _mgr_lines_for_check = other_lines[_mgr_check_start:]
                    _mgr_preview = cls._parse_manager_section(_mgr_lines_for_check)
                    if _mgr_preview:
                        # 判断某字段是否有实质内容（排除 "待关注，0人" 等仅含标签的空内容）
                        def _has_real_content(field_value):
                            v = (field_value or '').strip()
                            if not v:
                                return False
                            # "正向，0人" / "待关注，0人" → 无实质内容
                            if re.match(r'^(正向|待关注|负向|中立)[，,]\s*0\s*人[，,：:]*\s*$', v):
                                return False
                            # 只有标签无引用也算有内容（如 "待关注，3人"）
                            return True
                        
                        _has_positive = any(_has_real_content(m.get('正向', '')) for m in _mgr_preview)
                        _has_attention = any(_has_real_content(m.get('待关注', '')) for m in _mgr_preview)
                        if not _has_positive and _has_attention:
                            # 所有人都没有正向观点，只有待关注 → 共性不足
                            color_label = '共性不足'
                            judgment_lines = [re.sub(r'局部待关注|局部问题|局部不足|局部优势|共性优势', '共性不足', jl) for jl in judgment_lines]
                        elif _has_positive and not _has_attention:
                            # 所有人都没有待关注观点，只有正向 → 共性优势
                            color_label = '共性优势'
                            judgment_lines = [re.sub(r'局部待关注|局部问题|局部不足|共性不足|共性短板', '共性优势', jl) for jl in judgment_lines]
                        # 既有正向又有待关注 → 保留原始判断
                
                kind, dim_fc, dim_bg, dim_bd = cls._judgment_theme(color_label)
                if kind == 'default':
                    dim_fc, dim_bg, dim_bd = '#222', '#fff', '#e8e8e8'
                
                html_parts.append(f'<div style="margin:12px 0;padding:12px 16px;background:{dim_bg};border:2px solid {dim_bd};border-radius:10px;">')
                
                title_clean = re.sub(r'^#{1,4}\s*\**', '', dim_title).strip().rstrip('*')
                html_parts.append(f'<div style="margin:0 0 6px;font-size:14px;font-weight:700;color:{dim_fc};">{cls._fmt(title_clean)}</div>')
                
                # ── 兜底：如果 judgment_lines 只有标签行（如"共性优势"），无实质summary，
                #    从 other_lines 开头提取 summary 段落纳入虚线框 ──
                _jl_total_len = sum(len(jl.strip()) for jl in judgment_lines)
                _is_short_label = (len(judgment_lines) <= 1 and _jl_total_len < 20) or (len(judgment_lines) == 0)
                if _is_short_label and other_lines:
                    _absorbed = []
                    _remaining = []
                    _absorbing = True
                    for _ol in other_lines:
                        _os = _ol.strip()
                        if _absorbing:
                            # 跳过空行/空div
                            if not _os:
                                _absorbed.append(_ol)
                                continue
                            # 如果遇到"总结"/"各管理者情况"/表格标记/分隔线，停止吸纳
                            _os_plain = re.sub(r'\*+', '', _os)
                            if (re.match(r'^#{0,4}\s*\d+[）)]\s*总结', _os_plain) or
                                re.search(r'各管理者情况', _os_plain) or
                                _os.startswith('|') or _os == '---' or
                                re.match(r'^#{1,4}\s', _os)):
                                _absorbing = False
                                _remaining.append(_ol)
                                continue
                            # 正文段落 → 纳入 judgment
                            judgment_lines.append(_os)
                            _absorbed.append(_ol)
                        else:
                            _remaining.append(_ol)
                    other_lines = _remaining
                
                if judgment_lines:
                    html_parts.append(f'<div style="margin:6px 0 8px;padding:8px 10px;background:rgba(255,255,255,0.55);border:1px dashed {dim_bd};border-radius:8px;">')
                    for jl in judgment_lines:
                        clean = re.sub(r'^\d+[）)]\s*整体判断[：:]\s*', '', jl).strip()
                        clean = clean.strip('（）()')
                        if clean.startswith('整体判断：') or clean.startswith('整体判断:'):
                            clean = re.sub(r'^整体判断[：:]\s*', '', clean)
                        icon = ''
                        if clean.startswith('✅'):
                            icon = '✅ '
                            clean = clean.lstrip('✅ ').strip()
                        elif clean.startswith('⚠️') or clean.startswith('⚠'):
                            icon = '⚠️ '
                            clean = clean.lstrip('⚠️⚠ ').strip()
                        html_parts.append(f'<div style="margin:0 0 4px;font-size:13px;line-height:1.7;color:#222;">{icon}{cls._fmt(clean)}</div>')
                    html_parts.append('</div>')
                
                body_parts = cls._render_body_lines_with_manager_table(other_lines, max_quotes=2, dim_sentiment_color=(dim_fc, dim_bg, dim_bd))
                html_parts.extend(body_parts)
                html_parts.append('</div>')
        
        return '\n'.join(html_parts)

    @classmethod
    def org_to_html(cls, text):
        """组织侧：适配新格式 — 四维度概览表格 + 四大章节卡片（业务定位/上下游关系/战略方向/组织设计）。
        
        新格式结构：
        1. 顶部概览表格（维度 | 是否存在问题 | 描述）
        2. 正文四章节（### 一、业务定位 / ### 二、上下游关系 / ### 三、战略方向 / ### 四、组织设计）
        3. 每章开头可选有单行表格小结（| 暂无明显问题... |）
        4. 底部可选有综合结论表格、附录信源汇总表、注释
        """
        if not text:
            return ''
        
        lines = text.split('\n')
        
        # ── 第一步：定位关键区域 ──
        
        # (1) 找顶部概览表格（维度 | 是否存在问题 | 描述）
        overview_headers = []
        overview_rows = []
        overview_table_end = -1
        
        for i, line in enumerate(lines):
            s = line.strip()
            if s.startswith('|') and s.endswith('|') and '维度' in s and s.count('|') >= 3:
                if i + 1 < len(lines) and re.match(r'^\s*\|[\s\-:]+\|', lines[i + 1]):
                    overview_headers, overview_rows, overview_table_end = cls._parse_md_table(lines, i)
                    break
        
        # (2) 找正文章节（### 一、xxx / ## 一、xxx）
        section_starts = []  # [(line_idx, title_text), ...]
        for i, line in enumerate(lines):
            s = line.strip()
            if re.match(r'^#{1,4}\s*[一二三四五六七八九十]+、\s*\**', s):
                title = re.sub(r'^#{1,4}\s*', '', s).strip()
                section_starts.append((i, title))
        
        # (3) 找底部区域（综合结论/总结概览/附录/注释）
        bottom_start = len(lines)
        for i, line in enumerate(lines):
            s = line.strip()
            # 综合结论 / 总结概览（底部重复表格）
            if re.match(r'^#{1,4}\s*(综合结论|总结概览|总结|结论)', s) and i > overview_table_end:
                # 确保不是正文中的章节（检查是否在最后一个正文章节之后）
                if section_starts and i > section_starts[-1][0]:
                    bottom_start = min(bottom_start, i)
            # 附录/信源汇总
            if re.match(r'^#{1,4}\s*附[：:]', s) or '信源汇总' in s:
                bottom_start = min(bottom_start, i)
        
        # 找底部注释（> **注**：...）
        note_lines = []
        for i in range(max(0, bottom_start - 5), len(lines)):
            s = lines[i].strip()
            if s.startswith('>') and re.search(r'注[：:]|注\*\*[：:]', s):
                cleaned = re.sub(r'^>\s*\*?\*?注\*?\*?[：:]\s*', '', s).strip()
                if cleaned:
                    note_lines.append(cleaned)
                # 继续收集下一行
                for j in range(i + 1, len(lines)):
                    sj = lines[j].strip()
                    if sj.startswith('>'):
                        cl = sj.lstrip('> ').strip()
                        if cl:
                            note_lines.append(cl)
                    else:
                        break
                break
        
        # 也检查非引用块形式的注释（如结尾处的 > **注**：...）
        if not note_lines:
            for i in range(len(lines) - 1, max(0, len(lines) - 10), -1):
                s = lines[i].strip()
                if s.startswith('>') and re.search(r'注[：:]|注\*\*[：:]', s):
                    cleaned = re.sub(r'^>\s*\*?\*?注\*?\*?[：:]\s*', '', s).strip()
                    if cleaned:
                        note_lines.append(cleaned)
                    break
        
        # (4) 找附录信源汇总表
        appendix_headers = []
        appendix_rows = []
        for i in range(bottom_start, len(lines)):
            s = lines[i].strip()
            if '信源汇总' in s:
                # 在其后找表格
                for j in range(i + 1, min(i + 5, len(lines))):
                    sj = lines[j].strip()
                    if sj.startswith('|') and sj.endswith('|') and sj.count('|') >= 3:
                        if j + 1 < len(lines) and re.match(r'^\s*\|[\s\-:]+\|', lines[j + 1]):
                            appendix_headers, appendix_rows, _ = cls._parse_md_table(lines, j)
                            break
                break
        
        # (5) 构建维度问题映射（从概览表格中提取每个维度是否有问题）
        dim_has_problem = {}  # { '业务定位': True/False, ... }
        if overview_headers and overview_rows:
            # 找到"是否存在问题"列的索引
            problem_col = -1
            dim_col = -1
            for ci, h in enumerate(overview_headers):
                h_clean = re.sub(r'\*+', '', h).strip()
                if '是否' in h_clean or '存在问题' in h_clean:
                    problem_col = ci
                if '维度' in h_clean:
                    dim_col = ci
            if problem_col >= 0 and dim_col >= 0:
                for row in overview_rows:
                    if dim_col < len(row) and problem_col < len(row):
                        dim_name = re.sub(r'\*+', '', row[dim_col]).strip()
                        problem_val = re.sub(r'\*+', '', row[problem_col]).strip()
                        dim_has_problem[dim_name] = ('是' in problem_val)
        
        # ── 第二步：构建HTML ──
        html_parts = []
        
        # (A) 概览表格置顶 — 自定义渲染（"是"红色，"否"绿色）
        if overview_headers and overview_rows:
            html_parts.append('<div style="margin-bottom:16px;">')
            html_parts.append(cls._render_org_overview_table(overview_headers, overview_rows))
            html_parts.append('</div>')
        
        # (B) 正文章节 — 每个章节渲染为彩色卡片
        for sec_idx, (start_line, title_text) in enumerate(section_starts):
            # 确定章节内容范围
            if sec_idx + 1 < len(section_starts):
                end_line = section_starts[sec_idx + 1][0]
            else:
                end_line = bottom_start
            
            content_lines = lines[start_line + 1:end_line]
            
            # 清理标题
            title_clean = re.sub(r'\*+', '', title_text).strip()
            # 提取维度名（去掉"一、"前缀）
            dim_key = re.sub(r'^[一二三四五六七八九十]+、\s*', '', title_clean).strip()
            
            # 判断是否存在问题（从概览表映射中查找）
            has_problem = None
            for dk, hp in dim_has_problem.items():
                if dim_key in dk or dk in dim_key:
                    has_problem = hp
                    break
            
            # 如果概览表没有映射到，默认无问题（不做文本猜测）
            if has_problem is None:
                has_problem = False
            
            # 选择颜色方案
            if has_problem:
                fc = cls.CLR_NEG          # 红色文字
                bg = '#ffffff'            # 白色背景
                bd = '#e8e8e8'            # 灰色边框
                title_accent = cls.CLR_NEG  # 标题左边条红色
                icon = '⚠️'
                heading_style = 'red'
            else:
                fc = '#1890ff'            # 蓝色文字
                bg = '#f0f5ff'            # 淡蓝背景
                bd = '#91d5ff'            # 蓝色边框
                title_accent = '#1890ff'
                icon = '✅'
                heading_style = 'blue'
            
            # 渲染章节卡片
            html_parts.append(f'<div style="margin:12px 0;padding:12px 16px;background:{bg};border:1px solid {bd};border-radius:10px;">')
            html_parts.append(f'<div style="margin:0 0 8px;padding:8px 14px;font-size:14px;font-weight:700;color:{fc};border-left:4px solid {title_accent};border-radius:0 6px 6px 0;">{icon} {cls._fmt(title_clean)}</div>')
            
            # 渲染章节内容
            rendered_body = ''.join(cls._render_body_lines(content_lines, max_quotes=3, org_number_heading_style=heading_style, section_has_problem=has_problem))
            
            # 如果是红色章节，将子标题也染红
            if has_problem:
                rendered_body = rendered_body.replace(
                    'margin:20px 0 10px;font-size:15px;font-weight:700;color:#1a1a1a;border-left:4px solid #1890ff;padding-left:10px;',
                    f'margin:20px 0 10px;font-size:15px;font-weight:700;color:{fc};border-left:4px solid {title_accent};padding-left:10px;'
                )
            
            html_parts.append(rendered_body)
            html_parts.append('</div>')
        
        # (C) 附录信源汇总表 — 跳过，因为各维度内部已有信源引用，避免重复
        # if appendix_headers and appendix_rows:
        #     html_parts.append('<div style="margin:16px 0 8px;">')
        #     html_parts.append(f'<div style="font-size:13px;font-weight:700;color:#888;margin-bottom:8px;">📋 信源汇总</div>')
        #     html_parts.append(cls._render_table_html(appendix_headers, appendix_rows, highlight_sentiment=False, force_left_align=True))
        #     html_parts.append('</div>')
        
        # (D) 底部注释
        if note_lines:
            html_parts.append('<div style="margin:12px 0 0;padding:10px 14px;background:#fafafa;border:1px solid #e8e8e8;border-radius:6px;font-size:12px;color:#999;line-height:1.7;">')
            for nl in note_lines:
                html_parts.append(f'{cls._fmt(nl)}')
            html_parts.append('</div>')
        
        return '\n'.join(html_parts)
    
    @classmethod
    def _render_org_overview_table(cls, headers, rows):
        """渲染组织侧概览表格 — 蓝底白字表头，无纵向边框，横向灰色分割线，有问题行第二列红字。"""
        import html as html_mod
        
        # 找到"是否存在问题"列索引
        problem_col = -1
        for ci, h in enumerate(headers):
            h_clean = re.sub(r'\*+', '', h).strip()
            if '是否' in h_clean or '存在问题' in h_clean:
                problem_col = ci
                break
        
        header_texts = [re.sub(r'\*+', '', (h or '')).strip() for h in headers]
        
        # 列宽分配：第1列(维度)12%、第2列(是否存在问题)12%、第3列(描述)76%
        col_widths = ['12%', '12%', '76%'] if len(header_texts) >= 3 else None
        
        # 表头 — 蓝底白字，无纵向边框，整个表格加外框
        table = '<table style="width:100%;border-collapse:collapse;font-size:13px;table-layout:fixed;border:1px solid #d9d9d9;border-radius:6px;">'
        if col_widths:
            table += '<colgroup>'
            for w in col_widths:
                table += f'<col style="width:{w};">'
            table += '</colgroup>'
        table += '<thead><tr>'
        for hi, ht in enumerate(header_texts):
            # 前两列居中，最后一列靠左
            th_align = 'text-align:center;' if hi < len(header_texts) - 1 else 'text-align:left;'
            table += f'<th style="padding:10px 12px;background:#1a6fb5;color:#fff;font-weight:600;{th_align}border:none;border-bottom:2px solid #1a6fb5;">{html_mod.escape(ht)}</th>'
        table += '</tr></thead><tbody>'
        
        for ri, row in enumerate(rows):
            # 判断这一行是否有问题
            has_problem = False
            if problem_col >= 0 and problem_col < len(row):
                val = re.sub(r'\*+', '', row[problem_col]).strip()
                has_problem = ('是' in val)
            
            border_bottom = 'border-bottom:1px solid #e8e8e8;'
            
            table += '<tr style="background:#fff;">'
            for ci, cell in enumerate(row):
                cell_clean = re.sub(r'\*+', '', (cell or '')).strip()
                cell_html = cls._fmt(cell_clean)
                cell_html = cell_html.replace('&lt;br&gt;', '<br>')
                cell_html = cell_html.replace('&lt;br/&gt;', '<br>')
                
                # 无纵向边框，仅横向灰色分割线
                base_style = f'padding:10px 12px;border:none;{border_bottom}line-height:1.6;vertical-align:top;'
                
                if ci == problem_col:
                    # "是否存在问题"列：有问题红字，没问题绿字，居中
                    fc = cls.CLR_NEG if has_problem else cls.CLR_POS
                    table += f'<td style="{base_style}color:{fc};font-weight:700;white-space:nowrap;text-align:center;">{cell_html}</td>'
                elif ci == 0:
                    # 维度列：居中
                    table += f'<td style="{base_style}color:#333;font-weight:600;text-align:center;">{cell_html}</td>'
                else:
                    # 最后一列（描述）：靠左
                    table += f'<td style="{base_style}color:#555;text-align:left;">{cell_html}</td>'
            table += '</tr>'
        
        table += '</tbody></table>'
        return table


def generate_report_data(diag_row: List, diag_row_idx: int, 
                        feedback_reader: ExcelReader,
                        jm_loader=None, jm_detail_loader=None,
                        yd_loader=None, resignation_analyzer=None,
                        bp_loader=None, jiangan_loader=None,
                        open_fb_loader=None, jm_open_loader=None,
                        wc_loader=None) -> Dict:
    """生成报告数据"""
    
    org_full_path = diag_row[2] if len(diag_row) > 2 else ""
    # 清理零宽字符（U+200B等），避免匹配失败
    org_full_path = re.sub(r'[\u200b\u200c\u200d\u200e\u200f\ufeff]', '', org_full_path)
    org_name = simplify_org_name(org_full_path)
    leader_name = diag_row[5] if len(diag_row) > 5 else "未知"
    
    # K=全面反馈, L=异动, M=敬满, N=最终结果
    k_val = diag_row[10] if len(diag_row) > 10 else ""
    l_val = diag_row[11] if len(diag_row) > 11 else ""
    m_val = diag_row[12] if len(diag_row) > 12 else ""  # M列：敬满
    n_val = diag_row[13] if len(diag_row) > 13 else ""
    
    k_color = k_val.replace('灯', '') if k_val else ""
    l_color = l_val.replace('灯', '') if l_val else ""
    m_color = m_val.replace('灯', '') if m_val and m_val != '缺失' else ""
    
    warning_level = n_val if n_val else "未知"
    # 构建三维度状态
    jingman_text = color_to_text(m_color) if m_color else ('数据缺失' if m_val == '缺失' else '暂无数据')
    
    # 严重预警/预警的解释（拆分为两行，去掉括号）
    warning_explain_line1 = "严重预警：触发多个异常指标，或偏离程度较大"
    warning_explain_line2 = "预警：触发部分异常指标，且偏离程度不大"
    
    # --- 生成三维度评估描述 ---
    # 维度名称映射
    dim_names = {'full_feedback': '全面反馈', 'abnormal': '异动', 'jingman': '敬满'}
    dim_colors = {'full_feedback': k_color, 'abnormal': l_color, 'jingman': m_color}
    dim_missing = {'full_feedback': False, 'abnormal': False, 
                   'jingman': (not m_color and (m_val == '缺失' or not m_val))}
    
    # 分类各维度
    severe_dims = [k for k, v in dim_colors.items() if v == '红']   # 严重预警（红灯）
    warning_dims = [k for k, v in dim_colors.items() if v == '黄']  # 预警（黄灯）
    normal_dims = [k for k, v in dim_colors.items() if v == '绿']   # 正常（绿灯）
    missing_dims = [k for k, v in dim_missing.items() if v]         # 数据缺失
    
    all_warning_dims = severe_dims + warning_dims  # 所有预警维度（红+黄）
    n_warnings = len(all_warning_dims)
    n_severe = len(severe_dims)
    
    def dim_name(k): return dim_names[k]
    def dim_names_join(keys): return '和'.join(dim_name(k) for k in keys)
    
    if n_warnings == 0 and len(missing_dims) == 0:
        # 情况1：所有维度都正常 → 全绿
        tri_dim_desc = "没有维度预警，定为【全绿】"
    elif n_warnings == 0 and len(missing_dims) > 0:
        # 情况3：无预警但有缺失
        tri_dim_desc = f"没有维度预警，但{dim_names_join(missing_dims)}维度缺失，定为【无预警】"
    elif n_warnings == 1:
        # 情况2：仅1个维度预警
        the_dim = all_warning_dims[0]
        level_word = "严重预警" if the_dim in severe_dims else "预警"
        tri_dim_desc = f"仅{dim_name(the_dim)}单维度{level_word}，综合判定无预警"
    elif n_warnings == 2 and n_severe == 0:
        # 情况4：2个维度预警，无严重预警 → 三级预警
        tri_dim_desc = f"{dim_names_join(all_warning_dims)}维度预警，但无严重预警，定为【三级预警】"
    elif n_warnings == 2 and n_severe == 1:
        # 情况5：2个维度预警，1个严重预警 → 二级预警
        w_dim = warning_dims[0]
        s_dim = severe_dims[0]
        tri_dim_desc = f"{dim_name(w_dim)}预警，{dim_name(s_dim)}严重预警，定为【二级预警】"
    elif n_warnings == 2 and n_severe == 2:
        # 情况9：2个严重预警 → 一级预警
        tri_dim_desc = f"{dim_names_join(severe_dims)}为严重预警，定为【一级预警】"
    elif n_warnings == 3 and n_severe == 0:
        # 情况6：三维度均预警，无严重预警 → 一级预警
        tri_dim_desc = "三维度均预警，定为【一级预警】"
    elif n_warnings == 3 and n_severe == 3:
        # 情况8：三维度均严重预警 → 一级预警
        tri_dim_desc = "三维度均严重预警，定为【一级预警】"
    elif n_warnings == 3 and n_severe > 0:
        # 情况7：三维度均预警，部分严重预警 → 一级预警
        tri_dim_desc = f"三维度均预警，且{dim_names_join(severe_dims)}为严重预警，定为【一级预警】"
    else:
        tri_dim_desc = f"综合判定{warning_level}"
    
    # 兼容旧字段
    warning_desc = tri_dim_desc
    warning_note = tri_dim_desc
    
    report = {
        'org_full_path': org_full_path,
        'org_name': org_name,
        'leader_name': leader_name,
        'warning_level': warning_level,
        'warning_desc': warning_desc,
        'warning_note': warning_note,
        'tri_dim_desc': tri_dim_desc,
        'warning_explain_line1': warning_explain_line1,
        'warning_explain_line2': warning_explain_line2,
        'dimensions': {
            'full_feedback': {'color': k_color, 'text': color_to_text(k_color)},
            'abnormal': {'color': l_color, 'text': color_to_text(l_color)},
            'jingman': {'color': m_color, 'text': color_to_text(m_color) if m_color else ('数据缺失' if m_val == '缺失' else '暂无数据')}
        }
    }
    
    # 查找负责人反馈
    if leader_name and leader_name != "未知":
        matches = feedback_reader.find_rows(0, leader_name)
        if matches:
            leader_feedback = matches[0][1]
            
            # 保存原始数值
            total_rank_raw = leader_feedback[2] if len(leader_feedback) > 2 else ''
            peer_rank_raw = leader_feedback[3] if len(leader_feedback) > 3 else ''
            subordinate_raw = leader_feedback[5] if len(leader_feedback) > 5 else ''
            peer_coop_raw = leader_feedback[6] if len(leader_feedback) > 6 else ''
            competent_raw = leader_feedback[7] if len(leader_feedback) > 7 else ''
            negative_raw = leader_feedback[8] if len(leader_feedback) > 8 else '0'
            
            report['leader_feedback'] = {
                'total_rank': format_percentage(total_rank_raw, True),
                'peer_rank': format_percentage(peer_rank_raw, True),
                'subordinate_follow': format_percentage(subordinate_raw),
                'peer_cooperate': format_percentage(peer_coop_raw),
                'competent': format_percentage(competent_raw),
                'negative_feedback': str(negative_raw),
                # 保存原始值用于变色判断
                '_total_rank_raw': safe_float(total_rank_raw),
                '_peer_rank_raw': safe_float(peer_rank_raw),
                '_subordinate_raw': safe_float(subordinate_raw),
                '_peer_coop_raw': safe_float(peer_coop_raw),
                '_competent_raw': safe_float(competent_raw),
                '_negative_raw': safe_int(negative_raw),
                # 雷达图数据
                'radar': {
                    'allin': {
                        'insight_see': {
                            'subordinate': safe_float(leader_feedback[73] if len(leader_feedback) > 73 else None),
                            'peer': safe_float(leader_feedback[74] if len(leader_feedback) > 74 else None)
                        },
                        'insight_decide': {
                            'subordinate': safe_float(leader_feedback[76] if len(leader_feedback) > 76 else None),
                            'peer': safe_float(leader_feedback[77] if len(leader_feedback) > 77 else None)
                        },
                        'inspire_cooperate': {
                            'subordinate': 0,  # 无下级数据
                            'peer': safe_float(leader_feedback[82] if len(leader_feedback) > 82 else None)
                        },
                        'inspire_lead': {
                            'subordinate': safe_float(leader_feedback[84] if len(leader_feedback) > 84 else None),
                            'peer': 0  # 无同级数据
                        },
                        'win_result': {
                            'subordinate': safe_float(leader_feedback[88] if len(leader_feedback) > 88 else None),
                            'peer': 0  # 无同级数据
                        },
                        'win_manage': {
                            'subordinate': safe_float(leader_feedback[90] if len(leader_feedback) > 90 else None),
                            'peer': 0  # 无同级数据
                        }
                    },
                    'values': {
                        'user': {
                            'subordinate': safe_float(leader_feedback[93] if len(leader_feedback) > 93 else None),
                            'peer': safe_float(leader_feedback[94] if len(leader_feedback) > 94 else None)
                        },
                        'integrity': {
                            'subordinate': safe_float(leader_feedback[96] if len(leader_feedback) > 96 else None),
                            'peer': safe_float(leader_feedback[97] if len(leader_feedback) > 97 else None)
                        },
                        'cooperation': {
                            'subordinate': safe_float(leader_feedback[99] if len(leader_feedback) > 99 else None),
                            'peer': safe_float(leader_feedback[100] if len(leader_feedback) > 100 else None)
                        },
                        'progress': {
                            'subordinate': safe_float(leader_feedback[102] if len(leader_feedback) > 102 else None),
                            'peer': safe_float(leader_feedback[103] if len(leader_feedback) > 103 else None)
                        },
                        'create': {
                            'subordinate': safe_float(leader_feedback[105] if len(leader_feedback) > 105 else None),
                            'peer': safe_float(leader_feedback[106] if len(leader_feedback) > 106 else None)
                        }
                    }
                }
            }
    
    # 查找-1层管理者
    minus_one_managers_col = diag_row[9] if len(diag_row) > 9 else ""
    if minus_one_managers_col:
        managers = []
        manager_entries = minus_one_managers_col.split('；')
        
        for entry in manager_entries:
            entry = entry.strip()
            if not entry:
                continue
            
            if '：' in entry:
                name_part, light_part = entry.split('：', 1)
                name_part = name_part.strip()
                light_result = light_part.strip()
                
                matches = feedback_reader.find_rows(0, name_part)
                if matches:
                    mgr_feedback = matches[0][1]
                    
                    # 保存原始数值用于排序和变色判断
                    total_rank_raw = mgr_feedback[2] if len(mgr_feedback) > 2 else ''
                    peer_rank_raw = mgr_feedback[3] if len(mgr_feedback) > 3 else ''
                    look_clear_raw = mgr_feedback[4] if len(mgr_feedback) > 4 else None
                    subordinate_raw = mgr_feedback[5] if len(mgr_feedback) > 5 else ''
                    peer_coop_raw = mgr_feedback[6] if len(mgr_feedback) > 6 else ''
                    competent_raw = mgr_feedback[7] if len(mgr_feedback) > 7 else ''
                    negative_raw = mgr_feedback[8] if len(mgr_feedback) > 8 else '0'
                    
                    managers.append({
                        'name': name_part,
                        'result': light_result,
                        'total_rank': format_percentage(total_rank_raw, True),
                        'peer_rank': format_percentage(peer_rank_raw, True),
                        'look_clear_score': f"{safe_float(look_clear_raw):.2f}" if safe_float(look_clear_raw) != 0 or (look_clear_raw and look_clear_raw not in ['N/A', 'NA', '']) else 'N/A',
                        'subordinate_follow': format_percentage(subordinate_raw),
                        'peer_cooperate': format_percentage(peer_coop_raw),
                        'competent': format_percentage(competent_raw),
                        'negative_feedback': str(negative_raw),
                        # 保存原始值用于排序和变色
                        '_total_rank_raw': safe_float(total_rank_raw, 999),
                        '_peer_rank_raw': safe_float(peer_rank_raw, 999),
                        '_look_clear_raw': safe_float(look_clear_raw),
                        '_subordinate_raw': safe_float(subordinate_raw),
                        '_peer_coop_raw': safe_float(peer_coop_raw),
                        '_competent_raw': safe_float(competent_raw),
                        '_negative_raw': safe_int(negative_raw)
                    })
        
        if managers:
            # 排序：先按结果（红>黄>绿），再按总分排名（从差到好，即数值从大到小）
            color_order = {'红': 0, '黄': 1, '绿': 2}
            
            def get_result_color(result_str):
                if '红' in result_str:
                    return 0
                elif '黄' in result_str:
                    return 1
                elif '绿' in result_str:
                    return 2
                else:
                    return 3
            
            managers.sort(key=lambda m: (get_result_color(m['result']), -m['_total_rank_raw']))
            
            # 为每个管理者添加标签（BP点赞/BP提醒关注/兼岗）
            for mgr in managers:
                tags = []
                if bp_loader:
                    bp_tags = bp_loader.get_tags(org_full_path, mgr['name'])
                    tags.extend(bp_tags)
                if jiangan_loader and jiangan_loader.is_jiangan(org_full_path, mgr['name']):
                    tags.append('兼岗')
                mgr['tags'] = tags
            
            report['minus_one_managers'] = managers
    
    # 生成组织架构图数据
    org_chart = build_org_chart_data(org_full_path, leader_name_override=leader_name)
    if org_chart:
        report['org_chart'] = org_chart
    
    # 生成异动数据
    if yd_loader is not None:
        yd_data = yd_loader.build_yidong_data(org_full_path)
        if yd_data:
            report['yidong'] = yd_data
        core_data = yd_loader.build_core_talent_data(org_full_path)
        if core_data:
            report['core_talent'] = core_data
    
    # 生成敬满数据
    if jm_loader is not None:
        jm_data = build_jingman_data(org_full_path, jm_loader)
        if jm_data:
            report['jingman'] = jm_data
    
    # 生成敬满逐题详情数据
    if jm_detail_loader is not None:
        detail_row = jm_detail_loader.find_dept(org_full_path)
        if detail_row is not None:
            # 如果还没有jingman基础数据，先建一个
            if 'jingman' not in report:
                report['jingman'] = {'available': True, 'core_table': [], 'risk_table': []}
            
            # 用全量敬满数据重新计算核心维度表（覆盖之前的）
            detail_core_table = jm_detail_loader.build_core_table(detail_row)
            if detail_core_table:
                report['jingman']['core_table'] = detail_core_table
            
            # 用全量敬满数据重新计算风险区间表（覆盖之前的，包含GB/其他拆分）
            detail_risk_table = jm_detail_loader.build_risk_table(detail_row)
            if detail_risk_table:
                report['jingman']['risk_table'] = detail_risk_table
            
            # 柱状图数据
            chart_data = jm_detail_loader.build_chart_data(detail_row)
            report['jingman']['chart_data'] = chart_data
            
            n_front = sum(1 for it in chart_data if it['fav_bg'] >= 50)
            n_total = len(chart_data)
            n_sat = sum(1 for it in chart_data if it['family'] == '满意度')
            n_eng = sum(1 for it in chart_data if it['family'] == '敬业度')
            
            # BG内最小击败率（用于"倒数第一"红线）
            bg_name = str(detail_row.get("所属bg", ""))
            bg_min = jm_detail_loader.get_bg_min_fav_bg(bg_name)
            
            report['jingman']['chart_summary'] = {
                'n_front': n_front,
                'n_total': n_total,
                'n_sat': n_sat,
                'n_eng': n_eng,
                'bg_min': bg_min,
            }
            
            # 分析文字
            analysis = jm_detail_loader.build_analysis_text(detail_row, chart_data, detail_risk_table)
            report['jingman']['analysis'] = analysis
            
            # 末10%题目明细
            bottom_table = jm_detail_loader.build_bottom_table(detail_row)
            report['jingman']['bottom_table'] = bottom_table
            
            # 细分项大表
            subdiv_table = jm_detail_loader.build_subdivision_table(detail_row)
            report['jingman']['subdiv_table'] = subdiv_table
            
            report['jingman']['has_detail'] = True
    
    # 2.2.3 离职原因分析
    resignation_data = build_resignation_data(org_full_path, resignation_analyzer)
    if resignation_data:
        report['resignation'] = resignation_data
    
    # 2.1.3 全面反馈开放题总结
    if open_fb_loader:
        cadre_text, org_text = open_fb_loader.get_data(org_full_path)
        if cadre_text or org_text:
            # 清洗 AI 编造的中文名：收集数据源中的标准写法，替换不匹配的
            _known_names = {}  # eng_lower -> 'name(cn)' 标准写法
            for mgr in report.get('minus_one_managers', []):
                nm = re.match(r'(\w+)\(([^)]+)\)', mgr.get('name', ''))
                if nm:
                    _known_names[nm.group(1).lower()] = nm.group(0)
            # 负责人
            _leader_m = re.match(r'(\w+)\(([^)]+)\)', report.get('leader_name', ''))
            if _leader_m:
                _known_names[_leader_m.group(1).lower()] = _leader_m.group(0)
            
            def _sanitize_ai_names(text):
                """将 AI 编造的 engname（假中文名）/ engname(假中文名) 替换为纯英文名"""
                if not text or not _known_names:
                    return text
                def _fix_fullwidth(m):
                    en = m.group(1)
                    cn = m.group(2)
                    std = _known_names.get(en.lower())
                    if std:
                        std_cn = re.match(r'\w+\(([^)]+)\)', std)
                        if std_cn and std_cn.group(1) != cn:
                            return en  # AI 编的名字，只保留英文名
                    return m.group(0)
                def _fix_halfwidth(m):
                    en = m.group(1)
                    cn = m.group(2)
                    std = _known_names.get(en.lower())
                    if std:
                        std_cn = re.match(r'\w+\(([^)]+)\)', std)
                        if std_cn and std_cn.group(1) != cn:
                            return en  # AI 编的名字，只保留英文名
                    return m.group(0)
                # 全角括号
                text = re.sub(r'(\w+)（([^）]+)）', _fix_fullwidth, text)
                # 半角括号（只处理短内容，避免误伤代码/URL）
                text = re.sub(r'(\w+)\(([^)]{1,15})\)', _fix_halfwidth, text)
                return text
            
            if cadre_text:
                cadre_text = _sanitize_ai_names(cadre_text)
            if org_text:
                org_text = _sanitize_ai_names(org_text)
            
            report['open_feedback'] = {
                'cadre_html': OpenFeedbackLoader.cadre_to_html(cadre_text) if cadre_text else '',
            }
    
    # 2.3.3 敬满开放题分析
    if jm_open_loader:
        jm_open_html = jm_open_loader.get_html(org_full_path)
        if jm_open_html:
            report['jingman_open_html'] = jm_open_html
    
    # 2.3.3 词云数据
    if wc_loader:
        wc_json, wc_dept = wc_loader.get_wordcloud_data(org_full_path)
        if wc_json:
            report['wordcloud_json'] = wc_json
            report['wordcloud_dept'] = wc_dept
    
    return report


# ============================================================
# 敬满开放题分析 — 加载 *_敬满开放题分析.xlsx
# ============================================================

class JingmanOpenLoader:
    """加载【敬满】目录下的 *_敬满开放题分析.xlsx 文件，
    解析 E 列中 AI 生成的 markdown 分析结果为结构化数据，
    渲染为 demo_jingman_open.html 同款 HTML。"""

    def __init__(self, jingman_dir='【敬满】'):
        self._jingman_dir = jingman_dir
        self._data = {}          # dept_name -> E 列原始文本
        self._scanned_bgs = set()

    def _load_bg(self, bg_short):
        """按需加载某个BG的敬满开放题分析文件，优先特定文件，找不到则回退通用模板"""
        if bg_short in self._scanned_bgs:
            return
        self._scanned_bgs.add(bg_short)

        pattern = os.path.join(self._jingman_dir, f'{bg_short}_敬满开放题分析.xlsx')
        import glob
        files = glob.glob(pattern)
        if not files:
            # 也尝试小写匹配
            for f in os.listdir(self._jingman_dir):
                if f.lower().startswith(bg_short.lower()) and '敬满开放题分析' in f and f.endswith('.xlsx') and not f.startswith('~$'):
                    files.append(os.path.join(self._jingman_dir, f))
        # fallback to generic template
        if not files:
            fallback = os.path.join(self._jingman_dir, '模板_敬满开放题分析.xlsx')
            if os.path.exists(fallback):
                files = [fallback]
        
        for filepath in files:
            try:
                self._load_file(filepath)
            except Exception as e:
                print(f"  ⚠ 敬满开放题分析加载失败 {os.path.basename(filepath)}: {e}")

    def _load_file(self, filepath):
        """加载单个分析文件"""
        try:
            from openpyxl import load_workbook
        except ImportError:
            import subprocess, sys
            subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'openpyxl'])
            from openpyxl import load_workbook

        wb = load_workbook(filepath, data_only=True)
        ws = wb[wb.sheetnames[0]]
        count = 0
        for row in ws.iter_rows(min_row=2, max_col=5, values_only=False):
            dept = row[0].value
            e_val = row[4].value if len(row) > 4 and row[4].value else None
            if dept and e_val:
                dept_name = str(dept).strip()
                self._data[dept_name] = str(e_val)
                count += 1
        wb.close()
        print(f"  ✓ 敬满开放题分析已加载: {os.path.basename(filepath)} ({count} 个部门)")

    def _match_dept(self, org_full_path):
        """根据 org_full_path 匹配部门名"""
        import re
        # org_full_path 格式如 "BG示例/业务线/产品部"
        # 统一分隔符：半角/全角斜杠、全角破折号 都视为路径分隔符
        norm = org_full_path
        for ch in ['\\', '/', '／', '－', '—', '｜', '|']:
            norm = norm.replace(ch, '/')
        parts = [p.strip() for p in norm.split('/') if p.strip()]
        if not parts:
            return None
        
        dept_name = parts[-1]
        
        # 精确匹配
        if dept_name in self._data:
            return dept_name
        
        # 尝试去掉"线"后缀匹配，如 "业务线_产品部" -> "产品部"
        for key in self._data:
            if dept_name.endswith(key) or key.endswith(dept_name):
                return key
            # 也尝试下划线拼接匹配
            if len(parts) >= 2:
                combo = parts[-2] + '_' + parts[-1]
                if combo == key or key in combo or combo in key:
                    return key
        
        # 模糊匹配：部门名包含在key中或反过来
        for key in self._data:
            if dept_name in key or key in dept_name:
                return key
        
        return None

    def get_html(self, org_full_path):
        """获取指定部门的敬满开放题 HTML 渲染结果"""
        # 先按需加载BG数据
        bg_short = ''
        m = re.match(r'([A-Za-z0-9]+)', org_full_path)
        if m:
            bg_short = m.group(1).upper()
            if bg_short == 'OVERSEAS':
                bg_short = 'OFS'
        if not bg_short:
            parts = org_full_path.split('/')
            if parts:
                m2 = re.match(r'([A-Za-z0-9]+)', parts[0])
                if m2:
                    bg_short = m2.group(1).upper()
                    if bg_short == 'OVERSEAS':
                        bg_short = 'OFS'
        
        if bg_short:
            self._load_bg(bg_short)
        
        dept_key = self._match_dept(org_full_path)
        if not dept_key:
            return None
        
        raw_text = self._data.get(dept_key, '')
        if not raw_text:
            return None
        
        return self._render_html(raw_text)

    # ── 解析 markdown -> 结构化数据 ──

    @staticmethod
    def _parse_analysis(text):
        """将 E 列 markdown 文本解析为结构化维度列表。
        
        Returns:
            list of dict:
            [
              {
                'name': '工作负荷与工作生活平衡',
                'categories': [
                  {
                    'type': 'concern',   # or 'good'
                    'findings': [
                      {
                        'title': '工作时长过长、作息不健康、加班常态化',
                        'count': 23,
                        'quotes': ['引用1', '引用2', ...]
                      }, ...
                    ]
                  }, ...
                ]
              }, ...
            ]
        """
        lines = text.split('\n')
        dimensions = []
        current_dim = None
        current_cat = None
        current_finding = None
        
        for line in lines:
            stripped = line.strip()
            if not stripped:
                continue
            
            # 检测维度标题
            # 格式1: ### 【维度一】工作负荷与工作生活平衡
            # 格式2: ### 【维度一：管理与组织机制】
            # 格式3: 【维度一】组织文化与管理风格 （无###）
            # 格式4: ### 【维度一】有待关注的方面  （分类合并到维度标题）
            dim_match = re.match(r'^(?:#{1,4}\s*)?【维度[一二三四五六七八九十百]+[：:]?\s*(.*?)】\s*(.*)', stripped)
            if dim_match:
                dim_name_inside = dim_match.group(1).strip()
                dim_name_after = dim_match.group(2).strip()
                dim_name = dim_name_inside if dim_name_inside else dim_name_after
                dim_name = dim_name.strip('：: ')
                # 识别"维度标题里合并了分类"的特殊格式（如 PC游戏平台部）
                merged_cat_type = None
                if '做得好' in dim_name:
                    merged_cat_type = 'good'
                elif '有待关注' in dim_name or '改进' in dim_name or '负向' in dim_name:
                    merged_cat_type = 'concern'
                current_dim = {'name': dim_name, 'categories': []}
                dimensions.append(current_dim)
                if merged_cat_type:
                    # 合并分类情况：自动创建对应 category
                    current_cat = {'type': merged_cat_type, 'findings': []}
                    current_dim['categories'].append(current_cat)
                else:
                    current_cat = None
                current_finding = None
                continue
            
            # 检测分类标题
            # 格式: ●【有待关注的方面】 或 #### ●【做得好的方面】
            cat_match = re.match(r'^(?:#{1,4}\s*)?●\s*【(.*?)】', stripped)
            if cat_match:
                cat_label = cat_match.group(1).strip()
                cat_type = 'good' if '做得好' in cat_label else 'concern'
                if current_dim is None:
                    # 没有维度标题，创建默认维度
                    current_dim = {'name': '综合反馈', 'categories': []}
                    dimensions.append(current_dim)
                current_cat = {'type': cat_type, 'findings': []}
                current_dim['categories'].append(current_cat)
                current_finding = None
                continue
            
            # 检测关键发现
            # 格式1: 1. **工作时长过长...**，填答人数为 **23人**，
            # 格式2: 1、管理透明度与公平性存在质疑，填答人数为4人，
            # 格式3: 1、xxx，提及人数为1人，（青腾等部门使用"提及"）
            finding_match = re.match(
                r'^(\d+)[\.、]\s*\*{0,2}(.*?)\*{0,2}[，,]\s*(?:填答|填写|提及)人数为\s*\*{0,2}(\d+)\s*人?\s*\*{0,2}',
                stripped
            )
            # 格式4: ● **工作生活平衡**，填答人数为 **28人** （PC游戏平台部等使用 ● 替代数字编号）
            if not finding_match:
                finding_match_bullet = re.match(
                    r'^●\s*\*{0,2}(.*?)\*{0,2}[，,]\s*(?:填答|填写|提及)人数为\s*\*{0,2}(\d+)\s*人?\s*\*{0,2}',
                    stripped
                )
                if finding_match_bullet:
                    # 统一封装为 finding_match 的接口（复用后续代码）
                    class _M:
                        def __init__(self, title, count, end):
                            self._title, self._count, self._end = title, count, end
                        def group(self, i):
                            return {2: self._title, 3: str(self._count)}.get(i, '')
                        def end(self):
                            return self._end
                    finding_match = _M(
                        finding_match_bullet.group(1),
                        int(finding_match_bullet.group(2)),
                        finding_match_bullet.end()
                    )
            # 格式5: 1、xxx，负向反馈共 **6 人**提及  （编程系统中心等使用"N人提及/反馈"倒装）
            if not finding_match:
                finding_match_inv = re.match(
                    r'^(\d+)[\.、]\s*(.+?)[，,].*?\*{0,2}\s*(\d+)\s*\*{0,2}\s*人\s*\*{0,2}\s*(?:提及|反馈|提出)',
                    stripped
                )
                if finding_match_inv:
                    class _M2:
                        def __init__(self, title, count, end):
                            self._title, self._count, self._end = title, count, end
                        def group(self, i):
                            return {2: self._title, 3: str(self._count)}.get(i, '')
                        def end(self):
                            return self._end
                    finding_match = _M2(
                        finding_match_inv.group(2),
                        int(finding_match_inv.group(3)),
                        finding_match_inv.end()
                    )
            if finding_match and current_cat is not None:
                title = finding_match.group(2).strip().strip('*').strip()
                count = int(finding_match.group(3))
                current_finding = {'title': title, 'count': count, 'quotes': []}
                current_cat['findings'].append(current_finding)
                # 同行内联引用：finding 与 quote 在同一行的情况
                # 例: 1、xxx，填答人数为1人，"典型原文引用"："引用内容"。
                tail = stripped[finding_match.end():]
                # 去掉 "典型原文引用" 等提示词，只保留真实引用
                tail = re.sub(r'[\u201c\u201d"]*\s*典型原文引用\s*[\u201c\u201d"]*\s*[：:]\s*', '', tail)
                # 提取所有被成对引号包围的内容
                for q_match in re.finditer(
                    r'[\u201c"\u2018](.+?)[\u201d"\u2019]',
                    tail,
                ):
                    q_text = q_match.group(1).strip().rstrip('\u201c\u201d"\u2018\u2019 。.')
                    if q_text:
                        current_finding['quotes'].append(q_text)
                continue
            
            # 检测引用行 — 支持多种格式：
            # 格式1: - "引用内容"
            # 格式2: "引用内容"  （直接以引号开头）
            # 格式3: > "引用内容" （blockquote）
            # 排除 "典型原文引用" 标记行
            if re.match(r'^[\u201c\u201d"]*典型原文引用[\u201c\u201d"]*', stripped):
                continue
            quote_match = re.match(r'^\s*(?:[-\u2013\u2014>]\s*)?[\u201c\u201d"\u2018\u2019](.+?)[\u201c\u201d"\u2018\u2019]?\s*$', stripped)
            if not quote_match:
                quote_match = re.match(r'^\s*(?:[-\u2013\u2014>]\s*)?[\u201c\u201d"\u2018\u2019](.+)', stripped)
            if quote_match and current_finding is not None:
                quote_text = quote_match.group(1).strip().rstrip('\u201c\u201d"\u2018\u2019 ')
                if quote_text:
                    current_finding['quotes'].append(quote_text)
                continue
        
        return dimensions

    # ── 渲染 HTML ──

    @classmethod
    def _render_html(cls, raw_text):
        """将原始 markdown 分析文本渲染为 HTML"""
        dimensions = cls._parse_analysis(raw_text)
        if not dimensions:
            return None
        
        # 对每个维度分类：only_concern / mixed / only_good
        # 注意：分类标题存在但内容为空时不算有效分类
        # - 【做得好】内容为空 → 判为 有待关注（而非中立）
        # - 【有待关注】内容为空 → 判为 做得好（而非中立）
        dim_infos = []
        for dim in dimensions:
            has_concern = any(c['type'] == 'concern' and len(c.get('findings', [])) > 0 for c in dim['categories'])
            has_good = any(c['type'] == 'good' and len(c.get('findings', [])) > 0 for c in dim['categories'])
            
            if has_concern and not has_good:
                dim_type = 'only_concern'
                sort_key = 0
            elif has_concern and has_good:
                dim_type = 'mixed'
                sort_key = 1
            elif has_good and not has_concern:
                dim_type = 'only_good'
                sort_key = 2
            else:
                dim_type = 'only_concern'
                sort_key = 0
            
            dim_infos.append({
                'dim': dim,
                'type': dim_type,
                'sort_key': sort_key,
            })
        
        # 按三类排序
        dim_infos.sort(key=lambda x: x['sort_key'])
        
        # 过滤：维度下所有关键发现的累计人数 < 2 的维度不呈现
        dim_infos = [info for info in dim_infos
                     if sum(f['count'] for cat in info['dim']['categories'] for f in cat.get('findings', [])) >= 2]
        
        if not dim_infos:
            return None
        
        # 类型标签映射
        type_labels = {
            'only_concern': ('待关注', 'only-concern'),
            'mixed': ('中立', 'mixed'),
            'only_good': ('做得好', 'only-good'),
        }
        card_classes = {
            'only_concern': 'card-concern',
            'mixed': 'card-mixed',
            'only_good': 'card-good',
        }
        
        html_parts = []
        
        # 总览表
        html_parts.append('<div class="jm-open-overview">')
        html_parts.append('<div class="jm-open-overview-header">总览</div>')
        html_parts.append('<table class="jm-open-overview-table">')
        html_parts.append('<thead><tr><th style="width:5%">#</th><th style="width:35%">维度</th><th style="width:20%">类型</th><th style="width:40%">关键发现概要</th></tr></thead>')
        html_parts.append('<tbody>')
        
        for idx, info in enumerate(dim_infos, 1):
            dim = info['dim']
            label_text, label_class = type_labels[info['type']]
            # 关键发现概要：取每个 category 前2个 finding 的 title+count
            summaries = []
            for cat in dim['categories']:
                for f in cat['findings'][:2]:
                    summaries.append(f"{f['title']}({f['count']}人)")
            summary_text = '、'.join(summaries[:3])
            
            html_parts.append(f'<tr>')
            html_parts.append(f'  <td>{idx}</td>')
            html_parts.append(f'  <td class="jm-open-dim-name">{dim["name"]}</td>')
            html_parts.append(f'  <td><span class="jm-open-type-badge {label_class}">{label_text}</span></td>')
            html_parts.append(f'  <td class="jm-open-finding-summary">{summary_text}</td>')
            html_parts.append(f'</tr>')
        
        html_parts.append('</tbody></table></div>')
        
        # 维度卡片
        cn_nums = ['一', '二', '三', '四', '五', '六', '七', '八', '九', '十',
                    '十一', '十二', '十三', '十四', '十五']
        
        for idx, info in enumerate(dim_infos):
            dim = info['dim']
            card_cls = card_classes[info['type']]
            num_str = cn_nums[idx] if idx < len(cn_nums) else str(idx + 1)
            
            html_parts.append(f'<div class="jm-open-card {card_cls}">')
            html_parts.append(f'<div class="jm-open-card-header">【维度{num_str}】{dim["name"]}</div>')
            html_parts.append('<div class="jm-open-card-body">')
            
            rendered_cat_count = 0
            for ci, cat in enumerate(dim['categories']):
                # 跳过没有实际内容的空分类
                if not cat.get('findings'):
                    continue
                if rendered_cat_count > 0:
                    html_parts.append('<hr class="jm-open-divider">')
                rendered_cat_count += 1
                
                cat_label = '有待关注' if cat['type'] == 'concern' else '做得好'
                cat_cls = 'concern' if cat['type'] == 'concern' else 'good'
                
                html_parts.append(f'<div class="jm-open-cat-row"><span class="jm-open-cat-tag {cat_cls}">{cat_label}</span></div>')
                html_parts.append('<table class="jm-open-finding-table">')
                html_parts.append('<thead><tr><th>#</th><th>关键发现</th><th>典型原文引用</th></tr></thead>')
                html_parts.append('<tbody>')
                
                for fi, finding in enumerate(cat['findings'], 1):
                    count = finding['count']
                    # 颜色等级：做得好统一绿色；其余保持原有分级颜色
                    if cat['type'] == 'good':
                        count_cls = 'good'
                    elif count >= 10:
                        count_cls = 'high'
                    elif count >= 5:
                        count_cls = 'medium'
                    else:
                        count_cls = 'low'
                    
                    quotes_html = f'<span class="jm-open-count {count_cls}">{count}人</span>，如：'
                    for q in finding['quotes'][:3]:
                        q_escaped = q.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
                        quotes_html += f'<span class="jm-open-quote-line">&ldquo;{q_escaped}&rdquo;</span>'
                    
                    title_escaped = finding['title'].replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
                    html_parts.append(f'<tr>')
                    html_parts.append(f'  <td>{fi}</td>')
                    html_parts.append(f'  <td><span class="jm-open-finding-title">{title_escaped}</span></td>')
                    html_parts.append(f'  <td><div class="jm-open-quotes">{quotes_html}</div></td>')
                    html_parts.append(f'</tr>')
                
                html_parts.append('</tbody></table>')
            
            html_parts.append('</div></div>')
        
        return '\n'.join(html_parts)


# ============================================================
# 词云数据加载器 — 加载 WXG_敬满开放题关键词分析.xlsx（E列带词性标注）
# ============================================================

# 词性映射：xlsx中的词性标注 → JS词云的sentiment字段
_WC_SENTIMENT_MAP = {
    '正向': 'positive',
    '负向': 'negative',
    '中性': 'neutral',
    '中立': 'neutral',
}


_WC_FREQ_THRESHOLD = 20  # ≥2的词不足此数时降级取全部词

# 英文停用词 — 词云中过滤掉的无意义虚词/介词/连词
_WC_EN_STOPWORDS = {
    'a', 'an', 'the', 'to', 'and', 'or', 'of', 'in', 'on', 'at',
    'is', 'it', 'be', 'for', 'as', 'by', 'no', 'not', 'n', 'but',
    'with', 'that', 'this', 'are', 'was', 'were', 'has', 'have', 'had',
    'if', 'so', 'do', 'up', 'we', 'my', 'me', 'he', 'she', 'i', 'its',
    'can', 'will', 'just', 'should', 'would', 'could', 'may', 'might',
    'am', 'been', 'being', 'did', 'does', 'doing', 'each', 'few',
    'from', 'get', 'got', 'her', 'him', 'his', 'how', 'into',
    'more', 'most', 'much', 'must', 'nor', 'our', 'out', 'own',
    'same', 'some', 'such', 'than', 'them', 'then', 'too', 'very',
    'what', 'when', 'who', 'whom', 'why', 'you', 'your',
    'c',
}

def _wc_is_stopword(word):
    """判断是否为词云应过滤的停用词"""
    w = word.strip().lower()
    if w in _WC_EN_STOPWORDS:
        return True
    if re.fullmatch(r'\d+\.?', w):
        return True
    if re.fullmatch(r'[^a-zA-Z\u4e00-\u9fff\u3040-\u309f\u30a0-\u30ff]+', w):
        return True
    return False


class WordCloudLoader:
    """加载【敬满】/*_敬满开放题关键词分析.xlsx 的 B 列，
    解析 '关键词 词性 频次' 格式（；分隔），
    按部门提供词云数据（word list + freq + sentiment）。
    渲染层词频降级：≥2的词不足20个时取全部词。"""

    def __init__(self, jingman_dir='【敬满】'):
        self._jingman_dir = jingman_dir
        # dept_name -> list of (word, sentiment, freq) tuples
        self._dept_data = {}
        self._loaded_bgs = set()

    def _parse_keyword_text(self, text):
        """从 B 列文本中解析 '关键词 词性 频次' 数据（；分隔）。
        完整输出所有词（≥1），词频过滤由 get_wordcloud_data 按阈值降级处理。"""
        if not text:
            return []
        
        KW_RE = re.compile(r'^(.+?)\s+(正向|负向|中性|中立)\s+(\d+)$')
        entries = []
        for item in re.split(r'[；;]', text):
            item = item.strip()
            if not item:
                continue
            m = KW_RE.match(item)
            if m:
                word = m.group(1).strip()
                sentiment_cn = m.group(2)
                freq = int(m.group(3))
                sentiment = _WC_SENTIMENT_MAP.get(sentiment_cn, 'neutral')
                if word and freq >= 1:
                    if not _wc_is_stopword(word):
                        entries.append((word, sentiment, freq))
        return entries

    def _load_bg(self, bg_short):
        """按需加载某个 BG 的词云关键词分析文件"""
        if bg_short in self._loaded_bgs:
            return
        self._loaded_bgs.add(bg_short)
        
        import glob as _glob
        # 搜索 BG_敬满开放题关键词分析.xlsx，优先特定文件，找不到则回退通用模板
        pattern = os.path.join(self._jingman_dir, f'{bg_short}_敬满开放题关键词分析.xlsx')
        files = _glob.glob(pattern)
        if not files:
            # 大小写不敏感搜索
            for f in os.listdir(self._jingman_dir):
                if f.lower().startswith(bg_short.lower()) and '关键词分析' in f and f.endswith('.xlsx') and not f.startswith('~$'):
                    files.append(os.path.join(self._jingman_dir, f))
        # fallback to generic template
        if not files:
            fallback = os.path.join(self._jingman_dir, '模板_敬满开放题关键词分析.xlsx')
            if os.path.exists(fallback):
                files = [fallback]
        
        for filepath in files:
            try:
                from openpyxl import load_workbook
                wb = load_workbook(filepath, data_only=True)
                ws = wb[wb.sheetnames[0]]
                count = 0
                for row in ws.iter_rows(min_row=2, values_only=True):
                    dept = str(row[0]).strip() if row[0] else ''
                    # 优先读 B 列（关键词列），兼容旧格式 E 列
                    b_val = str(row[1]).strip() if len(row) > 1 and row[1] else ''
                    if b_val in ('无关键词', '无高频关键词', '无', 'nan', 'None', ''):
                        b_val = ''
                    if not dept or not b_val:
                        continue
                    parsed = self._parse_keyword_text(b_val)
                    if parsed:
                        self._dept_data[dept] = parsed
                        count += 1
                wb.close()
                print(f"  ✓ 词云关键词分析已加载: {count} 个部门")
            except Exception as e:
                print(f"  ⚠ 词云关键词分析加载失败: {e}")

    def _match_dept(self, org_full_path):
        """复用 JingmanOpenLoader 同款匹配逻辑"""
        import re
        norm = org_full_path
        for ch in ['\\', '/', '／', '－', '—', '｜', '|']:
            norm = norm.replace(ch, '/')
        parts = [p.strip() for p in norm.split('/') if p.strip()]
        if not parts:
            return None
        dept_name = parts[-1]
        if dept_name in self._dept_data:
            return dept_name
        for key in self._dept_data:
            if dept_name.endswith(key) or key.endswith(dept_name):
                return key
            if len(parts) >= 2:
                combo = parts[-2] + '_' + parts[-1]
                if combo == key or key in combo or combo in key:
                    return key
        for key in self._dept_data:
            if dept_name in key or key in dept_name:
                return key
        return None

    def get_wordcloud_data(self, org_full_path, top_n=80):
        """返回 (words_json_str, dept_display_name) 或 (None, None)
        
        词频阈值降级逻辑：
        - 优先取频次≥2的关键词
        - 若≥2的词不足20个，降级取全部词（≥1）
        """
        # 按需加载对应 BG 的词云数据
        bg_m = re.match(r'([A-Za-z0-9]+)', org_full_path)
        if bg_m:
            _bg = bg_m.group(1).upper()
            if _bg == 'OVERSEAS':
                _bg = 'OFS'
            self._load_bg(_bg)
        dept_key = self._match_dept(org_full_path)
        if not dept_key:
            return None, None
        entries = self._dept_data.get(dept_key, [])
        if not entries:
            return None, None
        # 词频阈值降级：≥2的词不足20个时取全部词
        high_freq_entries = [(w, s, f) for w, s, f in entries if f >= 2]
        if len(high_freq_entries) >= _WC_FREQ_THRESHOLD:
            use_entries = high_freq_entries
        else:
            use_entries = entries
        actual_top_n = min(top_n, len(use_entries))
        top_entries = use_entries[:actual_top_n]
        words_data = []
        for word, sentiment, freq in top_entries:
            words_data.append({'text': word, 'freq': freq, 'sentiment': sentiment})
        return json.dumps(words_data, ensure_ascii=False), dept_key


def build_org_chart_data(org_path: str, leader_name_override: str = "") -> Dict:
    """构建组织架构图数据
    
    Args:
        org_path: 组织全路径，例如 "示例BG/示例业务线/示例部门"
        leader_name_override: 诊断筛查表中的负责人名（优先级高于组织机构信息表）
    
    Returns:
        组织架构树数据字典
    """
    try:
        # 读取组织架构信息（动态查找列号）
        org_reader = ExcelReader('【组织架构信息】/组织机构信息.xlsx')
        col_map = {}
        needed = ['组织全路径', '组织负责人', '组织负责人上级', '组织架构层级',
                  '组织名称', '员工数', '组织类型']
        for i, h in enumerate(org_reader.headers):
            hs = str(h).strip()
            if hs in needed:
                col_map[hs] = i
        
        # 校验必需列
        missing = [k for k in ['组织全路径', '组织负责人', '组织名称', '员工数', '组织类型'] if k not in col_map]
        if missing:
            print(f"⚠️  组织机构信息表缺少列: {missing}")
            return None
        
        ci_path = col_map['组织全路径']
        ci_leader = col_map['组织负责人']
        ci_parent = col_map.get('组织负责人上级', -1)
        ci_level = col_map.get('组织架构层级', -1)
        ci_name = col_map['组织名称']
        ci_count = col_map['员工数']
        ci_type = col_map['组织类型']
        
        # 读取全面反馈数据（用于组织架构图亮灯）—— 只按英文名匹配
        feedback_reader = ExcelReader('【全面反馈】/全面反馈25H2v2.xlsx')
        leader_to_light = {}  # eng_name_lower -> color
        for row in feedback_reader.rows[1:]:
            if len(row) > 1:
                leader_raw = str(row[0]).strip() if row[0] else ""
                light_result = str(row[1]).strip() if row[1] else ""
                if not leader_raw:
                    continue
                # 提取英文名
                m = re.match(r'(\w+)\(', leader_raw)
                eng = m.group(1).lower() if m else leader_raw.lower().strip()
                
                if "红" in light_result:
                    color = "red"
                elif "黄" in light_result:
                    color = "yellow"
                elif "绿" in light_result:
                    color = "green"
                else:
                    color = "gray"
                
                # 如果同一人已有更严重的颜色，保留更严重的
                severity = {'red': 3, 'yellow': 2, 'green': 1, 'gray': 0}
                if eng not in leader_to_light or severity.get(color, 0) > severity.get(leader_to_light[eng], 0):
                    leader_to_light[eng] = color
        
        def _get_light(leader_str):
            """按英文名查亮灯颜色"""
            m = re.match(r'(\w+)\(', str(leader_str).strip())
            eng = m.group(1).lower() if m else str(leader_str).lower().strip()
            return leader_to_light.get(eng, "gray")
        
        # 收集目标组织及子组织（仅实体组织）
        target_org_nodes = []
        for row in org_reader.rows[1:]:
            row_org_path = str(row[ci_path]).strip() if ci_path < len(row) and row[ci_path] else ""
            org_type = str(row[ci_type]).strip() if ci_type < len(row) and row[ci_type] else ""
            
            if not row_org_path.startswith(org_path):
                continue
            if org_type != '实体组织':
                continue
            
            leader = str(row[ci_leader]).strip() if ci_leader < len(row) and row[ci_leader] else ""
            parent_leader = str(row[ci_parent]).strip() if ci_parent >= 0 and ci_parent < len(row) and row[ci_parent] else ""
            level = str(row[ci_level]).strip() if ci_level >= 0 and ci_level < len(row) and row[ci_level] else ""
            org_name = str(row[ci_name]).strip() if ci_name < len(row) and row[ci_name] else ""
            emp_count = row[ci_count] if ci_count < len(row) and row[ci_count] else ""
            if isinstance(emp_count, float):
                emp_count = int(emp_count)
            
            target_org_nodes.append({
                'org_path': row_org_path,
                'org_name': org_name,
                'leader': leader,
                'parent_leader': parent_leader,
                'level': level,
                'employee_count': emp_count,
                'light_color': _get_light(leader),
            })
        
        if not target_org_nodes:
            return None
        
        # 找根节点
        root_node = None
        for node in target_org_nodes:
            if node['org_path'] == org_path:
                root_node = node
                break
        if not root_node:
            return None
        
        # 根节点负责人优先用诊断筛查表
        if leader_name_override:
            root_node['leader'] = leader_name_override
            root_node['light_color'] = _get_light(leader_name_override)
        
        # 构建子树（纯路径层级判断，不依赖 parent_leader）
        def build_tree(parent_node):
            children = []
            parent_path = parent_node['org_path']
            
            for node in target_org_nodes:
                if node['org_path'] != parent_path and node['org_path'].startswith(parent_path + '/'):
                    remaining = node['org_path'][len(parent_path)+1:]
                    if '/' not in remaining:
                        child = node.copy()
                        child['children'] = build_tree(node)
                        children.append(child)
            return children
        
        root_with_children = root_node.copy()
        direct_children = build_tree(root_node)
        
        # 虚拟分管节点（方案B）
        # 对根节点的直接子组织，按 parent_leader 分组
        root_leader_str = str(root_node['leader']).strip()
        root_eng_m = re.match(r'(\w+)\(', root_leader_str)
        root_eng = root_eng_m.group(1).lower() if root_eng_m else root_leader_str.lower()
        
        # 分组：上级==根负责人的直接挂根，否则按上级分组
        direct_group = []   # 直接挂根的
        delegate_groups = {}  # parent_leader_str -> [children]
        
        for child in direct_children:
            pl = str(child.get('parent_leader', '')).strip()
            pl_m = re.match(r'(\w+)\(', pl)
            pl_eng = pl_m.group(1).lower() if pl_m else pl.lower()
            
            # 如果上级是根负责人、或上级为空/无效 -> 直接挂根
            if pl_eng == root_eng or not pl or pl in ('#N/A', 'N/A', 'None', ''):
                direct_group.append(child)
            else:
                delegate_groups.setdefault(pl, []).append(child)
        
        # 为每个分管人创建虚拟节点
        final_children = list(direct_group)
        for delegate_leader, sub_children in delegate_groups.items():
            virtual_node = {
                'org_path': '',
                'org_name': '',  # 虚拟节点不显示组织名
                'leader': delegate_leader,
                'parent_leader': root_leader_str,
                'level': '分管',
                'employee_count': '',
                'light_color': _get_light(delegate_leader),
                'is_virtual': True,
                'children': sub_children,
            }
            final_children.append(virtual_node)
        
        root_with_children['children'] = final_children
        return root_with_children
        
    except Exception as e:
        print(f"⚠️  组织架构图数据构建失败: {e}")
        import traceback
        traceback.print_exc()
        return None


# ============================================================
# 异动数据处理
# ============================================================

class YidongDataLoader:
    """异动数据加载器：读取2025全年高异动部门级数据.xlsx 和 部门&中心结论导出-OTD.xlsx 和 A2026核心人才数据"""
    
    def __init__(self, rate_file, otd_file, core_file=None):
        """
        Args:
            rate_file: 2025全年高异动部门级数据.xlsx 路径
            otd_file: 部门&中心结论导出-OTD.xlsx 路径
            core_file: A2026011210130001_修复版v0.4_交付.xlsx 路径（核心人才数据）
        """
        self.rate_file = rate_file
        self.otd_file = otd_file
        self.core_file = core_file
        
        # 加载异动率数据
        self.rate_reader = ExcelReader(rate_file)
        self._rate_index = {}  # 部门级组织 -> row
        for row in self.rate_reader.rows[1:]:
            dept_name = str(row[1]).strip() if len(row) > 1 and row[1] else ""
            if dept_name:
                self._rate_index[dept_name] = row
        
        # 加载OTD数据（部门sheet）
        self.otd_reader = ExcelReader(otd_file, sheet_name='部门')
        self._otd_index = {}  # 部门级组织 -> row
        for row in self.otd_reader.rows[1:]:
            dept_name = str(row[0]).strip() if len(row) > 0 and row[0] else ""
            if dept_name:
                self._otd_index[dept_name] = row
        
        # 加载核心人才数据
        self._core_index = {}       # 组织名称(C2) -> row
        self._core_bg_index = {}    # BG名称 -> BG行 (C7='BG')
        self._core_line_index = {}  # 线名称 -> 线行 (C7='线')
        if core_file:
            self.core_reader = ExcelReader(core_file)
            for row in self.core_reader.rows[1:]:
                org_name = str(row[2]).strip() if len(row) > 2 and row[2] else ""
                level = str(row[7]).strip() if len(row) > 7 and row[7] else ""
                if org_name:
                    self._core_index[org_name] = row
                # 建立BG索引：C7='BG' 时，用C3(BG名称)作为key
                if level == 'BG':
                    bg_name = str(row[3]).strip() if len(row) > 3 and row[3] else ""
                    if bg_name:
                        self._core_bg_index[bg_name] = row
                # 建立线索引：C7='线' 时，用C4(线名称)作为key
                elif level == '线':
                    line_name = str(row[4]).strip() if len(row) > 4 and row[4] else ""
                    if line_name:
                        self._core_line_index[line_name] = row
    
    def _org_path_to_dept_name(self, org_full_path: str) -> str:
        """将组织全路径(示例BG/示例业务线/示例部门)转换为空格分隔格式"""
        if not org_full_path:
            return ""
        return org_full_path.replace('/', ' ')
    
    def _find_dept_row(self, index: dict, org_full_path: str):
        """在指定索引中查找部门行，支持精确匹配和模糊匹配"""
        dept_name = self._org_path_to_dept_name(org_full_path)
        if not dept_name:
            return None
        
        # 精确匹配
        if dept_name in index:
            return index[dept_name]
        
        # 模糊匹配：尝试用最后一个部分匹配
        parts = dept_name.split()
        if len(parts) >= 2:
            last_part = parts[-1]
            for key, row in index.items():
                if key.endswith(last_part):
                    return row
        
        return None
    
    def _has_line(self, dept_name_spaced: str) -> bool:
        """判断部门是否有"线"（空格数>=2表示有线）"""
        return dept_name_spaced.count(' ') >= 2
    
    def _safe_float(self, val, default=0.0):
        """安全转换为float"""
        if val is None:
            return default
        try:
            v = float(val)
            if v != v:  # NaN check
                return default
            return v
        except (ValueError, TypeError):
            return default
    
    def _safe_int(self, val, default=0):
        """安全转换为int"""
        if val is None:
            return default
        try:
            return int(float(val))
        except (ValueError, TypeError):
            return default
    
    def _fmt_pct(self, val):
        """格式化百分比"""
        return f"{val * 100:.1f}%"
    
    def _judge_color(self, dept_rate, bg_rate, line_rate, has_line):
        """
        判断颜色：
        - 绿色：同时低于BG和线（无线则只看BG）
        - 黄色：高于任一个，但同时低于BG和线的2倍
        - 红色：高于线或BG的2倍
        """
        if has_line:
            if dept_rate < bg_rate and dept_rate < line_rate:
                return 'green'
            elif dept_rate >= bg_rate * 2 or dept_rate >= line_rate * 2:
                return 'red'
            else:
                return 'yellow'
        else:
            if dept_rate < bg_rate:
                return 'green'
            elif dept_rate >= bg_rate * 2:
                return 'red'
            else:
                return 'yellow'
    
    def _judge_text(self, dept_rate, bg_rate, line_rate, has_line):
        """生成判断文字"""
        if has_line:
            if dept_rate < bg_rate and dept_rate < line_rate:
                return "本部门整体异动率低于BG和线水平"
            elif dept_rate >= bg_rate * 2 and dept_rate >= line_rate * 2:
                return "本部门整体异动率高于BG和线水平两倍"
            elif dept_rate >= bg_rate * 2:
                return "本部门整体异动率高于BG水平两倍"
            elif dept_rate >= line_rate * 2:
                return "本部门整体异动率高于线水平两倍"
            elif dept_rate >= bg_rate and dept_rate >= line_rate:
                return "本部门整体异动率高于BG和线水平"
            elif dept_rate >= bg_rate:
                return "本部门整体异动率高于BG水平"
            elif dept_rate >= line_rate:
                return "本部门整体异动率高于线水平"
            else:
                return "本部门整体异动率低于BG和线水平"
        else:
            if dept_rate < bg_rate:
                return "本部门整体异动率低于BG水平"
            elif dept_rate >= bg_rate * 2:
                return "本部门整体异动率高于BG水平两倍"
            else:
                return "本部门整体异动率高于BG水平"
    
    def build_yidong_data(self, org_full_path: str) -> Dict:
        """构建2.2.1异动数据"""
        rate_row = self._find_dept_row(self._rate_index, org_full_path)
        otd_row = self._find_dept_row(self._otd_index, org_full_path)
        
        if rate_row is None and otd_row is None:
            return None
        
        dept_name_spaced = self._org_path_to_dept_name(org_full_path)
        has_line = self._has_line(dept_name_spaced)
        
        # 从 rate_row 取异动率
        dept_rate = self._safe_float(rate_row[5]) if rate_row and len(rate_row) > 5 else 0  # C6: 异动率
        bg_rate = self._safe_float(rate_row[9]) if rate_row and len(rate_row) > 9 else 0    # C10: BG异动率
        line_rate = self._safe_float(rate_row[7]) if rate_row and len(rate_row) > 7 else 0  # C8: 部门上级组织异动率
        
        # 判断颜色和文字
        color = self._judge_color(dept_rate, bg_rate, line_rate, has_line)
        judge_text = self._judge_text(dept_rate, bg_rate, line_rate, has_line)
        
        # 参考值文字
        if has_line:
            ref_text = f"参考值：BG整体{self._fmt_pct(bg_rate)}；线整体{self._fmt_pct(line_rate)}"
        else:
            ref_text = f"参考值：BG整体{self._fmt_pct(bg_rate)}"
        
        # 从 otd_row 取离职/活水明细
        # OTD columns: C1:部门级组织, C3:已活水, C4:活水有意向, C5:活水流程中,
        #              C7:离职已离职, C8:离职有意向, C9:离职流程中
        resigned = self._safe_int(otd_row[6]) if otd_row and len(otd_row) > 6 else 0       # C7: 离职已离职
        resign_process = self._safe_int(otd_row[8]) if otd_row and len(otd_row) > 8 else 0  # C9: 离职流程中
        resign_intent = self._safe_int(otd_row[7]) if otd_row and len(otd_row) > 7 else 0   # C8: 离职有意向
        
        transferred = self._safe_int(otd_row[2]) if otd_row and len(otd_row) > 2 else 0     # C3: 已活水
        transfer_process = self._safe_int(otd_row[4]) if otd_row and len(otd_row) > 4 else 0 # C5: 活水流程中
        transfer_intent = self._safe_int(otd_row[3]) if otd_row and len(otd_row) > 3 else 0  # C4: 活水有意向
        
        return {
            'available': True,
            'dept_rate': dept_rate,
            'dept_rate_str': self._fmt_pct(dept_rate),
            'bg_rate': bg_rate,
            'line_rate': line_rate,
            'has_line': has_line,
            'color': color,
            'judge_text': judge_text,
            'ref_text': ref_text,
            # 离职明细
            'resigned': resigned,
            'resign_process': resign_process,
            'resign_intent': resign_intent,
            # 活水明细
            'transferred': transferred,
            'transfer_process': transfer_process,
            'transfer_intent': transfer_intent,
        }
    
    def build_core_talent_data(self, org_full_path: str) -> Dict:
        """构建2.2.2核心人才异动数据"""
        if not self._core_index:
            return None
        
        # 用 org_full_path 直接匹配 C2（组织名称）—— A2026用的是/分隔
        dept_row = self._core_index.get(org_full_path)
        if dept_row is None:
            return None
        if len(dept_row) <= 60:
            return None
        
        # 本部门核心/非核心人才主动异动率
        core_rate = self._safe_float(dept_row[56])    # C56: 核心人才主动异动率
        non_core_rate = self._safe_float(dept_row[60]) # C60: 非核心人才主动异动率
        
        # 获取BG名称和线名称
        bg_name = str(dept_row[3]).strip() if len(dept_row) > 3 and dept_row[3] else ""
        line_name = str(dept_row[4]).strip() if len(dept_row) > 4 and dept_row[4] else ""
        has_line = bool(line_name)
        
        # 获取BG参考值
        bg_core_rate = 0.0
        bg_row = self._core_bg_index.get(bg_name)
        if bg_row:
            bg_core_rate = self._safe_float(bg_row[56])
        
        # 获取线参考值
        line_core_rate = 0.0
        if has_line:
            line_row = self._core_line_index.get(line_name)
            if line_row:
                line_core_rate = self._safe_float(line_row[56])
        
        # 第一步：基础颜色判断（核心人才率 vs BG/线）
        color = self._judge_color(core_rate, bg_core_rate, line_core_rate, has_line)
        
        # 第二步：如果核心 > 非核心，颜色升级
        if core_rate > non_core_rate:
            if color == 'green':
                color = 'yellow'
            elif color == 'yellow':
                color = 'red'
            # red stays red
        
        # 参考值文字
        if has_line:
            ref_text = f"参考值：BG整体{self._fmt_pct(bg_core_rate)}；线整体{self._fmt_pct(line_core_rate)}"
        else:
            ref_text = f"参考值：BG整体{self._fmt_pct(bg_core_rate)}"
        
        return {
            'available': True,
            'core_rate': core_rate,
            'core_rate_str': self._fmt_pct(core_rate),
            'non_core_rate': non_core_rate,
            'non_core_rate_str': self._fmt_pct(non_core_rate),
            'bg_core_rate': bg_core_rate,
            'line_core_rate': line_core_rate,
            'has_line': has_line,
            'color': color,
            'ref_text': ref_text,
            'core_higher_than_non_core': core_rate > non_core_rate,
        }


# ============================================================
# 敬满数据处理
# ============================================================

class JingmanDataLoader:
    """敬满数据加载器：读取敬满总分相关指标.xlsx"""
    
    # 列索引映射
    COL_ORG_NAME = 0            # 组织名称
    COL_ORG_ID = 1              # 组织
    COL_IS_TOP = 2              # 是否top组织
    COL_IS_PROBLEM = 3          # 是否问题组织
    COL_BOTTOM10_PCT = 4        # 后10%短板题占比
    COL_BOTTOM20_PCT = 5        # 后20%短板题占比
    # 敬业度
    COL_ENG_FAV = 9             # 关键指数_敬业度指数_赞成%
    COL_ENG_FAV_2023 = 10       # 关键指数_敬业度指数_赞成%_2023
    COL_ENG_FAV_GAP = 11        # 关键指数_敬业度指数_赞成%_gap
    COL_ENG_FAV_GAP_PCT = 12    # 关键指数_敬业度指数_赞成%_gap%
    COL_ENG_FAV_RANK = 13       # 关键指数_敬业度指数_赞成%_分位值
    COL_ENG_FAV_RANK_2023 = 14  # 关键指数_敬业度指数_赞成%_2023_分位值
    # 满意度
    COL_SAT_FAV = 15            # 关键指数_满意度指数_赞成%
    COL_SAT_FAV_2023 = 16       # 关键指数_满意度指数_赞成%_2023
    COL_SAT_FAV_GAP = 17        # 关键指数_满意度指数_赞成%_gap
    COL_SAT_FAV_GAP_PCT = 18    # 关键指数_满意度指数_赞成%_gap%
    COL_SAT_FAV_RANK = 19       # 关键指数_满意度指数_赞成%_分位值
    COL_SAT_FAV_RANK_2023 = 20  # 关键指数_满意度指数_赞成%_2023_分位值
    
    def __init__(self, filepath):
        self.filepath = filepath
        self.reader = ExcelReader(filepath)
        self._build_index()
    
    def _normalize_jm_path(self, path):
        """标准化敬满组织路径：腾讯集团/CDG/xxx -> CDG/xxx"""
        parts = path.split('/')
        if parts and parts[0] == '腾讯集团':
            parts = parts[1:]
        # 移除 "直管-xxx" 中间层
        parts = [p for p in parts if not p.startswith('直管-')]
        return '/'.join(parts)
    
    def _normalize_diag_path(self, path):
        """标准化诊断组织路径：BG名称/xxx -> BG缩写/xxx"""
        parts = path.split('/')
        if parts:
            first = parts[0]
            match = re.match(r'^([A-Z]+\d*)', first)
            if match:
                parts[0] = match.group(1)
            match2 = re.match(r'^(S\d+)', first)
            if match2:
                parts[0] = match2.group(1)
            if first.startswith('Overseas'):
                parts[0] = 'OFS'
        return '/'.join(parts)
    
    def _get_bg_from_jm(self, path):
        """从敬满路径提取BG"""
        parts = path.replace('腾讯集团/', '').split('/')
        return parts[0] if parts else ''
    
    def _build_index(self):
        """构建多策略查找索引"""
        self.by_norm = {}       # 标准化路径 -> row
        self.by_bg_last = {}    # (bg, 末端组织名) -> row
        self.by_last2 = {}      # 末2段路径 -> row
        
        for row in self.reader.rows[1:]:
            if not row or not row[0]:
                continue
            path = row[0]
            norm = self._normalize_jm_path(path)
            self.by_norm[norm] = row
            
            bg = self._get_bg_from_jm(path)
            parts = norm.split('/')
            if len(parts) >= 2:
                self.by_bg_last[(bg, parts[-1])] = row
                self.by_last2['/'.join(parts[-2:])] = row
            elif len(parts) == 1:
                self.by_bg_last[(bg, parts[0])] = row
    
    def find_dept(self, diag_org_path):
        """根据诊断数据的组织路径查找敬满数据行"""
        norm = self._normalize_diag_path(diag_org_path)
        
        # 策略1：精确标准化匹配
        if norm in self.by_norm:
            return self.by_norm[norm]
        
        # 策略2：BG + 末端组织名
        parts = norm.split('/')
        bg = parts[0] if parts else ''
        last = parts[-1] if parts else ''
        if (bg, last) in self.by_bg_last:
            return self.by_bg_last[(bg, last)]
        
        # 策略3：末2段路径
        if len(parts) >= 2:
            key = '/'.join(parts[-2:])
            if key in self.by_last2:
                return self.by_last2[key]
        
        return None
    
    def get_cell_float(self, row, col_idx, default=None):
        """安全地从行中获取浮点数"""
        if row is None or col_idx >= len(row):
            return default
        val = row[col_idx]
        if not val or val == '' or val == 'NA' or val == 'N/A':
            return default
        try:
            return float(val)
        except (ValueError, TypeError):
            return default


def build_jingman_data(diag_org_path, jm_loader):
    """构建敬满报告数据
    
    Args:
        diag_org_path: 诊断数据中的组织全路径
        jm_loader: JingmanDataLoader 实例
    
    Returns:
        敬满数据字典，如果找不到数据则返回 None
    """
    if jm_loader is None:
        return None
    
    row = jm_loader.find_dept(diag_org_path)
    if row is None:
        return None
    
    gf = jm_loader.get_cell_float
    
    # 核心维度数据
    eng_rank = gf(row, JingmanDataLoader.COL_ENG_FAV_RANK)
    eng_rank_2023 = gf(row, JingmanDataLoader.COL_ENG_FAV_RANK_2023)
    eng_score = gf(row, JingmanDataLoader.COL_ENG_FAV)
    eng_score_2023 = gf(row, JingmanDataLoader.COL_ENG_FAV_2023)
    eng_gap = gf(row, JingmanDataLoader.COL_ENG_FAV_GAP)
    eng_gap_pct = gf(row, JingmanDataLoader.COL_ENG_FAV_GAP_PCT)
    
    sat_rank = gf(row, JingmanDataLoader.COL_SAT_FAV_RANK)
    sat_rank_2023 = gf(row, JingmanDataLoader.COL_SAT_FAV_RANK_2023)
    sat_score = gf(row, JingmanDataLoader.COL_SAT_FAV)
    sat_score_2023 = gf(row, JingmanDataLoader.COL_SAT_FAV_2023)
    sat_gap = gf(row, JingmanDataLoader.COL_SAT_FAV_GAP)
    sat_gap_pct = gf(row, JingmanDataLoader.COL_SAT_FAV_GAP_PCT)
    
    bottom10_pct = gf(row, JingmanDataLoader.COL_BOTTOM10_PCT)
    bottom20_pct = gf(row, JingmanDataLoader.COL_BOTTOM20_PCT)
    
    def fmt_rank(v):
        """分位值转排名百分比：分位值0.85 -> 在BG排名前15%"""
        if v is None: return "-"
        rank_pct = (1 - v) * 100
        return f"{rank_pct:.0f}%"
    
    def fmt_rank_change(curr, prev):
        """排名变化：curr/prev都是分位值"""
        if curr is None: return "-"
        if prev is None: return "/"
        curr_rank = (1 - curr) * 100
        prev_rank = (1 - prev) * 100
        change = curr_rank - prev_rank
        sign = "+" if change >= 0 else ""
        return f"{sign}{change:.0f}%（去年{prev_rank:.0f}%）"
    
    def fmt_score(v):
        if v is None: return "-"
        return f"{v:.1f}"
    
    def fmt_yoy(gap, gap_pct):
        if gap is None and gap_pct is None: return "-"
        parts = []
        if gap is not None:
            sign = "+" if gap >= 0 else ""
            parts.append(f"{sign}{gap:.1f}")
        if gap_pct is not None:
            sign = "+" if gap_pct >= 0 else ""
            parts.append(f"（{sign}{gap_pct*100:.1f}%）")
        return "".join(parts)
    
    def fmt_pct(v):
        if v is None: return "-"
        return f"{v*100:.0f}%"
    
    core_table = [
        {
            'metric': '敬业度',
            'in_bg_rank': fmt_rank(eng_rank),
            'rank_change': fmt_rank_change(eng_rank, eng_rank_2023),
            'score': fmt_score(eng_score),
            'yoy_change': fmt_yoy(eng_gap, eng_gap_pct),
            '_rank_raw': eng_rank,
        },
        {
            'metric': '满意度',
            'in_bg_rank': fmt_rank(sat_rank),
            'rank_change': fmt_rank_change(sat_rank, sat_rank_2023),
            'score': fmt_score(sat_score),
            'yoy_change': fmt_yoy(sat_gap, sat_gap_pct),
            '_rank_raw': sat_rank,
        },
    ]
    
    # 风险区间
    risk_table = []
    if bottom20_pct is not None:
        # 47道题中末20%的数量 ≈ bottom20_pct * 47
        count_20 = round(bottom20_pct * 47) if bottom20_pct else 0
        risk_table.append({
            'zone': 'BG末20%',
            'count': count_20,
            'pct': fmt_pct(bottom20_pct),
        })
    if bottom10_pct is not None:
        count_10 = round(bottom10_pct * 47) if bottom10_pct else 0
        risk_table.append({
            'zone': 'BG末10%',
            'count': count_10,
            'pct': fmt_pct(bottom10_pct),
        })
    
    return {
        'available': True,
        'org_name_jm': row[0] if row else '',
        'core_table': core_table,
        'risk_table': risk_table,
    }


# ============================================================
# 敬满逐题详细数据加载器（全量敬满数据 + BG数据 + 题目对照表）
# ============================================================

class JingmanDetailLoader:
    """加载report_tool_final/data/下的三个数据文件，提供逐题详情"""
    
    # 47道子题 key 列表
    Q47 = [
        "say_q1", "stay_q1", "stay_q2", "strive_q1",
        "gb_中干_q1",
        "gb_直接上级_q1", "gb_直接上级_q2", "gb_直接上级_q3", "gb_直接上级_q4",
        "gb_直接上级_q5", "gb_直接上级_q6", "gb_直接上级_q7", "gb_直接上级_q8",
        "gj_job_q1", "gj_job_q2", "gj_job_q3", "gj_job_q4", "gj_job_q5",
        "gr_绩效_q1", "gr_薪酬_q1", "gr_薪酬_q2", "gr_福利_q1", "gr_晋升_q1",
        "gc_卓越团队_q1", "gc_卓越团队_q2", "gc_卓越团队_q3",
        "gc_卓越团队_q4", "gc_卓越团队_q5", "gc_卓越团队_q6",
        "gc_协作信任_q1", "gc_协作信任_q2",
        "gc_工作支持_q1", "gc_工作支持_q2", "gc_工作支持_q3",
        "gc_文化价值观_q1", "gc_文化价值观_q2", "gc_文化价值观_q3",
        "gc_客户导向_q1", "gc_客户导向_q2",
        "gc_创造_q1",
        "gc_沟通渠道_q1", "gc_沟通渠道_q2",
        "gc_公司未来_q1", "gc_公司未来_q2",
        "gc_组织活力_q1",
        "gc_人才管理_q1",
        "gc_多样性_q1",
    ]
    
    ENG_Q_KEYS = {"say_q1", "stay_q1", "stay_q2", "strive_q1"}
    
    def __init__(self, dept_file, bg_file, var_file):
        """
        Args:
            dept_file: 全量敬满数据.xlsx 路径
            bg_file: BG相关数据.xlsx 路径
            var_file: 题目与标题对照表.xlsx 路径
        """
        try:
            import pandas as pd
            self.pd = pd
            import numpy as np
            self.np = np
        except ImportError:
            raise ImportError("需要安装pandas: pip3 install pandas openpyxl")
        
        self.dept_df = pd.read_excel(dept_file)
        self.bg_df = pd.read_excel(bg_file)
        
        var_df = pd.read_excel(var_file)
        self.var_map = {}
        for _, row in var_df.iterrows():
            self.var_map[str(row["key"])] = {
                "short": str(row["remark"]),
                "category": str(row["remark1"]),
                "full": str(row["remark2"]),
            }
        
        # 构建部门名->行 的索引
        self._dept_index = {}
        for idx, row in self.dept_df.iterrows():
            dept_name = str(row.get("部门", ""))
            if dept_name:
                self._dept_index[dept_name] = row
        
    def _normalize_diag_path(self, path):
        """将诊断数据路径转为敬满数据的部门名格式
        诊断: 示例BG/示例业务线/... -> BG缩写/示例业务线/...
        """
        parts = path.split('/')
        if parts:
            first = parts[0]
            match = re.match(r'^([A-Z]+\d*)', first)
            if match:
                parts[0] = match.group(1)
            match2 = re.match(r'^(S\d+)', first)
            if match2:
                parts[0] = match2.group(1)
            if first.startswith('Overseas'):
                parts[0] = 'OFS'
        return '/'.join(parts)
    
    def find_dept(self, diag_org_path):
        """根据诊断路径查找全量敬满数据行
        返回 pd.Series 或 None
        """
        norm = self._normalize_diag_path(diag_org_path)
        
        # 精确匹配
        if norm in self._dept_index:
            return self._dept_index[norm]
        
        # BG+末端匹配
        parts = norm.split('/')
        bg = parts[0] if parts else ''
        last = parts[-1] if parts else ''
        
        for dept_name, row in self._dept_index.items():
            d_parts = dept_name.split('/')
            if d_parts and d_parts[0] == bg and d_parts[-1] == last:
                return row
        
        # 末2段匹配
        if len(parts) >= 2:
            target = '/'.join(parts[-2:])
            for dept_name, row in self._dept_index.items():
                if dept_name.endswith(target):
                    return row
        
        return None
    
    def get_q_values(self, row, key):
        """提取某道题的所有关键数值"""
        pd = self.pd
        np = self.np
        
        def v(col):
            val = row.get(col, np.nan)
            return np.nan if pd.isna(val) else val
        
        return {
            "fav":        v(f"{key}_fav"),
            "fav2024":    v(f"{key}_fav2024"),
            "fav_bg":     v(f"{key}_fav_bg"),
            "fav2024_bg": v(f"{key}_fav2024_bg"),
            "diff":       v(f"{key}_fav_diff"),
            "growth":     v(f"{key}_fav_growth"),
        }
    
    def build_core_table(self, row):
        """从全量敬满数据计算核心维度表
        
        注意：fav_bg是击败率（如83表示击败83%），
        但"在BG的排名"应显示排名分位（如17%表示前17%），
        即 rank_pct = 100 - fav_bg
        """
        pd = self.pd
        np = self.np
        
        def v(col):
            val = row.get(col, np.nan)
            return np.nan if pd.isna(val) else val
        
        def fmt_rank_pct(fav_bg_val):
            """击败率转排名分位：83% → 17%（排在前17%）"""
            if pd.isna(fav_bg_val): return "-"
            rank_pct = 100 - fav_bg_val
            return f"{rank_pct:.0f}%"
        
        def fmt_rank_change(curr_bg, prev_bg):
            """排名变化：curr_bg/prev_bg都是击败率"""
            if pd.isna(curr_bg): return "-"
            if pd.isna(prev_bg): return "/"
            # 排名分位变化
            curr_rank = 100 - curr_bg
            prev_rank = 100 - prev_bg
            change = curr_rank - prev_rank  # 负值=排名上升（好事），正值=排名下降
            sign = "+" if change >= 0 else ""
            return f"{sign}{change:.0f}%（去年{prev_rank:.0f}%）"
        
        def fmt_score(val):
            if pd.isna(val): return "-"
            return f"{val:.1f}"
        
        def fmt_diff_growth(diff, growth):
            if pd.isna(diff) and pd.isna(growth): return "/"
            parts = []
            if not pd.isna(diff):
                sign = "+" if diff >= 0 else ""
                parts.append(f"{sign}{diff:.1f}")
            if not pd.isna(growth):
                sign = "+" if growth >= 0 else ""
                parts.append(f"（{sign}{growth:.1f}%）")
            return "".join(parts)
        
        results = []
        for metric in ["敬业度", "满意度"]:
            fav = v(f"{metric}_fav")
            fav_bg = v(f"{metric}_fav_bg")
            fav2024_bg = v(f"{metric}_fav2024_bg")
            diff = v(f"{metric}_fav_diff")
            growth = v(f"{metric}_fav_growth")
            
            results.append({
                'metric': metric,
                'in_bg_rank': fmt_rank_pct(fav_bg),
                'rank_change': fmt_rank_change(fav_bg, fav2024_bg),
                'score': fmt_score(fav),
                'yoy_change': fmt_diff_growth(diff, growth),
                '_rank_raw_bg': float(fav_bg) if not pd.isna(fav_bg) else None,  # 保留原始击败率
            })
        return results
    
    def build_risk_table(self, row):
        """从全量敬满数据计算风险区间表（与module2_risk_table逻辑一致）"""
        pd = self.pd
        np = self.np
        
        results = []
        for threshold in [20, 10]:
            count = 0
            count_gb = 0
            for key in self.Q47:
                fav_bg = row.get(f"{key}_fav_bg", np.nan)
                if pd.isna(fav_bg):
                    continue
                if fav_bg <= threshold:
                    count += 1
                    if key.startswith("gb_"):
                        count_gb += 1
            
            count_other = count - count_gb
            
            pct_col = f"pct_bottom{threshold}"
            pct_val = row.get(pct_col, np.nan)
            pct_str = f"{pct_val:.0f}%" if not pd.isna(pct_val) else "-"
            
            results.append({
                'zone': f"BG末{threshold}%",
                'count': count,
                'count_gb': count_gb,
                'count_other': count_other,
                'pct': pct_str,
                'count_str': f"{count}题" if count > 0 else "无",
                'gb_str': f"{count_gb}题" if count_gb > 0 else "/",
                'other_str': f"{count_other}题" if count_other > 0 else "/",
            })
        return results
    
    def build_analysis_text(self, row, chart_data, risk_table):
        """根据数据生成分析文字"""
        pd = self.pd
        np = self.np
        
        texts = {}
        
        # 核心维度分析文字
        eng_bg = row.get("敬业度_fav_bg", np.nan)
        sat_bg = row.get("满意度_fav_bg", np.nan)
        eng_2024 = row.get("敬业度_fav2024_bg", np.nan)
        sat_2024 = row.get("满意度_fav2024_bg", np.nan)
        
        # 判断排名等级
        def rank_level(bg_val):
            if pd.isna(bg_val): return None
            if bg_val >= 80: return "前列（前20%的分位）"
            elif bg_val >= 50: return "中上水平"
            elif bg_val >= 20: return "中下水平"
            else: return "靠后（后20%的分位）"
        
        eng_level = rank_level(eng_bg) if not pd.isna(eng_bg) else None
        sat_level = rank_level(sat_bg) if not pd.isna(sat_bg) else None
        
        core_parts = []
        if eng_level and sat_level:
            if eng_level == sat_level:
                core_parts.append(f"该部门敬业度与满意度总分都在BG内排名{eng_level}")
            else:
                core_parts.append(f"该部门敬业度在BG内排名{eng_level}，满意度在BG内排名{sat_level}")
        
        # 判断趋势
        has_trend = not pd.isna(eng_2024) or not pd.isna(sat_2024)
        if not has_trend:
            core_parts.append("，因为24年数据缺失无法查看趋势。")
        else:
            trend_parts = []
            if not pd.isna(eng_bg) and not pd.isna(eng_2024):
                change = eng_bg - eng_2024
                if change > 5:
                    trend_parts.append("敬业度排名上升")
                elif change < -5:
                    trend_parts.append("敬业度排名下降")
            if not pd.isna(sat_bg) and not pd.isna(sat_2024):
                change = sat_bg - sat_2024
                if change > 5:
                    trend_parts.append("满意度排名上升")
                elif change < -5:
                    trend_parts.append("满意度排名下降")
            if trend_parts:
                core_parts.append("，".join(trend_parts) + "。")
            else:
                core_parts.append("排名基本持平。")
        
        texts['core_analysis'] = "".join(core_parts) if core_parts else ""
        
        # 风险区间分析文字
        n_front = sum(1 for it in chart_data if it['fav_bg'] >= 50)
        n_total = len(chart_data)
        n_top20 = sum(1 for it in chart_data if it['fav_bg'] >= 80)
        n_bottom20 = risk_table[0]['count'] if risk_table else 0
        n_bottom10 = risk_table[1]['count'] if len(risk_table) > 1 else 0
        
        if n_front == n_total:
            risk_text = f"敬满{n_total}道子题全部都在BG内排名靠前"
            if n_top20 > 0:
                risk_text += f"，其中有{n_top20}题都位于BG内前20%。"
            else:
                risk_text += "。"
        elif n_front > n_total * 0.7:
            risk_text = f"敬满{n_total}道子题中{n_front}题在BG内排名靠前。"
        else:
            prefix = "仅" if n_front <= 10 else "有"
            risk_text = f"敬满{n_total}道子题中{prefix}{n_front}题在BG内排名靠前"
            if n_bottom20 > 0:
                risk_text += f"，有{n_bottom20}题位于BG末20%"
                if n_bottom10 > 0:
                    risk_text += f"（其中{n_bottom10}题位于末10%）"
            risk_text += "。"
        
        texts['risk_analysis'] = risk_text
        
        return texts
    
    def get_bg_min_fav_bg(self, bg_name):
        """获取该BG所有部门中47道子题的最小击败率（用于"倒数第一"红线）"""
        pd = self.pd
        np = self.np
        bg_subset = self.dept_df[self.dept_df["所属bg"] == bg_name]
        min_val = 100.0
        for key in self.Q47:
            col = f"{key}_fav_bg"
            if col in bg_subset.columns:
                vals = bg_subset[col].dropna()
                if not vals.empty:
                    min_val = min(min_val, vals.min())
        return float(min_val)
    
    def build_chart_data(self, row):
        """构建47道题柱状图数据，返回按fav_bg从高到低排序的列表"""
        pd = self.pd
        np = self.np
        items = []
        for key in self.Q47:
            fav_bg = row.get(f"{key}_fav_bg", np.nan)
            if pd.isna(fav_bg):
                fav_bg = 0.0
            family = "敬业度" if key in self.ENG_Q_KEYS else "满意度"
            info = self.var_map.get(key, {})
            items.append({
                "key": key,
                "fav_bg": float(fav_bg),
                "family": family,
                "short": info.get("short", key),
                "full": info.get("full", key),
            })
        items.sort(key=lambda x: x["fav_bg"], reverse=True)
        return items
    
    def build_bottom_table(self, row, threshold=10):
        """构建末10%题目明细表"""
        pd = self.pd
        np = self.np
        results = []
        for key in self.Q47:
            vals = self.get_q_values(row, key)
            fav_bg = vals["fav_bg"]
            if pd.isna(fav_bg) or fav_bg > threshold:
                continue
            info = self.var_map.get(key, {})
            
            # 格式化分值变化
            diff = vals["diff"]
            growth = vals["growth"]
            if pd.isna(diff) and pd.isna(growth):
                yoy_str = "-"
            else:
                parts = []
                if not pd.isna(diff):
                    sign = "+" if diff >= 0 else ""
                    parts.append(f"{sign}{diff:.1f}")
                if not pd.isna(growth):
                    sign = "+" if growth >= 0 else ""
                    parts.append(f"（{sign}{growth:.1f}%）")
                yoy_str = "".join(parts)
            
            results.append({
                "key": key,
                "short": info.get("short", key),
                "category": info.get("category", ""),
                "full": info.get("full", key),
                "fav_bg": float(fav_bg),
                "bg_rank_str": f"{fav_bg:.0f}%",
                "yoy_change": yoy_str,
            })
        results.sort(key=lambda x: x["fav_bg"])
        return results
    
    def build_subdivision_table(self, row):
        """构建细分项大表（4个分组）"""
        pd = self.pd
        np = self.np
        
        bg = str(row.get("所属bg", ""))
        bg_subset = self.bg_df[self.bg_df["bg"] == bg].set_index("question")
        
        # 部门内排名
        fav_scores = {}
        for key in self.Q47:
            v = row.get(f"{key}_fav", np.nan)
            fav_scores[key] = v if not pd.isna(v) else -999
        
        sorted_keys = sorted(fav_scores.keys(), key=lambda k: fav_scores[k], reverse=True)
        dept_rank_map = {}
        rank = 1
        for i, k in enumerate(sorted_keys):
            if i > 0 and fav_scores[k] < fav_scores[sorted_keys[i - 1]]:
                rank = i + 1
            dept_rank_map[k] = rank
        
        RANK_GAP = 10
        GROWTH_GAP = 10.0
        GROUP_LABELS = {
            "bg_higher": "BG排名较高的题（这些题在BG整体表现好，但本部门相对落后，值得重点关注）",
            "dept_higher": "部门排名较高的题（这些题本部门表现优于BG整体，是部门的相对优势项）",
            "dept_lag": "部门增幅落后BG（这些题BG整体涨幅明显，但本部门改善较慢，需警惕差距拉大）",
            "dept_lead": "部门增幅领先BG（这些题本部门改善速度明显快于BG整体，是近期进步亮点）",
        }
        
        results = []
        for key in self.Q47:
            vals = self.get_q_values(row, key)
            dept_rank = dept_rank_map.get(key, 99)
            dept_growth = vals["growth"]
            
            if key not in bg_subset.index:
                continue
            bg_row = bg_subset.loc[key]
            bg_rank = int(bg_row["rank_2025"])
            bg_growth = bg_row["growth"] * 100  # BG的growth是小数
            
            rank_diff = dept_rank - bg_rank
            if pd.isna(dept_growth):
                growth_diff = np.nan
            else:
                growth_diff = dept_growth - bg_growth
            
            info = self.var_map.get(key, {})
            
            def fmt_gd(gd):
                if pd.isna(gd): return "-"
                sign = "+" if gd >= 0 else ""
                return f"{sign}{gd:.1f}%"
            
            base = {
                "key": key,
                "short": info.get("short", key),
                "full": info.get("full", key),
                "dept_rank": dept_rank,
                "bg_rank": bg_rank,
                "rank_diff": rank_diff,
                "growth_diff": float(growth_diff) if not pd.isna(growth_diff) else None,
                "growth_diff_str": fmt_gd(growth_diff),
            }
            
            if dept_rank - bg_rank > RANK_GAP:
                results.append({**base, "group": "bg_higher", "group_label": GROUP_LABELS["bg_higher"]})
            if bg_rank - dept_rank > RANK_GAP:
                results.append({**base, "group": "dept_higher", "group_label": GROUP_LABELS["dept_higher"]})
            if not pd.isna(growth_diff) and bg_growth - dept_growth > GROWTH_GAP:
                results.append({**base, "group": "dept_lag", "group_label": GROUP_LABELS["dept_lag"]})
            if not pd.isna(growth_diff) and dept_growth - bg_growth > GROWTH_GAP:
                results.append({**base, "group": "dept_lead", "group_label": GROUP_LABELS["dept_lead"]})
        
        group_order = list(GROUP_LABELS.keys())
        results.sort(key=lambda x: (
            group_order.index(x["group"]),
            -abs(x["rank_diff"]) if x["group"] in ("bg_higher", "dept_higher")
            else -abs(x["growth_diff"]) if x["growth_diff"] is not None else 0
        ))
        return results


def generate_html_report(report: Dict, output_path: str):
    """生成HTML报告"""
    
    # 颜色映射
    color_class_map = {
        '红': 'red',
        '黄': 'yellow',
        '绿': 'green'
    }
    
    # 预警等级对应的勾选（"其它"等同于"无预警"，合并为5类）
    warning_checks = {
        '一级预警': [True, False, False, False, False],
        '二级预警': [False, True, False, False, False],
        '三级预警': [False, False, True, False, False],
        '无预警': [False, False, False, True, False],
        '其它': [False, False, False, True, False],
        '全绿': [False, False, False, False, True]
    }
    
    checks = warning_checks.get(report['warning_level'], [False]*5)
    
    html_content = f'''<!DOCTYPE html>
<html lang="zh-CN">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>组织诊断报告 - {report['org_name']}</title>
    <style>
        * {{ margin: 0; padding: 0; box-sizing: border-box; }}
        
        canvas {{
            max-width: 100% !important;
            height: auto !important;
        }}
        
        @media print {{
            body {{ margin: 0; padding: 10px; }}
            .no-print {{ display: none !important; }}
            .page-break {{ page-break-before: always; }}
            .container {{ max-width: 100%; padding: 20px; box-shadow: none; overflow: visible !important; }}
            canvas {{ max-width: 100% !important; height: auto !important; }}
            .org-chart-wrapper {{
                padding: 0 !important;
                margin: 0 !important;
                overflow: visible !important;
                page-break-inside: avoid;
                break-inside: avoid;
                background: none !important;
                border-radius: 0 !important;
            }}
            .org-chart-controls {{ display: none !important; }}
            .org-chart-container {{
                width: auto !important;
                max-height: none !important;
                overflow: visible !important;
                border: none !important;
                border-radius: 0 !important;
                padding: 0 !important;
            }}
            .org-chart {{
                /* JS beforeprint 会动态设置 zoom */
                transform: none !important;
                padding: 5px 0 !important;
            }}
            .org-child-container {{
                padding: 0 2px;
            }}
            .org-down-connector {{
                height: 10px;
            }}
            .org-connector {{
                height: 10px;
            }}
        }}
        
        body {{
            font-family: "Microsoft YaHei", "SimSun", Arial, sans-serif;
            background: #f5f5f5;
            padding: 20px;
            line-height: 1.8;
        }}
        
        .container {{
            max-width: 1200px;
            margin: 0 auto;
            background: white;
            padding: 40px;
            box-shadow: 0 2px 10px rgba(0,0,0,0.1);
        }}
        
        .header {{
            text-align: center;
            margin-bottom: 30px;
            padding-bottom: 20px;
            border-bottom: 3px solid #1890ff;
        }}
        
        .header h1 {{
            font-size: 24px;
            color: #333;
            margin-bottom: 15px;
            font-weight: normal;
        }}
        
        .header .org-name {{
            font-size: 28px;
            color: #1890ff;
            font-weight: bold;
            margin-bottom: 12px;
        }}
        
        .header .leader-info {{
            font-size: 15px;
            color: #666;
            margin-bottom: 8px;
        }}
        
        .header .time-info {{
            font-size: 13px;
            color: #999;
        }}
        
        .info-row {{
            margin: 15px 0;
            font-size: 16px;
        }}
        
        .info-row strong {{
            color: #1890ff;
            margin-right: 10px;
        }}
        
        h2 {{
            font-size: 24px;
            color: #333;
            margin: 30px 0 20px;
            padding-bottom: 10px;
            border-bottom: 2px solid #1890ff;
        }}
        
        h2.diagnosis-title {{
            text-align: center;
            border-bottom: none;
            padding-bottom: 0;
            margin-bottom: 20px;
        }}
        
        h3 {{
            font-size: 20px;
            color: #555;
            margin: 25px 0 15px;
        }}
        
        h4 {{
            font-size: 18px;
            color: #666;
            margin: 20px 0 10px;
        }}
        
        h5 {{
            font-size: 15px;
            font-weight: bold;
            color: #333;
            margin: 20px 0 10px;
        }}
        
        /* 段落级小标题（统一替代内联p标题） */
        .section-subtitle {{
            font-size: 15px;
            font-weight: bold;
            color: #333;
            margin: 25px 0 10px;
        }}
        
        /* 辅助说明文字（统一风格） */
        .note-text {{
            font-size: 13px;
            color: #888;
            font-style: italic;
            margin: 5px 0 15px;
        }}
        
        /* 缺失数据/空数据提示框 */
        .empty-hint {{
            margin: 20px 0;
            padding: 20px;
            background: #f5f5f5;
            border-left: 4px solid #faad14;
            color: #666;
            font-size: 14px;
        }}
        
        /* 备注框 */
        .remark-box {{
            margin-top: 15px;
            padding: 12px 15px;
            background: #f5f5f5;
            border-left: 4px solid #1890ff;
            font-size: 14px;
            color: #666;
        }}
        
        /* 统一表头：蓝系 */
        .th-blue {{
            background: #D6E4F0;
            color: #2E75B6;
        }}
        
        /* 统一表头：红系（预警/风险） */
        .th-red {{
            background: #FCE4D6;
            color: #C00000;
        }}
        
        .warning-box {{
            background: #e6f7ff;
            border-left: 4px solid #1890ff;
            padding: 20px;
            margin: 20px 0;
            border-radius: 4px;
        }}
        
        .warning-box.level-1 {{ background: #fff1f0; border-color: #ff4d4f; }}
        .warning-box.level-2 {{ background: #fff7e6; border-color: #faad14; }}
        .warning-box.level-3 {{ background: #fffbe6; border-color: #fadb14; }}
        .warning-box.level-green {{ background: #f6ffed; border-color: #52c41a; }}
        
        .warning-lights {{
            display: flex;
            gap: 30px;
            justify-content: center;
            margin: 0;
            padding: 15px 0;
            background: transparent;
            border-radius: 0;
        }}
        
        .light-item {{
            text-align: center;
            flex: 1;
        }}
        
        .light-circle {{
            width: 80px;
            height: 80px;
            border-radius: 50%;
            margin: 0 auto 15px;
            display: flex;
            align-items: center;
            justify-content: center;
            font-weight: bold;
            font-size: 13px;
            color: white;
            box-shadow: 0 4px 10px rgba(0,0,0,0.15);
            white-space: nowrap;
        }}
        
        .light-circle.red {{ background: linear-gradient(135deg, #ff4d4f 0%, #cf1322 100%); }}
        .light-circle.yellow {{ background: linear-gradient(135deg, #faad14 0%, #d48806 100%); }}
        .light-circle.green {{ background: linear-gradient(135deg, #52c41a 0%, #389e0d 100%); }}
        .light-circle.gray {{ background: linear-gradient(135deg, #d9d9d9 0%, #8c8c8c 100%); }}
        
        .light-label {{
            font-size: 16px;
            color: #666;
            font-weight: bold;
        }}
        
        .light-status {{
            font-size: 14px;
            color: #999;
            margin-top: 5px;
        }}
        
        table {{
            width: 100%;
            border-collapse: collapse;
            margin: 20px 0;
            background: white;
        }}
        
        table th, table td {{
            border: 1px solid #e8e8e8;
            padding: 12px;
            text-align: center;
        }}
        
        table th {{
            background: #D6E4F0;
            font-weight: bold;
            color: #2E75B6;
        }}
        
        table tr:hover {{
            background: #f5f5f5;
        }}
        
        .checkmark {{
            color: #52c41a;
            font-size: 20px;
            font-weight: bold;
        }}
        
        .checkmark-red {{
            color: #ff4d4f;
            font-size: 20px;
            font-weight: bold;
        }}
        
        .metric-grid {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
            gap: 15px;
            margin: 20px 0;
        }}
        
        .metric-card {{
            background: #fafafa;
            padding: 15px;
            border-radius: 8px;
            border-left: 4px solid #1890ff;
        }}
        
        .metric-card .label {{
            font-size: 14px;
            color: #666;
            margin-bottom: 5px;
        }}
        
        .metric-card .value {{
            font-size: 24px;
            font-weight: bold;
            color: #1890ff;
        }}
        
        .toolbar {{
            position: fixed;
            top: 20px;
            right: 20px;
            display: flex;
            gap: 10px;
            z-index: 1000;
        }}
        
        .btn {{
            padding: 12px 24px;
            border: none;
            border-radius: 6px;
            font-size: 14px;
            font-weight: bold;
            cursor: pointer;
            transition: all 0.3s;
            box-shadow: 0 2px 8px rgba(0,0,0,0.15);
        }}
        
        .btn-primary {{
            background: linear-gradient(135deg, #1890ff 0%, #0050b3 100%);
            color: white;
        }}
        
        .btn-success {{
            background: linear-gradient(135deg, #52c41a 0%, #389e0d 100%);
            color: white;
        }}
        
        .btn:hover {{
            transform: translateY(-2px);
            box-shadow: 0 4px 12px rgba(0,0,0,0.2);
        }}
        
        .footer {{
            text-align: center;
            margin-top: 40px;
            padding-top: 20px;
            border-top: 1px solid #e8e8e8;
            color: #999;
            font-size: 14px;
        }}
        
        /* 组织架构图样式 */
        .org-chart-wrapper {{
            margin: 30px 0;
            padding: 20px;
            background: #fafafa;
            border-radius: 8px;
            position: relative;
        }}
        
        .org-chart-controls {{
            display: flex;
            justify-content: center;
            gap: 10px;
            margin-bottom: 15px;
            padding: 10px;
            background: white;
            border-radius: 6px;
            box-shadow: 0 2px 6px rgba(0,0,0,0.08);
        }}
        
        .zoom-btn {{
            padding: 8px 16px;
            border: 1px solid #d9d9d9;
            background: white;
            border-radius: 4px;
            cursor: pointer;
            font-size: 14px;
            transition: all 0.3s;
        }}
        
        .zoom-btn:hover {{
            background: #f5f5f5;
            border-color: #1890ff;
        }}
        
        .zoom-level {{
            padding: 8px 16px;
            color: #666;
            font-size: 14px;
            font-weight: bold;
        }}
        
        .org-chart-container {{
            width: 100%;
            overflow: auto;
            border: 1px solid #e8e8e8;
            border-radius: 6px;
            background: white;
        }}
        
        .org-chart {{
            display: flex;
            flex-direction: column;
            align-items: center;
            min-width: max-content;
            padding: 40px;
            transform-origin: top left;
            transition: transform 0.3s ease;
        }}
        
        .org-node {{
            margin: 0;
            padding: 12px 16px;
            background: white;
            border: 2px solid #d9d9d9;
            border-radius: 6px;
            min-width: 150px;
            max-width: 200px;
            text-align: center;
            box-shadow: 0 2px 6px rgba(0,0,0,0.08);
            transition: all 0.3s;
            position: relative;
        }}
        
        .org-node:hover {{
            transform: translateY(-2px) scale(1.05);
            box-shadow: 0 4px 12px rgba(0,0,0,0.15);
            z-index: 10;
        }}
        
        .org-node.red {{
            border-color: #ff4d4f;
            background: #fff1f0;
        }}
        
        .org-node.yellow {{
            border-color: #faad14;
            background: #fffbe6;
        }}
        
        .org-node.green {{
            border-color: #52c41a;
            background: #f6ffed;
        }}
        
        .org-node.gray {{
            border-color: #d9d9d9;
            background: #fafafa;
        }}
        
        .org-node.virtual {{
            border-style: dashed;
            border-width: 2px;
            min-width: 100px;
            padding: 8px 12px;
        }}
        
        .org-node-name {{
            font-size: 14px;
            font-weight: bold;
            color: #333;
            margin-bottom: 4px;
            word-break: keep-all;
            white-space: nowrap;
            overflow: hidden;
            text-overflow: ellipsis;
        }}
        
        .org-node-leader {{
            font-size: 12px;
            color: #666;
            margin-bottom: 2px;
            word-break: break-all;
        }}
        
        .org-node-count {{
            font-size: 11px;
            color: #999;
        }}
        
        /* 子节点容器 - 包含横线和子节点列表 */
        .org-children-wrapper {{
            display: flex;
            flex-direction: column;
            align-items: center;
        }}
        
        /* 父节点到横线的垂直主干线 */
        .org-down-connector {{
            width: 2px;
            height: 20px;
            background: #c8c8c8;
        }}
        
        /* 横线容器 - 横跨所有子节点 */
        .org-children {{
            display: flex;
            flex-direction: row;
            justify-content: center;
        }}
        
        /* 每个子节点列 */
        .org-child-container {{
            display: flex;
            flex-direction: column;
            align-items: center;
            position: relative;
            padding: 0 8px;
        }}
        
        /* 用每个子节点顶部的伪元素画横线的左半段 */
        .org-child-container::before {{
            content: '';
            position: absolute;
            top: 0;
            left: 0;
            right: 50%;
            height: 2px;
            background: #c8c8c8;
        }}
        /* 用每个子节点顶部的伪元素画横线的右半段 */
        .org-child-container::after {{
            content: '';
            position: absolute;
            top: 0;
            left: 50%;
            right: 0;
            height: 2px;
            background: #c8c8c8;
        }}
        /* 第一个子节点不画左半段 */
        .org-child-container:first-child::before {{
            display: none;
        }}
        /* 最后一个子节点不画右半段 */
        .org-child-container:last-child::after {{
            display: none;
        }}
        /* 只有1个子节点时不画横线（两侧都不画） */
        .org-child-container:only-child::before,
        .org-child-container:only-child::after {{
            display: none;
        }}
        
        /* 每个子节点上方的竖线（从横线到节点） */
        .org-connector {{
            width: 2px;
            height: 20px;
            background: #c8c8c8;
        }}
        
        /* ===== 2.3.3 敬满开放题 ===== */
        .jm-open-overview {{
            margin-bottom: 26px;
            border: 1px solid #e8e8e8;
            border-radius: 8px;
            overflow: hidden;
        }}
        .jm-open-overview-header {{
            padding: 10px 16px;
            font-size: 14px;
            font-weight: 700;
            background: #f0f5ff;
            border-bottom: 1px solid #d6e4ff;
            color: #1d39c4;
        }}
        .jm-open-overview-table {{
            width: 100%;
            border-collapse: collapse;
            font-size: 13px;
        }}
        .jm-open-overview-table th {{
            padding: 8px 14px;
            text-align: center;
            font-weight: 600;
            color: #666;
            font-size: 12px;
            background: #fafafa;
            border-bottom: 1px solid #eee;
        }}
        .jm-open-overview-table th:last-child {{
            text-align: left;
        }}
        .jm-open-overview-table td {{
            padding: 7px 14px;
            border-bottom: 1px solid #f0f0f0;
            vertical-align: middle;
            text-align: center;
        }}
        .jm-open-overview-table td:last-child {{
            text-align: left;
        }}
        .jm-open-overview-table tr:last-child td {{
            border-bottom: none;
        }}
        .jm-open-dim-name {{
            font-weight: 600;
            color: #222;
        }}
        .jm-open-type-badge {{
            display: inline-block;
            padding: 2px 10px;
            border-radius: 4px;
            font-size: 11px;
            font-weight: 600;
        }}
        .jm-open-type-badge.only-concern {{
            background: #fff1f0;
            color: #cf1322;
            border: 1px solid #ffa39e;
        }}
        .jm-open-type-badge.mixed {{
            background: #fffbe6;
            color: #ad6800;
            border: 1px solid #ffe58f;
        }}
        .jm-open-type-badge.only-good {{
            background: #f6ffed;
            color: #389e0d;
            border: 1px solid #b7eb8f;
        }}
        .jm-open-finding-summary {{
            color: #666;
            font-size: 12px;
        }}
        
        /* 维度卡片 */
        .jm-open-card {{
            margin-bottom: 22px;
            border: 2px solid #e8e8e8;
            border-radius: 8px;
            overflow: hidden;
        }}
        .jm-open-card.card-concern {{
            border-color: #ffa39e;
        }}
        .jm-open-card.card-concern .jm-open-card-header {{
            background: #fff1f0;
            border-bottom-color: #ffa39e;
        }}
        .jm-open-card.card-mixed {{
            border-color: #ffe58f;
        }}
        .jm-open-card.card-mixed .jm-open-card-header {{
            background: #fffbe6;
            border-bottom-color: #ffe58f;
        }}
        .jm-open-card.card-good {{
            border-color: #b7eb8f;
        }}
        .jm-open-card.card-good .jm-open-card-header {{
            background: #f6ffed;
            border-bottom-color: #b7eb8f;
        }}
        .jm-open-card-header {{
            padding: 10px 16px;
            font-size: 14px;
            font-weight: 700;
            background: #f8f9fa;
            border-bottom: 2px solid #e8e8e8;
            color: #1a1a1a;
        }}
        .jm-open-card-body {{
            padding: 0;
        }}
        
        /* 类别标签行 */
        .jm-open-cat-row {{
            display: flex;
            align-items: center;
            padding: 8px 16px 4px;
            gap: 6px;
        }}
        .jm-open-cat-tag {{
            display: inline-block;
            padding: 2px 10px;
            border-radius: 4px;
            font-size: 12px;
            font-weight: 600;
        }}
        .jm-open-cat-tag.concern {{
            background: #fff1f0;
            color: #cf1322;
            border: 1px solid #ffa39e;
        }}
        .jm-open-cat-tag.good {{
            background: #f6ffed;
            color: #389e0d;
            border: 1px solid #b7eb8f;
        }}
        
        /* 发现表格 */
        .jm-open-finding-table {{
            width: 100%;
            border-collapse: collapse;
            font-size: 13px;
            line-height: 1.6;
        }}
        .jm-open-finding-table thead th {{
            padding: 7px 12px;
            text-align: left;
            font-weight: 600;
            color: #666;
            font-size: 12px;
            border-bottom: 1px solid #eee;
            background: #fafafa;
        }}
        .jm-open-finding-table thead th:nth-child(1) {{ width: 5%; text-align: center; }}
        .jm-open-finding-table thead th:nth-child(2) {{ width: 38%; }}
        .jm-open-finding-table thead th:nth-child(3) {{ width: 57%; }}
        .jm-open-finding-table tbody td {{
            padding: 8px 12px;
            vertical-align: top;
            border-bottom: 1px solid #f0f0f0;
            text-align: left;
        }}
        .jm-open-finding-table tbody tr:last-child td {{
            border-bottom: none;
        }}
        .jm-open-finding-table tbody td:nth-child(1) {{
            text-align: center;
            color: #999;
            font-size: 12px;
        }}
        .jm-open-finding-title {{
            font-weight: 600;
            color: #222;
        }}
        .jm-open-quotes {{
            color: #555;
            font-size: 12.5px;
            line-height: 1.8;
        }}
        .jm-open-count {{
            font-weight: 700;
            margin-right: 2px;
        }}
        .jm-open-count.high {{ color: #cf1322; }}
        .jm-open-count.medium {{ color: #d48806; }}
        .jm-open-count.low {{ color: #555; }}
        .jm-open-count.good {{ color: #389e0d; }}
        .jm-open-quote-line {{
            display: block;
            color: #666;
            padding-left: 2px;
        }}
        .jm-open-divider {{
            border: none;
            border-top: 1px dashed #e8e8e8;
            margin: 4px 16px;
        }}
        /* ─── 词云样式 ─── */
        .wc-section {{
            margin-top: 28px;
            page-break-inside: avoid;
        }}
        .wc-section-title {{
            text-align: center;
            font-size: 15px;
            font-weight: 700;
            color: #1a1a1a;
            margin-bottom: 4px;
        }}
        .wc-section-subtitle {{
            text-align: center;
            font-size: 12px;
            color: #999;
            margin-bottom: 14px;
        }}
        .wc-cloud-container {{
            position: relative;
            width: 100%;
            height: 480px;          /* 初始布局高度，JS 放置完后会自适应收缩 */
            border: 1px solid #eee;
            border-radius: 14px;
            background: #fff;
            overflow: hidden;
            transition: height .3s ease;
        }}
        .wc-word {{
            position: absolute;
            white-space: nowrap;
            cursor: default;
            font-weight: 700;
            transition: transform .15s ease, opacity .15s ease;
            user-select: none;
            transform-origin: center center;
        }}
        .wc-cloud-container:hover .wc-word {{
            opacity: 0.35;
        }}
        .wc-cloud-container:hover .wc-word:hover {{
            opacity: 1;
            transform: scale(1.12);
            z-index: 100;
        }}
        .wc-tooltip {{
            position: fixed;
            pointer-events: none;
            background: rgba(0,0,0,.78);
            color: #fff;
            font-size: 12px;
            padding: 5px 10px;
            border-radius: 6px;
            display: none;
            z-index: 999;
            white-space: nowrap;
        }}
        .wc-legend {{
            display: flex;
            justify-content: center;
            gap: 28px;
            margin-top: 12px;
        }}
        .wc-legend-item {{
            display: flex;
            align-items: center;
            gap: 7px;
            font-size: 12px;
            color: #777;
        }}
        .wc-legend-bar {{
            width: 32px;
            height: 8px;
            border-radius: 4px;
        }}
        .wc-legend-bar.positive {{ background: linear-gradient(90deg, #b7eb8f, #237804); }}
        .wc-legend-bar.neutral  {{ background: linear-gradient(90deg, #ffe58f, #ad6800); }}
        .wc-legend-bar.negative {{ background: linear-gradient(90deg, #ffa39e, #a8071a); }}
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>组织诊断报告</h1>
            <div class="org-name">{report['org_name']}</div>
            <div class="leader-info">部门负责人：{report['leader_name']}</div>
            <div class="time-info">{datetime.now().strftime("%Y年%m月%d日")}</div>
        </div>
        
        '''
    
    warning_style = 'color: #ff4d4f; font-weight: bold;' if report['warning_level'] in ['一级预警', '二级预警', '三级预警'] else 'color: #52c41a; font-weight: bold;'
    
    html_content += f'''
        <h2 class="diagnosis-title">诊断筛查结果：<span style="{warning_style}">【{report['warning_level']}】</span></h2>
        
        <table>
            <tr>
                <th style="background: #ffccc7; color: #333;">一级预警</th>
                <th style="background: #ffe7ba; color: #333;">二级预警</th>
                <th style="background: #fff1b8; color: #333;">三级预警</th>
                <th style="background: #e6f7e6; color: #333;">无预警</th>
                <th style="background: #d9f7be; color: #333;">全绿</th>
            </tr>
            <tr>
                <td style="background: white;">{'<span class="checkmark-red">✓</span>' if checks[0] else ''}</td>
                <td style="background: white;">{'<span class="checkmark-red">✓</span>' if checks[1] else ''}</td>
                <td style="background: white;">{'<span class="checkmark-red">✓</span>' if checks[2] else ''}</td>
                <td style="background: white;">{'<span class="checkmark">✓</span>' if checks[3] else ''}</td>
                <td style="background: white;">{'<span class="checkmark">✓</span>' if checks[4] else ''}</td>
            </tr>
        </table>
        
        <h3>三维度评估：<span style="font-weight:normal;color:#1a6fb5;">{report['tri_dim_desc']}</span></h3>
        
        <div style="background:#f7f8fa;border-radius:8px;padding:20px 15px 15px 15px;">
            <div class="warning-lights" style="margin:0;">
                <div class="light-item">
                    <div class="light-circle {color_class_map.get(report['dimensions']['full_feedback']['color'], 'gray')}">
                        全面反馈
                    </div>
                    <div class="light-label">{report['dimensions']['full_feedback']['text']}</div>
                </div>
                <div class="light-item">
                    <div class="light-circle {color_class_map.get(report['dimensions']['abnormal']['color'], 'gray')}">
                        异动
                    </div>
                    <div class="light-label">{report['dimensions']['abnormal']['text']}</div>
                </div>
                <div class="light-item">
                    <div class="light-circle {color_class_map.get(report['dimensions']['jingman']['color'], 'gray')}">
                        敬满
                    </div>
                    <div class="light-label">{report['dimensions']['jingman']['text']}</div>
                </div>
            </div>
            
            <div style="margin-top:16px;padding-top:12px;border-top:1px solid #e8e8e8;font-size:12px;color:#999;line-height:1.8;text-align:center;">
                *注：【严重预警】：触发多个异常指标，或偏离程度较大；【预警】：触发部分异常指标，且偏离程度不大
            </div>
        </div>
        
        <h2>具体数据情况</h2>
        
        <h3>1 全面反馈</h3>
'''
    
    # 添加负责人数据
    if 'leader_feedback' in report:
        lf = report['leader_feedback']
        
        # 第一行主指标变色判断
        # 总分排名
        total_rank_style = 'border-left: 4px solid #1890ff;'
        total_rank_value_style = ''
        if lf['_total_rank_raw'] >= 0.9:
            total_rank_style = 'border-left: 4px solid #ff4d4f; background: #ffccc7;'
            total_rank_value_style = 'color: #ff4d4f;'
        elif lf['_total_rank_raw'] >= 0.67:
            total_rank_style = 'border-left: 4px solid #d48806; background: #fff1b8;'
            total_rank_value_style = 'color: #d48806;'
        
        # 同级打分排名
        peer_rank_style = 'border-left: 4px solid #1890ff;'
        peer_rank_value_style = ''
        if lf['_peer_rank_raw'] >= 0.9:
            peer_rank_style = 'border-left: 4px solid #ff4d4f; background: #ffccc7;'
            peer_rank_value_style = 'color: #ff4d4f;'
        elif lf['_peer_rank_raw'] >= 0.67:
            peer_rank_style = 'border-left: 4px solid #d48806; background: #fff1b8;'
            peer_rank_value_style = 'color: #d48806;'
        
        # 第二行辅指标变色判断
        # 下属愿意跟随
        subordinate_style = 'border-left: 4px solid #1890ff;'
        subordinate_value_style = ''
        if lf['_subordinate_raw'] <= 0.8 and lf['_subordinate_raw'] > 0:
            subordinate_style = 'border-left: 4px solid #d48806; background: #fff1b8;'
            subordinate_value_style = 'color: #d48806;'
        
        # 同级愿意合作
        peer_coop_style = 'border-left: 4px solid #1890ff;'
        peer_coop_value_style = ''
        if lf['_peer_coop_raw'] <= 0.8 and lf['_peer_coop_raw'] > 0:
            peer_coop_style = 'border-left: 4px solid #d48806; background: #fff1b8;'
            peer_coop_value_style = 'color: #d48806;'
        
        # 专业能力胜任
        competent_style = 'border-left: 4px solid #1890ff;'
        competent_value_style = ''
        if lf['_competent_raw'] <= 0.8 and lf['_competent_raw'] > 0:
            competent_style = 'border-left: 4px solid #d48806; background: #fff1b8;'
            competent_value_style = 'color: #d48806;'
        
        # 负面反馈人次
        negative_style = 'border-left: 4px solid #1890ff;'
        negative_value_style = ''
        if lf['_negative_raw'] > 0:
            negative_style = 'border-left: 4px solid #d48806; background: #fff1b8;'
            negative_value_style = 'color: #d48806;'
        
        html_content += f'''
        <h4>1.1 负责人{report['leader_name']}的全面反馈分数</h4>
        
        <div style="margin: 20px 0;">
            <div style="display: grid; grid-template-columns: repeat(2, 1fr); gap: 20px; margin-bottom: 15px;">
                <div class="metric-card" style="{total_rank_style}">
                    <div class="label">总分_BG内排名</div>
                    <div class="value" style="{total_rank_value_style}">{lf['total_rank']}</div>
                </div>
                <div class="metric-card" style="{peer_rank_style}">
                    <div class="label">同级打分_BG内排名</div>
                    <div class="value" style="{peer_rank_value_style}">{lf['peer_rank']}</div>
                </div>
            </div>
            <div style="display: grid; grid-template-columns: repeat(4, 1fr); gap: 15px;">
                <div class="metric-card" style="{subordinate_style}">
                    <div class="label">下属愿意跟随</div>
                    <div class="value" style="{subordinate_value_style}">{lf['subordinate_follow']}</div>
                </div>
                <div class="metric-card" style="{peer_coop_style}">
                    <div class="label">同级愿意合作</div>
                    <div class="value" style="{peer_coop_value_style}">{lf['peer_cooperate']}</div>
                </div>
                <div class="metric-card" style="{competent_style}">
                    <div class="label">专业能力胜任</div>
                    <div class="value" style="{competent_value_style}">{lf['competent']}</div>
                </div>
                <div class="metric-card" style="{negative_style}">
                    <div class="label">负面反馈人次</div>
                    <div class="value" style="{negative_value_style}">{lf['negative_feedback']}</div>
                </div>
            </div>
        </div>
        
        <div style="margin: 30px 0; overflow: hidden;">
            <div style="display: grid; grid-template-columns: repeat(2, 1fr); gap: 30px; max-width: 100%;">
                <div style="overflow: hidden; min-width: 0;">
                    <h5 style="text-align: center; margin-bottom: 20px;">All In 分数</h5>
                    <canvas id="allinChart" style="max-width: 100%; height: auto;"></canvas>
                </div>
                <div style="overflow: hidden; min-width: 0;">
                    <h5 style="text-align: center; margin-bottom: 20px;">价值观</h5>
                    <canvas id="valuesChart" style="max-width: 100%; height: auto;"></canvas>
                </div>
            </div>
        </div>
        
        <script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.0/dist/chart.umd.min.js"></script>
        <script>
            // All In 分数雷达图
            const allinCtx = document.getElementById('allinChart').getContext('2d');
            new Chart(allinCtx, {{
                type: 'radar',
                data: {{
                    labels: ['Insight 看清楚', 'Insight 善决策', 'Inspire 擅协作', 'Inspire 会带兵', 'Win 拿结果', 'Win 巧经营'],
                    datasets: [{{
                        label: '下级',
                        data: [{lf['radar']['allin']['insight_see']['subordinate']}, {lf['radar']['allin']['insight_decide']['subordinate']}, {lf['radar']['allin']['inspire_cooperate']['subordinate']}, {lf['radar']['allin']['inspire_lead']['subordinate']}, {lf['radar']['allin']['win_result']['subordinate']}, {lf['radar']['allin']['win_manage']['subordinate']}],
                        borderColor: 'rgb(54, 162, 235)',
                        backgroundColor: 'rgba(54, 162, 235, 0.2)',
                        pointBackgroundColor: 'rgb(54, 162, 235)',
                        pointBorderColor: '#fff',
                        pointHoverBackgroundColor: '#fff',
                        pointHoverBorderColor: 'rgb(54, 162, 235)'
                    }}, {{
                        label: '同级',
                        data: [{lf['radar']['allin']['insight_see']['peer']}, {lf['radar']['allin']['insight_decide']['peer']}, {lf['radar']['allin']['inspire_cooperate']['peer']}, {lf['radar']['allin']['inspire_lead']['peer']}, {lf['radar']['allin']['win_result']['peer']}, {lf['radar']['allin']['win_manage']['peer']}],
                        borderColor: 'rgb(255, 99, 132)',
                        backgroundColor: 'rgba(255, 99, 132, 0.2)',
                        pointBackgroundColor: 'rgb(255, 99, 132)',
                        pointBorderColor: '#fff',
                        pointHoverBackgroundColor: '#fff',
                        pointHoverBorderColor: 'rgb(255, 99, 132)'
                    }}]
                }},
                options: {{
                    responsive: true,
                    maintainAspectRatio: true,
                    scales: {{
                        r: {{
                            beginAtZero: true,
                            max: 5,
                            ticks: {{
                                stepSize: 1
                            }}
                        }}
                    }},
                    plugins: {{
                        legend: {{
                            position: 'top'
                        }}
                    }}
                }}
            }});
            
            // 价值观雷达图
            const valuesCtx = document.getElementById('valuesChart').getContext('2d');
            new Chart(valuesCtx, {{
                type: 'radar',
                data: {{
                    labels: ['用户', '正直', '协作', '进取', '创造'],
                    datasets: [{{
                        label: '下级',
                        data: [{lf['radar']['values']['user']['subordinate']}, {lf['radar']['values']['integrity']['subordinate']}, {lf['radar']['values']['cooperation']['subordinate']}, {lf['radar']['values']['progress']['subordinate']}, {lf['radar']['values']['create']['subordinate']}],
                        borderColor: 'rgb(54, 162, 235)',
                        backgroundColor: 'rgba(54, 162, 235, 0.2)',
                        pointBackgroundColor: 'rgb(54, 162, 235)',
                        pointBorderColor: '#fff',
                        pointHoverBackgroundColor: '#fff',
                        pointHoverBorderColor: 'rgb(54, 162, 235)'
                    }}, {{
                        label: '同级',
                        data: [{lf['radar']['values']['user']['peer']}, {lf['radar']['values']['integrity']['peer']}, {lf['radar']['values']['cooperation']['peer']}, {lf['radar']['values']['progress']['peer']}, {lf['radar']['values']['create']['peer']}],
                        borderColor: 'rgb(255, 99, 132)',
                        backgroundColor: 'rgba(255, 99, 132, 0.2)',
                        pointBackgroundColor: 'rgb(255, 99, 132)',
                        pointBorderColor: '#fff',
                        pointHoverBackgroundColor: '#fff',
                        pointHoverBorderColor: 'rgb(255, 99, 132)'
                    }}]
                }},
                options: {{
                    responsive: true,
                    maintainAspectRatio: true,
                    scales: {{
                        r: {{
                            beginAtZero: true,
                            max: 5,
                            ticks: {{
                                stepSize: 1
                            }}
                        }}
                    }},
                    plugins: {{
                        legend: {{
                            position: 'top'
                        }}
                    }}
                }}
            }});
        </script>
'''
    else:
        # 没有负责人数据时显示提示信息
        html_content += f'''
        <h4>1.1 负责人{report['leader_name']}的全面反馈分数</h4>
        
        <div class="empty-hint">
            部门负责人全面反馈数据缺失
        </div>
'''
    
    # 添加-1层管理者数据
    if 'minus_one_managers' in report and report['minus_one_managers']:
        html_content += '''
        <h4>1.2 部门-1层管理者的全面反馈分数</h4>
        
        <table>
            <tr>
                <th class="th-blue">管理者</th>
                <th class="th-blue">结果</th>
                <th class="th-blue">总分_BG内排名</th>
                <th class="th-blue">同级打分_BG内排名</th>
                <th class="th-blue">看清楚-总分</th>
                <th class="th-blue">下属愿意跟随</th>
                <th class="th-blue">同级愿意合作</th>
                <th class="th-blue">专业能力胜任</th>
                <th class="th-blue">负面反馈人次</th>
            </tr>
'''
        for mgr in report['minus_one_managers']:
            # 结果列变色
            result_color = ''
            if '红' in mgr['result']:
                result_color = 'color: #ff4d4f; font-weight: bold;'
            elif '黄' in mgr['result']:
                result_color = 'color: #faad14; font-weight: bold;'
            elif '绿' in mgr['result']:
                result_color = 'color: #52c41a; font-weight: bold;'
            
            # 总分排名变色（后10%标红，后1/3标黄）
            total_rank_color = ''
            if mgr['_total_rank_raw'] >= 0.9:
                total_rank_color = 'background: #ffccc7;'
            elif mgr['_total_rank_raw'] >= 0.67:
                total_rank_color = 'background: #fff1b8;'
            
            # 同级打分排名变色（后10%标红，后1/3标黄）
            peer_rank_color = ''
            if mgr['_peer_rank_raw'] >= 0.9:
                peer_rank_color = 'background: #ffccc7;'
            elif mgr['_peer_rank_raw'] >= 0.67:
                peer_rank_color = 'background: #fff1b8;'
            
            # 看清楚变色（低于-0.15标黄）
            look_clear_color = ''
            if mgr['_look_clear_raw'] < -0.15:
                look_clear_color = 'background: #fff1b8;'
            
            # 下属愿意跟随变色（<=80%标黄）
            subordinate_color = ''
            if mgr['_subordinate_raw'] <= 0.8 and mgr['_subordinate_raw'] > 0:
                subordinate_color = 'background: #fff1b8;'
            
            # 同级愿意合作变色（<=80%标黄）
            peer_coop_color = ''
            if mgr['_peer_coop_raw'] <= 0.8 and mgr['_peer_coop_raw'] > 0:
                peer_coop_color = 'background: #fff1b8;'
            
            # 专业能力胜任变色（<=80%标黄）
            competent_color = ''
            if mgr['_competent_raw'] <= 0.8 and mgr['_competent_raw'] > 0:
                competent_color = 'background: #fff1b8;'
            
            # 负面反馈人次变色（>0标黄）
            negative_color = ''
            if mgr['_negative_raw'] > 0:
                negative_color = 'background: #fff1b8;'
            
            # 生成标签HTML
            tags_html = ''
            for tag in mgr.get('tags', []):
                if tag == 'BP点赞':
                    tags_html += '<span style="display:inline-block;background:#e6f7e6;color:#389e0d;font-size:11px;padding:1px 6px;border-radius:3px;margin-top:4px;margin-right:3px;">BP点赞</span>'
                elif tag == 'BP提醒关注':
                    tags_html += '<span style="display:inline-block;background:#fff1f0;color:#cf1322;font-size:11px;padding:1px 6px;border-radius:3px;margin-top:4px;margin-right:3px;">BP提醒关注</span>'
                elif tag == '兼岗':
                    tags_html += '<span style="display:inline-block;background:#f0f0f0;color:#666;font-size:11px;padding:1px 6px;border-radius:3px;margin-top:4px;margin-right:3px;">兼岗</span>'
            
            name_cell = f'{mgr["name"]}'
            if tags_html:
                name_cell += f'<br>{tags_html}'
            
            html_content += f'''
            <tr>
                <td style="vertical-align:middle;">{name_cell}</td>
                <td style="{result_color}">{mgr['result']}</td>
                <td style="{total_rank_color}">{mgr['total_rank']}</td>
                <td style="{peer_rank_color}">{mgr['peer_rank']}</td>
                <td style="{look_clear_color}">{mgr['look_clear_score']}</td>
                <td style="{subordinate_color}">{mgr['subordinate_follow']}</td>
                <td style="{peer_coop_color}">{mgr['peer_cooperate']}</td>
                <td style="{competent_color}">{mgr.get('competent', 'N/A')}</td>
                <td style="{negative_color}">{mgr['negative_feedback']}</td>
            </tr>
'''
        html_content += '''
        </table>
'''
    
    # 添加组织架构图（在2.1.2后面）
    if 'org_chart' in report and report['org_chart']:
        def render_org_node(node, is_root=False):
            """递归渲染组织节点"""
            is_virtual = node.get('is_virtual', False)
            if is_virtual:
                # 虚拟分管节点：虚线框，只显示负责人名
                node_html = f'''
            <div class="org-child-container">
                <div class="org-connector"></div>
                <div class="org-node virtual {node['light_color']}">
                    <div class="org-node-leader">{node['leader']}</div>
                </div>
            '''
            else:
                emp_display = f"{node['employee_count']}人" if node['employee_count'] not in ('', None) else ''
                node_html = f'''
            <div class="org-child-container">
                {'<div class="org-connector"></div>' if not is_root else ''}
                <div class="org-node {node['light_color']}">
                    <div class="org-node-name">{node['org_name']}</div>
                    <div class="org-node-leader">{node['leader']}</div>
                    <div class="org-node-count">{emp_display}</div>
                </div>
            '''
            
            if node.get('children'):
                node_html += '''
                <div class="org-children-wrapper">
                    <div class="org-down-connector"></div>
                    <div class="org-children">'''
                for child in node['children']:
                    node_html += render_org_node(child, False)
                node_html += '''
                    </div>
                </div>'''
            
            node_html += '</div>'
            return node_html
        
        html_content += f'''
        <p class="section-subtitle">组织架构</p>
        <p style="color: #999; font-size: 12px; margin-top: -8px; margin-bottom: 12px;">*组织架构信息为年底</p>
        
        <div class="org-chart-wrapper">
            <div style="margin-bottom: 15px; font-size: 13px; color: #888;">
                <strong>图例说明：</strong>
                <span style="display: inline-block; margin-left: 10px;">
                    <span style="color: #ff4d4f;">●</span> 一级/二级预警
                </span>
                <span style="display: inline-block; margin-left: 10px;">
                    <span style="color: #faad14;">●</span> 三级预警
                </span>
                <span style="display: inline-block; margin-left: 10px;">
                    <span style="color: #52c41a;">●</span> 正常
                </span>
            </div>
            
            <div class="org-chart-controls no-print">
                <button class="zoom-btn" onclick="zoomOut()">🔍− 缩小</button>
                <span class="zoom-level" id="zoomLevel">100%</span>
                <button class="zoom-btn" onclick="zoomIn()">🔍+ 放大</button>
                <button class="zoom-btn" onclick="resetZoom()">↻ 重置</button>
                <button class="zoom-btn" onclick="fitToScreen()">⤢ 适应屏幕</button>
            </div>
            
            <div class="org-chart-container" id="orgChartContainer">
                <div class="org-chart" id="orgChart">
                    {render_org_node(report['org_chart'], True)}
                </div>
            </div>
        </div>
        '''
    
    # ============================================================
    # 2.1.3 全面反馈开放题总结
    # ============================================================
    if 'open_feedback' in report:
        of = report['open_feedback']
        cadre_html = of.get('cadre_html', '')
        
        # 检查是否所有维度都是"缺失"（完全无实质数据）
        _all_missing = False
        if cadre_html:
            import re as _re_check
            _judgments = _re_check.findall(r'(?:共性优势|共性不足|局部待关注|局部优势|缺失)', cadre_html)
            if _judgments and all(j == '缺失' for j in _judgments):
                _all_missing = True
        
        if _all_missing or not cadre_html:
            html_content += '''
        <h4>1.3 开放题总结</h4>
        <div class="empty-hint">
            该部门全面反馈开放题数据缺失
        </div>
'''
        else:
            html_content += '''
        <h4>1.3 开放题总结</h4>
'''
            # 确保cadre div正确关闭：检测cadre_html的div平衡
            _cadre_opens = cadre_html.count('<div')
            _cadre_closes = cadre_html.count('</div>')
            _cadre_fix = '</div>\n' * (_cadre_opens - _cadre_closes) if _cadre_opens > _cadre_closes else ''
            html_content += '<div style="background: #fff; border: 1px solid #e8e8e8; border-radius: 8px; padding: 20px 24px; margin-bottom: 20px; line-height: 1.7;">\n'
            html_content += cadre_html
            html_content += _cadre_fix
            html_content += '\n</div>\n'
    
    # ============================================================
    # 2.2 异动部分
    # ============================================================
    if 'yidong' in report and report['yidong'].get('available'):
        yd = report['yidong']
        
        # 颜色映射
        yd_color_map = {
            'green': '#00B050',
            'yellow': '#FFC000',
            'red': '#C00000',
        }
        rate_color = yd_color_map.get(yd['color'], '#333')
        
        html_content += '''
        <h2>2 异动</h2>
        <p class="note-text">*数据日期：2025年1月1日-12月31日</p>
        
        <h3>2.1 整体异动率（含过程）</h3>
'''
        # 判断文字
        html_content += f'''
        <ul style="margin: 10px 0 10px; padding-left: 20px;">
            <li style="font-size: 14px; color: #333; line-height: 1.8;">{yd['judge_text']}</li>
        </ul>
        <p class="note-text" style="margin: 0 0 15px;">{yd['ref_text']}</p>
'''
        # 2.2.1 表格（本部门+异动率跨行合并，离职行+活水行）
        # 根据颜色设定"本部门"左侧单元格背景色
        yd_label_bg_map = {
            'green': '#4EA72E',
            'yellow': '#ED7D31',
            'red': '#C00000',
        }
        label_bg = yd_label_bg_map.get(yd['color'], '#2E75B6')
        
        html_content += f'''
        <table style="border-collapse: collapse; width: 100%; margin: 15px 0; border: none; table-layout: fixed;">
            <colgroup>
                <col style="width: 8%;">
                <col style="width: 10%;">
                <col style="width: 14%;">
                <col style="width: 6%;">
                <col style="width: 14%;">
                <col style="width: 6%;">
                <col style="width: 14%;">
                <col style="width: 6%;">
            </colgroup>
            <tr>
                <td rowspan="2" style="background: {label_bg}; color: #FFFFFF; font-weight: bold; font-size: 14px; padding: 10px 6px; border: 1px solid #e0e0e0; text-align: center; vertical-align: middle;">本部门</td>
                <td rowspan="2" style="padding: 10px 6px; border: 1px solid #e0e0e0; font-weight: bold; font-size: 20px; color: {rate_color}; text-align: center; vertical-align: middle;">{yd['dept_rate_str']}</td>
                <td style="padding: 6px 8px; border: 1px solid #e0e0e0; color: #888; font-size: 12px; text-align: center;">已离职</td>
                <td style="padding: 6px 4px; border: 1px solid #e0e0e0; color: #555; font-size: 13px; font-weight: 600; text-align: center;">{yd['resigned']}</td>
                <td style="padding: 6px 8px; border: 1px solid #e0e0e0; color: #888; font-size: 12px; text-align: center;">离职流程中</td>
                <td style="padding: 6px 4px; border: 1px solid #e0e0e0; color: #555; font-size: 13px; font-weight: 600; text-align: center;">{yd['resign_process']}</td>
                <td style="padding: 6px 8px; border: 1px solid #e0e0e0; color: #888; font-size: 12px; text-align: center;">离职有意向</td>
                <td style="padding: 6px 4px; border: 1px solid #e0e0e0; color: #555; font-size: 13px; font-weight: 600; text-align: center;">{yd['resign_intent']}</td>
            </tr>
            <tr>
                <td style="padding: 6px 8px; border: 1px solid #e0e0e0; color: #888; font-size: 12px; text-align: center;">已活水</td>
                <td style="padding: 6px 4px; border: 1px solid #e0e0e0; color: #555; font-size: 13px; font-weight: 600; text-align: center;">{yd['transferred']}</td>
                <td style="padding: 6px 8px; border: 1px solid #e0e0e0; color: #888; font-size: 12px; text-align: center;">活水流程中</td>
                <td style="padding: 6px 4px; border: 1px solid #e0e0e0; color: #555; font-size: 13px; font-weight: 600; text-align: center;">{yd['transfer_process']}</td>
                <td style="padding: 6px 8px; border: 1px solid #e0e0e0; color: #888; font-size: 12px; text-align: center;">活水有意向</td>
                <td style="padding: 6px 4px; border: 1px solid #e0e0e0; color: #555; font-size: 13px; font-weight: 600; text-align: center;">{yd['transfer_intent']}</td>
            </tr>
        </table>
'''
        
        # 2.2.2 核心人才（梯队或高绩效）主动异动率
        html_content += '''
        <h3>2.2 核心人才（梯队或高绩效）主动异动率（已发生）</h3>
'''
        if 'core_talent' in report and report['core_talent'].get('available'):
            ct = report['core_talent']
            
            ct_color_map = {
                'green': '#00B050',
                'yellow': '#FFC000',
                'red': '#C00000',
            }
            ct_label_bg_map = {
                'green': '#4EA72E',
                'yellow': '#ED7D31',
                'red': '#C00000',
            }
            ct_rate_color = ct_color_map.get(ct['color'], '#333')
            ct_label_bg = ct_label_bg_map.get(ct['color'], '#2E75B6')
            
            html_content += f'''
        <table style="border-collapse: collapse; width: 100%; margin: 15px 0; border: none; table-layout: fixed;">
            <colgroup>
                <col style="width: 50%;">
                <col style="width: 50%;">
            </colgroup>
            <tr>
                <td style="background: {ct_label_bg}; color: #FFFFFF; font-size: 13px; font-weight: bold; padding: 10px 15px; border: 1px solid #e0e0e0; text-align: center;">本部门核心人才主动异动率</td>
                <td style="background: #F2F2F2; color: #555; font-size: 13px; font-weight: bold; padding: 10px 15px; border: 1px solid #e0e0e0; text-align: center;">本部门非核心人才主动异动率</td>
            </tr>
            <tr>
                <td style="padding: 14px 15px; border: 1px solid #e0e0e0; text-align: center; vertical-align: top;">
                    <span style="font-size: 22px; font-weight: bold; color: {ct_rate_color};">{ct['core_rate_str']}</span>
                    <br><span style="font-size: 11px; color: #999; margin-top: 4px; display: inline-block;">{ct['ref_text']}</span>
                </td>
                <td style="padding: 14px 15px; border: 1px solid #e0e0e0; text-align: center; vertical-align: top;">
                    <span style="font-size: 22px; font-weight: bold; color: #555;">{ct['non_core_rate_str']}</span>
                </td>
            </tr>
        </table>
'''
        else:
            html_content += '''
        <p class="note-text">暂无数据</p>
'''
    
    # ============================================================
    # 2.2.3 离职原因分析
    # ============================================================
    if 'resignation' in report and report['resignation'].get('available'):
        resign = report['resignation']
        pie = resign['pie_data']
        
        html_content += '''
        <h3>2.3 离职原因分析</h3>
'''
        
        # 饼图：离职类型分布
        # 使用CSS绘制圆环饼图
        passive_pct = pie['passive_pct']
        active_pct = pie['active_pct']
        # conic-gradient角度：被动离职从0开始，主动离职接着
        passive_deg = passive_pct * 3.6  # 百分比转角度
        
        html_content += f'''
        <div style="display: flex; justify-content: center; margin: 20px 0;">
            <div style="text-align: center;">
                <div style="font-size: 16px; font-weight: bold; color: #333; margin-bottom: 15px;">离职类型分布</div>
                <div style="display: flex; justify-content: center; gap: 25px; margin-bottom: 15px; font-size: 13px; color: #666;">
                    <span><span style="display: inline-block; width: 12px; height: 12px; background: #C0C0C0; border-radius: 2px; margin-right: 4px; vertical-align: middle;"></span>被动离职</span>
                    <span><span style="display: inline-block; width: 12px; height: 12px; background: #7BA7CC; border-radius: 2px; margin-right: 4px; vertical-align: middle;"></span>主动离职</span>
                </div>
                <div style="position: relative; width: 240px; height: 240px; margin: 0 auto;">
                    <div style="width: 240px; height: 240px; border-radius: 50%; background: conic-gradient(#C0C0C0 0deg {passive_deg}deg, #7BA7CC {passive_deg}deg 360deg);"></div>
                    <div style="position: absolute; top: 50%; left: 50%; transform: translate(-50%, -50%); width: 0px; height: 0px; border-radius: 50%; background: white;"></div>
                    <!-- 百分比标签 -->
                    <div style="position: absolute; top: 0; left: 0; width: 100%; height: 100%;">
'''
        
        # 计算标签位置（在各扇区中心）
        import math as math_mod
        # 被动离职标签位置：在 0 ~ passive_deg 的中心
        if passive_pct > 0:
            passive_mid_deg = passive_deg / 2
            passive_rad = math_mod.radians(passive_mid_deg - 90)  # CSS从12点方向开始
            passive_x = 50 + 30 * math_mod.cos(passive_rad)
            passive_y = 50 + 30 * math_mod.sin(passive_rad)
            html_content += f'''                        <span style="position: absolute; top: {passive_y}%; left: {passive_x}%; transform: translate(-50%, -50%); font-size: 16px; font-weight: bold; color: #555;">{passive_pct}%</span>
'''
        
        # 主动离职标签位置：在 passive_deg ~ 360 的中心
        if active_pct > 0:
            active_mid_deg = passive_deg + (360 - passive_deg) / 2
            active_rad = math_mod.radians(active_mid_deg - 90)
            active_x = 50 + 30 * math_mod.cos(active_rad)
            active_y = 50 + 30 * math_mod.sin(active_rad)
            html_content += f'''                        <span style="position: absolute; top: {active_y}%; left: {active_x}%; transform: translate(-50%, -50%); font-size: 16px; font-weight: bold; color: #fff;">{active_pct}%</span>
'''
        
        html_content += f'''                    </div>
                </div>
                <div style="margin-top: 10px; font-size: 13px; color: #888;">共 {pie['total']} 人（被动离职 {pie['passive']} 人，主动离职 {pie['active']} 人）</div>
            </div>
        </div>
'''
        
        # 离职原因分析表格
        if resign.get('has_org_reasons') and resign.get('reason_analysis'):
            html_content += '''
        <table>
            <tr>
                <th colspan="3" style="background: #4472C4; color: white; text-align: left; padding: 10px 15px; font-size: 14px;">离职原因分析（与组织相关的）</th>
            </tr>
            <tr>
                <th style="background: #D6E4F0; color: #333; width: 15%;">归类</th>
                <th style="background: #D6E4F0; color: #333; width: 55%; text-align: left;">原文</th>
                <th style="background: #D6E4F0; color: #333; width: 30%;">离职人员信息</th>
            </tr>
'''
            for item in resign['reason_analysis']:
                # 归类列
                category = item['category']
                
                # 原文列：总结 + 原文引用
                quotes_html = f"<b>总结：{item['summary']}</b>"
                if item['quotes']:
                    quotes_html += "<br/><br/>原文举例："
                    for qi, q in enumerate(item['quotes'], 1):
                        # 截断过长的原文
                        q_display = q if len(q) <= 200 else q[:200] + "..."
                        quotes_html += f"<br/>{qi}. \"{q_display}\""
                
                # 人员信息列
                stats_html = item['people_stats'].replace('\n', '<br/>')
                
                html_content += f'''            <tr>
                <td style="padding: 10px 12px; border: 1px solid #e0e0e0; vertical-align: top; font-weight: bold; color: #333;">{category}</td>
                <td style="padding: 10px 12px; border: 1px solid #e0e0e0; vertical-align: top; font-size: 13px; color: #444; line-height: 1.6; text-align: left;">{quotes_html}</td>
                <td style="padding: 10px 12px; border: 1px solid #e0e0e0; vertical-align: top; font-size: 12px; color: #666; line-height: 1.8;">{stats_html}</td>
            </tr>
'''
            
            html_content += '''        </table>
'''
        else:
            html_content += '''
        <div style="background: #f8f9fa; border: 1px solid #e0e0e0; border-radius: 6px; padding: 20px; text-align: center; margin: 15px 0; color: #999; font-size: 14px;">
            暂无与组织相关的离职原因
        </div>
'''
    
    # ============================================================
    # 2.3 敬满部分
    # ============================================================
    if 'jingman' in report and report['jingman'].get('available'):
        jm = report['jingman']
        
        html_content += '''
        <h2>3 敬满</h2>
        
        <h3>3.1 总分与定位</h3>
        <p class="note-text">以下数据反映本部门在所属BG内的敬业度及满意度排名情况。</p>
'''
        
        analysis = jm.get('analysis', {})
        
        # 核心维度概览 小标题
        # 核心维度概览表
        if jm.get('core_table') and len(jm['core_table']) > 0:
            html_content += '''
        <p class="section-subtitle">核心维度概览</p>
        <table>
            <tr>
                <th class="th-blue">指标</th>
                <th class="th-blue">在BG的排名</th>
                <th class="th-blue">排名变化</th>
                <th class="th-blue">分值</th>
                <th class="th-blue">较去年变化</th>
            </tr>
'''
            for cr in jm['core_table']:
                row_bg = '#EBF3FB' if cr['metric'] == '敬业度' else '#FFFFFF'
                # 排名变色（基于击败率百分比）
                rank_style = ''
                raw_bg = cr.get('_rank_raw_bg')
                if raw_bg is not None:
                    if raw_bg >= 80:
                        rank_style = 'color: #389e0d; font-weight: bold;'
                    elif raw_bg <= 20:
                        rank_style = 'color: #ff4d4f; font-weight: bold;'
                else:
                    # 兼容旧的分位值格式
                    rv = cr.get('_rank_raw')
                    if rv is not None:
                        if rv >= 0.80:
                            rank_style = 'color: #389e0d; font-weight: bold;'
                        elif rv <= 0.20:
                            rank_style = 'color: #ff4d4f; font-weight: bold;'
                
                html_content += f'''
            <tr>
                <td style="background: {row_bg}; font-weight: bold;">{cr['metric']}</td>
                <td style="background: {row_bg}; text-align: center; {rank_style}">{cr['in_bg_rank']}</td>
                <td style="background: {row_bg}; text-align: center;">{cr['rank_change']}</td>
                <td style="background: {row_bg}; text-align: center;">{cr['score']}</td>
                <td style="background: {row_bg}; text-align: center;">{cr['yoy_change']}</td>
            </tr>
'''
            html_content += '''
        </table>
'''
        
        # 风险区间分布
        if jm.get('risk_table'):
            risk_table = jm['risk_table']
            has_gb_detail = any(isinstance(rr.get('count_gb'), int) for rr in risk_table)
            
            html_content += '''
        <p class="section-subtitle">风险区间分布</p>
'''
            if has_gb_detail:
                # 5列：风险区间 / 题数 / 占比 / Great Boss / 其他
                html_content += '''
        <table>
            <tr>
                <th class="th-blue">风险区间</th>
                <th class="th-blue">题数</th>
                <th class="th-blue">占比</th>
                <th class="th-blue">Great Boss</th>
                <th class="th-blue">其他</th>
            </tr>
'''
                for rr in risk_table:
                    count_str = rr.get('count_str', str(rr['count']))
                    gb_str = rr.get('gb_str', '/')
                    other_str = rr.get('other_str', '/')
                    pct_str = rr.get('pct', '/')
                    
                    html_content += f'''
            <tr>
                <td style="font-weight: bold;">{rr['zone']}</td>
                <td style="text-align: center;">{count_str}</td>
                <td style="text-align: center;">{pct_str}</td>
                <td style="text-align: center;">{gb_str}</td>
                <td style="text-align: center;">{other_str}</td>
            </tr>
'''
                html_content += '''
        </table>
'''
            else:
                # 旧版：只有 题数 / 占比
                html_content += '''
        <table>
            <tr>
                <th class="th-blue">风险区间</th>
                <th class="th-blue">题数</th>
                <th class="th-blue">占比</th>
            </tr>
'''
                for rr in risk_table:
                    html_content += f'''
            <tr>
                <td style="font-weight: bold;">{rr['zone']}</td>
                <td style="text-align: center;">{rr['count']}</td>
                <td style="text-align: center;">{rr['pct']}</td>
            </tr>
'''
                html_content += '''
        </table>
'''
        
        # === 47道题柱状图（单向，0-100，CSS渲染） ===
        if jm.get('chart_data'):
            html_content += '''
        <p class="section-subtitle">47道敬满子题 BG内击败率分布</p>
'''
            chart_items = jm['chart_data']
            summary = jm.get('chart_summary', {})
            n_front = summary.get('n_front', 0)
            n_total = summary.get('n_total', 47)
            n_sat = summary.get('n_sat', 43)
            n_eng = summary.get('n_eng', 4)
            bg_min = summary.get('bg_min', 0)
            
            chart_height = 400  # px
            
            # 动态标题
            if n_front == n_total:
                chart_title = f"该部门{n_total}道敬满题得分全部在BG排名靠前"
            elif n_front > n_total * 0.7:
                chart_title = f"该部门{n_total}道敬满子题{n_front}题排名靠前"
            elif n_front <= 10:
                chart_title = f"该部门{n_total}道敬满子题仅{n_front}题排名靠前"
            else:
                chart_title = f"该部门{n_total}道敬满子题有{n_front}题排名靠前"
            
            html_content += f'''
        <!-- 标题（图表上方） -->
        <p style="text-align:center; font-size:16px; color:#1F3864; font-weight:bold; margin:30px 0 5px;">
            {chart_title}
        </p>
        
        <div style="position:relative; margin: 5px 0 20px; padding-left: 45px; padding-right: 20px;">
            <!-- 图例（右上角） -->
            <div style="position:absolute; right:20px; top:0; z-index:10; background:rgba(255,255,255,0.9); border:1px solid #ddd; border-radius:4px; padding:6px 10px; font-size:12px; line-height:1.8;">
                <div><span style="display:inline-block; width:14px; height:10px; background:#4472C4; margin-right:5px; vertical-align:middle;"></span> 排名靠前（击败率 ≥ 50%）</div>
                <div><span style="display:inline-block; width:14px; height:10px; background:#F4A460; margin-right:5px; vertical-align:middle;"></span> 排名靠后（20% &lt; 击败率 &lt; 50%）</div>
                <div><span style="display:inline-block; width:14px; height:10px; background:#C00000; margin-right:5px; vertical-align:middle;"></span> 高危（击败率 ≤ 20%）</div>
            </div>
            
            <!-- 图表区域（单向：底部0，顶部100） -->
            <div style="position:relative; height:{chart_height}px; border-left:2px solid #999; border-bottom:2px solid #999;">
                
                <!-- Y轴刻度 -->
'''
            for yval in [0, 10, 20, 30, 40, 50, 60, 70, 80, 90, 100]:
                y_pos = (yval / 100) * chart_height  # 从底部算起的像素
                html_content += f'                <div style="position:absolute; left:-40px; bottom:{y_pos}px; width:35px; text-align:right; font-size:11px; color:#666; transform:translateY(50%);">{yval}</div>\n'
                # 网格线
                if yval > 0 and yval < 100:
                    line_color = '#1F3864' if yval == 50 else '#e8e8e8'
                    line_width = '2px' if yval == 50 else '1px'
                    html_content += f'                <div style="position:absolute; left:0; right:0; bottom:{y_pos}px; height:{line_width}; background:{line_color}; z-index:2;"></div>\n'
            
            # 50%线右侧延伸标注
            y50_pos = int(50 / 100 * chart_height)
            
            # 红线：倒数20%、倒数10%
            y20_pos = int(20 / 100 * chart_height)
            y10_pos = int(10 / 100 * chart_height)
            bg_min_pos = int(bg_min / 100 * chart_height)
            
            html_content += f'''
                <!-- 倒数20%红线 -->
                <div style="position:absolute; left:0; right:0; bottom:{y20_pos}px; height:2px; background:#FF0000; z-index:5;">
                    <span style="position:absolute; left:0; top:-16px; font-size:11px; color:#FF0000; font-weight:bold; font-style:italic;">倒数20%</span>
                </div>
                
                <!-- 倒数10%红线 -->
                <div style="position:absolute; left:0; right:0; bottom:{y10_pos}px; height:2px; background:#FF0000; z-index:5;">
                    <span style="position:absolute; left:0; top:-16px; font-size:11px; color:#FF0000; font-weight:bold; font-style:italic;">倒数10%</span>
                </div>
                
                <!-- 柱子容器 -->
                <div style="display:flex; align-items:stretch; height:100%; padding:0 2px; gap:2px;">
'''
            for it in chart_items:
                fav_bg = it['fav_bg']
                if fav_bg >= 50:
                    color = '#4472C4'
                elif fav_bg > 20:
                    color = '#F4A460'
                else:
                    color = '#C00000'
                
                tooltip = f"{it['short']}: 击败率{fav_bg:.0f}%"
                
                if fav_bg >= 50:
                    # 蓝色柱子：从50%线向上到 fav_bg% 的位置
                    # bottom = 50%（50%线位置，从底部算）
                    # top = (100 - fav_bg)%（从顶部算）
                    bar_top_pct = 100 - fav_bg
                    bar_style = f"position:absolute; top:{bar_top_pct}%; bottom:50%; left:0; right:0; background:{color}; border-radius:2px 2px 0 0;"
                else:
                    # 黄色/红色柱子：从50%线向下垂落到 fav_bg% 的位置
                    # top = 50%（50%线位置，从顶部算）
                    # bottom = fav_bg%（从底部算）
                    bar_bottom_pct = max(fav_bg, 0.5)
                    bar_style = f"position:absolute; top:50%; bottom:{bar_bottom_pct}%; left:0; right:0; background:{color}; border-radius:0 0 2px 2px;"
                
                html_content += f'                    <div title="{tooltip}" style="flex:1; min-width:4px; position:relative;"><div style="{bar_style}"></div></div>\n'
            
            html_content += f'''
                </div>
            </div>
            
            <!-- 底部标注 -->
            <p style="text-align:center; font-size:13px; color:#1F3864; margin-top:8px; font-weight:bold;">
                {n_sat}道满意度题 + {n_eng}道敬业度题
            </p>
        </div>
'''
        
        # === 末10%题目明细表 ===
        if jm.get('bottom_table') and len(jm['bottom_table']) > 0:
            html_content += f'''
        <p class="section-subtitle">BG末10%题目明细</p>
        
        <table>
            <tr>
                <th class="th-red" style="width:10%;">标题</th>
                <th class="th-red">题目</th>
                <th class="th-red" style="width:140px; white-space:nowrap;">BG内排名（倒数）</th>
                <th class="th-red" style="width:120px;">分值变化</th>
            </tr>
'''
            for bt in jm['bottom_table']:
                # 击败率颜色
                bg_pct = bt['fav_bg']
                if bg_pct <= 5:
                    rank_style = 'color: #C00000; font-weight: bold;'
                else:
                    rank_style = 'color: #E65100; font-weight: bold;'
                
                html_content += f'''
            <tr>
                <td style="font-weight: bold; white-space: nowrap;">{bt['short']}</td>
                <td style="font-size: 13px;">{bt['full']}</td>
                <td style="text-align: center; {rank_style}">{bt['bg_rank_str']}</td>
                <td style="text-align: center;">{bt['yoy_change']}</td>
            </tr>
'''
            html_content += '''
        </table>
'''
        
        # === 2.3.2 值得关注的细分项 ===
        if jm.get('subdiv_table') and len(jm['subdiv_table']) > 0:
            html_content += '''
        <h3>3.2 值得关注的细分项</h3>
        <p class="note-text">
            以下题目在部门排名与BG排名、或增幅之间存在显著差异（差值 &gt; 10）。
        </p>
'''
            # 单表 + 分组行（合并单元格）
            html_content += '''
        <table>
            <tr>
                <th class="th-blue">标题</th>
                <th class="th-blue">题目</th>
                <th class="th-blue">部门排名</th>
                <th class="th-blue">BG排名</th>
                <th class="th-blue">排名差</th>
                <th class="th-blue">增幅差</th>
            </tr>
'''
            current_group = None
            group_colors = {
                'bg_higher': ('#FFF2CC', '#E65100'),
                'dept_higher': ('#E8F5E9', '#2E7D32'),
                'dept_lag': ('#FCE4D6', '#C00000'),
                'dept_lead': ('#E3F2FD', '#1565C0'),
            }
            
            for sd in jm['subdiv_table']:
                if sd['group'] != current_group:
                    current_group = sd['group']
                    g_bg, g_color = group_colors.get(current_group, ('#F5F5F5', '#333'))
                    
                    html_content += f'''
            <tr>
                <td colspan="6" style="background: {g_bg}; color: {g_color}; font-weight: bold; font-size: 13px; padding: 8px 10px;">{sd['group_label']}</td>
            </tr>
'''
                rank_diff_val = sd['rank_diff']
                rd_style = ''
                if rank_diff_val > 10:
                    rd_style = 'color: #C00000; font-weight: bold;'
                elif rank_diff_val < -10:
                    rd_style = 'color: #2E7D32; font-weight: bold;'
                
                html_content += f'''
            <tr>
                <td style="font-weight: bold; white-space: nowrap;">{sd['short']}</td>
                <td style="font-size: 13px;">{sd['full'][:50]}</td>
                <td style="text-align: center;">{sd['dept_rank']}</td>
                <td style="text-align: center;">{sd['bg_rank']}</td>
                <td style="text-align: center; {rd_style}">{rank_diff_val:+d}</td>
                <td style="text-align: center;">{sd['growth_diff_str']}</td>
            </tr>
'''
            html_content += '''
        </table>
'''
        
        # 2.3.3 敬满开放题
        html_content += '''
        <h3>3.3 敬满开放题</h3>
'''
        if 'jingman_open_html' in report:
            html_content += report['jingman_open_html']
        else:
            html_content += '''
        <p class="note-text">（该部门暂无敬满开放题分析数据）</p>
'''
        # ─── 词云（嵌入到 2.3.3 底部） ───
        if 'wordcloud_json' in report:
            wc_json_str = report['wordcloud_json']
            wc_dept_safe = report.get('wordcloud_dept', '').replace('"', '&quot;').replace("'", '&#39;')
            html_content += f'''
        <div class="wc-section">
            <div class="wc-section-title">{wc_dept_safe} · 敬满开放题关键词云</div>
            <div class="wc-section-subtitle">字号 = 词频 &nbsp;&middot;&nbsp; 颜色深浅 = 词频强度 &nbsp;&middot;&nbsp; 色系 = 情感倾向</div>
            <div class="wc-cloud-container" id="wcCloudBox"></div>
            <div id="wcTooltip" class="wc-tooltip"></div>
            <div class="wc-legend">
                <div class="wc-legend-item"><div class="wc-legend-bar positive"></div>正向词</div>
                <div class="wc-legend-item"><div class="wc-legend-bar neutral"></div>中性词</div>
                <div class="wc-legend-item"><div class="wc-legend-bar negative"></div>负向词</div>
            </div>
        </div>
        <script>
        (function() {{
            var wcData = {wc_json_str};
            function lerpColor(c1, c2, t) {{
                return [
                    Math.round(c1[0] + (c2[0] - c1[0]) * t),
                    Math.round(c1[1] + (c2[1] - c1[1]) * t),
                    Math.round(c1[2] + (c2[2] - c1[2]) * t),
                ];
            }}
            function sentimentColor(sentiment, freq, maxF, minF) {{
                var logMax = Math.log(maxF), logMin = Math.log(minF);
                var t = logMax === logMin ? 0.5 : (Math.log(freq) - logMin) / (logMax - logMin);
                var rgb;
                if (sentiment === 'positive') {{
                    rgb = lerpColor([183, 235, 143], [35, 120, 4], t);
                }} else if (sentiment === 'negative') {{
                    rgb = lerpColor([255, 163, 158], [168, 7, 26], t);
                }} else {{
                    rgb = lerpColor([255, 229, 143], [173, 104, 0], t);
                }}
                return 'rgb(' + rgb[0] + ',' + rgb[1] + ',' + rgb[2] + ')';
            }}
            var sentimentLabel = {{ positive: '正向', negative: '负向', neutral: '中性' }};
            var maxFreq = Math.max.apply(null, wcData.map(function(w){{ return w.freq; }}));
            var minFreq = Math.min.apply(null, wcData.map(function(w){{ return w.freq; }}));
            function fontSize(freq) {{
                var logMax = Math.log(maxFreq), logMin = Math.log(minFreq);
                var t = logMax === logMin ? 0.5 : (Math.log(freq) - logMin) / (logMax - logMin);
                return 13 + t * 44;
            }}
            var box = document.getElementById('wcCloudBox');
            var W = box.clientWidth, H = box.clientHeight;
            window._wcRenderWidth = W;
            var cx = W / 2, cy = H / 2;
            var _mc = document.createElement('canvas').getContext('2d');
            wcData.sort(function(a, b) {{ return b.freq - a.freq; }});
            var placed = [];
            function measure(text, size) {{
                _mc.font = '700 ' + size + 'px "PingFang SC","Microsoft YaHei",sans-serif';
                return {{ w: _mc.measureText(text).width + 4, h: size * 1.25 + 2 }};
            }}
            function collides(r) {{
                if (r.x < 2 || r.y < 2 || r.x + r.w > W - 2 || r.y + r.h > H - 2) return true;
                for (var i = 0; i < placed.length; i++) {{
                    var p = placed[i];
                    if (!(r.x >= p.x + p.w || r.x + r.w <= p.x || r.y >= p.y + p.h || r.y + r.h <= p.y)) return true;
                }}
                return false;
            }}
            for (var wi = 0; wi < wcData.length; wi++) {{
                var word = wcData[wi];
                var size = fontSize(word.freq);
                var m = measure(word.text, size);
                var ok = false;
                var a = Math.random() * Math.PI * 2;
                for (var i = 0; i < 12000 && !ok; i++) {{
                    var r = i * 0.04;
                    var x = cx + r * Math.cos(a) - m.w / 2;
                    var y = cy + r * Math.sin(a) - m.h / 2;
                    a += 0.12;
                    var rect = {{ x: x, y: y, w: m.w, h: m.h }};
                    if (!collides(rect)) {{
                        placed.push(rect);
                        var el = document.createElement('span');
                        el.className = 'wc-word';
                        el.textContent = word.text;
                        el.style.cssText = 'left:' + x + 'px;top:' + y + 'px;font-size:' + size + 'px;color:' + sentimentColor(word.sentiment, word.freq, maxFreq, minFreq) + ';';
                        el.setAttribute('data-freq', word.freq);
                        el.setAttribute('data-sentiment', sentimentLabel[word.sentiment]);
                        box.appendChild(el);
                        ok = true;
                    }}
                }}
            }}
            // ── 自适应高度：计算实际 bounding box 并收缩容器 ──
            if (placed.length > 0) {{
                var minY = Infinity, maxY = -Infinity;
                for (var pi = 0; pi < placed.length; pi++) {{
                    if (placed[pi].y < minY) minY = placed[pi].y;
                    if (placed[pi].y + placed[pi].h > maxY) maxY = placed[pi].y + placed[pi].h;
                }}
                var contentH = maxY - minY;
                var pad = 28;  // 上下留白
                var newH = Math.max(contentH + pad * 2, 160);
                // 垂直偏移：把所有词整体移到新容器正中
                var offsetY = pad - minY + (newH - pad * 2 - contentH) / 2;
                var spans = box.querySelectorAll('.wc-word');
                for (var si = 0; si < spans.length; si++) {{
                    var curTop = parseFloat(spans[si].style.top);
                    spans[si].style.top = (curTop + offsetY) + 'px';
                }}
                box.style.height = newH + 'px';
                window._wcRenderHeight = newH;
            }}
            var tip = document.getElementById('wcTooltip');
            box.addEventListener('mousemove', function(e) {{
                var t = e.target.closest('.wc-word');
                if (t) {{
                    tip.textContent = t.textContent + '  ×' + t.getAttribute('data-freq') + '  ' + t.getAttribute('data-sentiment');
                    tip.style.display = 'block';
                    tip.style.left = (e.clientX + 12) + 'px';
                    tip.style.top  = (e.clientY - 30) + 'px';
                }} else {{
                    tip.style.display = 'none';
                }}
            }});
            box.addEventListener('mouseleave', function() {{ tip.style.display = 'none'; }});
        }})();
        </script>
'''
    else:
        # 无敬满数据时的提示
        html_content += '''
        <h2>3 敬满</h2>
        <div class="empty-hint">
            该部门敬满数据暂无或无法匹配
        </div>
'''
    
    html_content += f'''
        <div class="footer">
            <p>本报告由组织诊断报告生成系统自动生成</p>
            <p>生成时间：{datetime.now().strftime("%Y-%m-%d %H:%M:%S")}</p>
        </div>
    </div>
    
    <script>
        // 组织架构图缩放控制
        let currentZoom = 1.0;
        const zoomStep = 0.1;
        const minZoom = 0.3;
        const maxZoom = 2.0;
        
        function updateZoom(zoom) {{
            currentZoom = Math.max(minZoom, Math.min(maxZoom, zoom));
            const orgChart = document.getElementById('orgChart');
            const container = document.getElementById('orgChartContainer');
            if (orgChart) {{
                orgChart.style.transform = `scale(${{currentZoom}})`;
                document.getElementById('zoomLevel').textContent = Math.round(currentZoom * 100) + '%';
                // 自适应容器高度
                requestAnimationFrame(function() {{
                    if (container) {{
                        const chartHeight = orgChart.scrollHeight;
                        container.style.height = Math.ceil(chartHeight * currentZoom + 40) + 'px';
                    }}
                }});
            }}
        }}
        
        function zoomIn() {{
            updateZoom(currentZoom + zoomStep);
        }}
        
        function zoomOut() {{
            updateZoom(currentZoom - zoomStep);
        }}
        
        function resetZoom() {{
            updateZoom(1.0);
        }}
        
        function fitToScreen() {{
            const container = document.getElementById('orgChartContainer');
            const orgChart = document.getElementById('orgChart');
            
            if (container && orgChart) {{
                // 暂时重置缩放以获取真实尺寸
                orgChart.style.transform = 'scale(1)';
                
                const containerWidth = container.clientWidth;
                const chartWidth = orgChart.scrollWidth;
                
                // 仅按宽度计算缩放比例，高度自适应
                const scale = Math.min(containerWidth / (chartWidth + 80), 1.0);
                
                updateZoom(scale);
                
                // 根据缩放后的实际高度设置容器高度，避免空白
                requestAnimationFrame(function() {{
                    const chartHeight = orgChart.scrollHeight;
                    container.style.height = Math.ceil(chartHeight * currentZoom + 40) + 'px';
                }});
            }}
        }}
        
        // 页面加载时自动适应屏幕
        window.addEventListener('load', function() {{
            setTimeout(fitToScreen, 100);
        }});
        
        // 打印时自动缩放组织架构图适应页面，打印后恢复
        let savedZoom = 1.0;
        // 保存样式用于恢复
        let savedWrapperCss = '';
        let savedChartCss = '';
        let savedContainerCss = '';
        
        // 记录组织架构图在屏幕上渲染时的实际自然宽度（无缩放时）
        let orgChartNaturalWidth = 0;
        (function() {{
            const orgChart = document.getElementById('orgChart');
            if (orgChart) {{
                // 临时去掉transform以测量真实宽度
                const saved = orgChart.style.transform;
                orgChart.style.transform = 'scale(1)';
                orgChartNaturalWidth = orgChart.scrollWidth;
                orgChart.style.transform = saved;
                console.log('[OrgChart] naturalWidth =', orgChartNaturalWidth);
            }}
        }})();
        
        window.addEventListener('beforeprint', function() {{
            // 幂等保护：如果已经在打印模式，不要重复处理
            if (window._isPrintMode) {{
                console.log('[Print] beforeprint skipped (already in print mode)');
                return;
            }}
            window._isPrintMode = true;
            
            savedZoom = currentZoom;
            const orgChart = document.getElementById('orgChart');
            const container = document.getElementById('orgChartContainer');
            const wrapper = document.querySelector('.org-chart-wrapper');
            
            // 保存样式（只在第一次 beforeprint 时保存）
            savedWrapperCss = wrapper ? wrapper.style.cssText || '' : '';
            savedChartCss = orgChart ? orgChart.style.cssText || '' : '';
            savedContainerCss = container ? container.style.cssText || '' : '';
            
            if (orgChart) {{
                orgChart.style.transform = 'none';
                orgChart.style.transformOrigin = '';
                orgChart.style.zoom = '';
            }}
            if (container) {{
                container.style.maxHeight = 'none';
                container.style.overflow = 'visible';
                container.style.height = 'auto';
                container.style.border = 'none';
            }}
            
            if (wrapper && orgChartNaturalWidth > 0) {{
                // 测量wrapper在打印布局下的实际可用宽度
                var wrapperWidth = wrapper.offsetWidth;
                console.log('[OrgChart] print wrapper offsetWidth =', wrapperWidth, 'chart natural =', orgChartNaturalWidth);
                
                // 如果读到的宽度>800px，说明可能读到了屏幕宽度而非打印宽度
                // 此时用保守的A4可用宽度：
                // A4@96dpi = 794px, 减body padding 20px, 减container padding 40px = 734px
                if (wrapperWidth > 800) {{
                    wrapperWidth = 734;
                    console.log('[OrgChart] fallback to A4 width =', wrapperWidth);
                }}
                
                if (orgChartNaturalWidth > wrapperWidth && wrapperWidth > 0) {{
                    // 将zoom应用在orgChartContainer上
                    // zoom改变布局尺寸，container缩小后不会溢出wrapper
                    var zoomVal = wrapperWidth / orgChartNaturalWidth;
                    if (container) {{
                        container.style.zoom = zoomVal;
                        container.style.overflow = 'visible';
                    }}
                    console.log('[OrgChart] applied zoom on CONTAINER =', zoomVal);
                }}
            }}
            
            // 词云：用 zoom 缩放（与组织架构图相同策略）
            // zoom 同时改变视觉大小和布局尺寸，不会溢出
            var wcRenderWidth = window._wcRenderWidth || 0;
            var wcBox = document.getElementById('wcCloudBox');
            if (wcBox && wcRenderWidth > 0) {{
                // 保存原始样式
                window._wcSavedStyle = wcBox.style.cssText || '';
                // 打印时 wcBox 的可用宽度（受 A4 纸宽限制）
                // 用 wc-section 父容器的宽度，它受 .container padding 约束
                var wcSection = wcBox.closest('.wc-section');
                var availWidth = wcSection ? wcSection.offsetWidth : 734;
                // 如果拿到的还是屏幕宽度（>800），用保守的 A4 宽度
                if (availWidth > 800) availWidth = 734;
                
                if (wcRenderWidth > availWidth) {{
                    var wcZoom = availWidth / wcRenderWidth;
                    wcBox.style.zoom = wcZoom;
                    console.log('[WordCloud] applied zoom =', wcZoom, 'renderW=', wcRenderWidth, 'availW=', availWidth);
                }}
            }}
        }});
        window.addEventListener('afterprint', function() {{
            // 重置打印模式标记
            window._isPrintMode = false;
            
            const orgChart = document.getElementById('orgChart');
            const container = document.getElementById('orgChartContainer');
            const wrapper = document.querySelector('.org-chart-wrapper');
            
            // 恢复所有样式
            if (orgChart) orgChart.style.cssText = savedChartCss;
            if (container) container.style.cssText = savedContainerCss;
            if (wrapper) wrapper.style.cssText = savedWrapperCss;
            
            updateZoom(savedZoom);
            // 恢复词云样式
            var wcBox = document.getElementById('wcCloudBox');
            if (wcBox && window._wcSavedStyle !== undefined) {{
                wcBox.style.cssText = window._wcSavedStyle;
                window._wcSavedStyle = undefined;
            }}
        }});
    </script>
</body>
</html>
'''
    
    # 写入文件
    with open(output_path, 'w', encoding='utf-8') as f:
        f.write(html_content)


def _load_all_data():
    """加载所有报告生成所需的数据源，返回字典"""
    print("\n正在加载数据...")
    
    diag_reader = ExcelReader('【组织诊断结果】/组织诊断2025全年最终版.xlsx', '亮灯明细')
    print(f"  ✓ 组织诊断: {len(diag_reader.rows)} 行")
    
    feedback_reader = ExcelReader('【全面反馈】/全面反馈25H2v2.xlsx')
    print(f"  ✓ 全面反馈: {len(feedback_reader.rows)} 行")
    
    jm_loader = None
    try:
        jm_loader = JingmanDataLoader('【敬满】/敬满总分相关指标.xlsx')
        print(f"  ✓ 敬满基础数据")
    except Exception as e:
        print(f"  ⚠ 敬满基础: {e}")
    
    jm_detail_loader = None
    jm_detail_dir = '【敬满】/report_tool_final/data'
    jm_dept_file = os.path.join(jm_detail_dir, '全量敬满数据.xlsx')
    jm_bg_file = os.path.join(jm_detail_dir, 'BG相关数据.xlsx')
    jm_var_file = os.path.join(jm_detail_dir, '题目与标题对照表.xlsx')
    if all(os.path.exists(f) for f in [jm_dept_file, jm_bg_file, jm_var_file]):
        try:
            jm_detail_loader = JingmanDetailLoader(jm_dept_file, jm_bg_file, jm_var_file)
            print(f"  ✓ 敬满详情 (report_tool_final)")
        except Exception as e:
            print(f"  ⚠ 敬满详情: {e}")
    
    yd_loader = None
    yd_rate_file = '【异动】/2025全年高异动部门级数据.xlsx'
    yd_otd_file = '【异动】/部门&中心结论导出-OTD.xlsx'
    yd_core_file = '【异动】/A2026011210130001_修复版v0.4_交付.xlsx'
    if os.path.exists(yd_rate_file) and os.path.exists(yd_otd_file):
        try:
            core_path = yd_core_file if os.path.exists(yd_core_file) else None
            yd_loader = YidongDataLoader(yd_rate_file, yd_otd_file, core_path)
            print(f"  ✓ 异动数据")
        except Exception as e:
            print(f"  ⚠ 异动: {e}")
    
    bp_loader = None
    try:
        bp_loader = BPObservationLoader('【全面反馈】')
        print(f"  ✓ BP观察")
    except:
        pass
    
    jiangan_loader = None
    try:
        jiangan_loader = JianGangLoader('【组织架构信息】/岗位信息表.xlsx')
        print(f"  ✓ 岗位信息")
    except:
        pass
    
    open_fb_loader = None
    try:
        open_fb_loader = OpenFeedbackLoader('【全面反馈】')
        print(f"  ✓ 全面反馈开放题")
    except:
        pass
    
    jm_open_loader = None
    try:
        jm_open_loader = JingmanOpenLoader('【敬满】')
        print(f"  ✓ 敬满开放题分析")
    except:
        pass
    
    wc_loader = None
    try:
        wc_loader = WordCloudLoader('【敬满】')
        print(f"  ✓ 词云关键词")
    except:
        pass
    
    return {
        'diag_reader': diag_reader,
        'feedback_reader': feedback_reader,
        'jm_loader': jm_loader,
        'jm_detail_loader': jm_detail_loader,
        'yd_loader': yd_loader,
        'bp_loader': bp_loader,
        'jiangan_loader': jiangan_loader,
        'open_fb_loader': open_fb_loader,
        'jm_open_loader': jm_open_loader,
        'wc_loader': wc_loader,
    }


# ── BG 配置 ──
BG_CONFIG = {
    'CDG':  {'output': '报告/cdg/html',  'match_col': 1, 'match': 'CDG'},
    'IEG':  {'output': '报告/ieg/html',  'match_col': 1, 'match': 'IEG'},
    'PCG':  {'output': '报告/pcg/html',  'match_col': 1, 'match': 'PCG'},
    'TEG':  {'output': '报告/teg/html',  'match_col': 1, 'match': 'TEG'},
    'WXG':  {'output': '报告/wxg/html',  'match_col': 1, 'match': 'WXG'},
    'CSIG': {'output': '报告/csig/html', 'match_col': 1, 'match': 'CSIG'},
    'S1':   {'output': '报告/s1/html',   'match_col': 1, 'match': 'S1'},
    'S2':   {'output': '报告/s2/html',   'match_col': 1, 'match': 'S2'},
    'S3':   {'output': '报告/s3/html',   'match_col': 1, 'match': 'S3'},
    'OFS':  {'output': '报告/ofs/html',  'match_col': 1, 'match': 'Overseas'},
}


def batch_generate(bg_list, data):
    """批量生成指定BG的全部报告"""
    diag_reader = data['diag_reader']
    feedback_reader = data['feedback_reader']
    
    for bg_key in bg_list:
        bg_key_upper = bg_key.upper()
        if bg_key_upper not in BG_CONFIG:
            print(f"⚠ 未知BG: {bg_key}，可选: {', '.join(BG_CONFIG.keys())}")
            continue
        
        cfg = BG_CONFIG[bg_key_upper]
        output_dir = cfg['output']
        match_str = cfg['match']
        match_col = cfg['match_col']
        os.makedirs(output_dir, exist_ok=True)
        
        # 筛选部门
        rows = []
        for i, row in enumerate(diag_reader.rows):
            if i == 0:
                continue
            col_val = row[match_col].strip() if len(row) > match_col and row[match_col] else ''
            if match_str in col_val:
                rows.append((i, row))
        
        print(f"\n{'='*60}")
        print(f"  批量生成 {bg_key_upper} - {len(rows)} 个部门")
        print(f"{'='*60}\n")
        
        success = 0
        fail = 0
        
        for row_idx, row in rows:
            org_path = row[2].strip() if len(row) > 2 else ''
            dept_short = org_path.split('/')[-1] if '/' in org_path else org_path
            print(f"  [{success+fail+1}/{len(rows)}] {dept_short}...", end=' ', flush=True)
            
            try:
                resignation_analyzer = None
                resign_file = ResignationAnalyzer.find_file_for_bg(org_path)
                if resign_file:
                    try:
                        resignation_analyzer = ResignationAnalyzer(resign_file)
                    except:
                        pass
                
                report_data = generate_report_data(
                    row, row_idx, feedback_reader,
                    data['jm_loader'], data['jm_detail_loader'],
                    data['yd_loader'], resignation_analyzer,
                    data['bp_loader'], data['jiangan_loader'],
                    data['open_fb_loader'], data['jm_open_loader'],
                    data['wc_loader']
                )
                
                org_id = str(row[0]).strip() if row[0] else ''
                org_path_safe = report_data['org_name'].replace('/', '-')
                org_path_safe = re.sub(r'[|\\:*?"<>]', '', org_path_safe)
                html_file = os.path.join(output_dir, f'{org_id}_{org_path_safe}_组织诊断报告.html')
                generate_html_report(report_data, html_file)
                
                print("✓")
                success += 1
            except Exception as e:
                print(f"✗ ({e})")
                fail += 1
        
        print(f"\n  {bg_key_upper} 完成: {success} 成功, {fail} 失败")
        print(f"  输出: {output_dir}/")


def interactive_generate(data):
    """交互式单部门生成"""
    diag_reader = data['diag_reader']
    feedback_reader = data['feedback_reader']
    
    while True:
        print("\n" + "=" * 80)
        dept_name = input("请输入部门名称 (输入 'q' 退出): ").strip()
        
        if dept_name.lower() == 'q':
            print("\n再见！")
            break
        
        if not dept_name:
            continue
        
        print(f"\n🔍 正在搜索 '{dept_name}'...")
        matches = diag_reader.find_rows(2, dept_name)
        
        if len(matches) == 0:
            print(f"✗ 未找到包含 '{dept_name}' 的部门")
            continue
        
        selected_row_idx = None
        selected_row = None
        
        if len(matches) == 1:
            selected_row_idx, selected_row = matches[0]
            print(f"✓ 找到唯一匹配: {selected_row[2]}")
        else:
            print(f"\n找到 {len(matches)} 个匹配的部门:")
            for i, (row_idx, row) in enumerate(matches, 1):
                print(f"  {i}. {row[2]}")
            
            while True:
                try:
                    choice = input(f"\n请选择 (1-{len(matches)}): ").strip()
                    choice_idx = int(choice) - 1
                    if 0 <= choice_idx < len(matches):
                        selected_row_idx, selected_row = matches[choice_idx]
                        break
                except ValueError:
                    print("请输入数字")
        
        print(f"\n⚙️  正在生成HTML报告...")
        try:
            org_full_path_for_resign = selected_row[2] if len(selected_row) > 2 else ""
            resignation_analyzer = None
            resign_file = ResignationAnalyzer.find_file_for_bg(org_full_path_for_resign)
            if resign_file:
                try:
                    resignation_analyzer = ResignationAnalyzer(resign_file)
                except Exception as e:
                    print(f"⚠ 离职明细加载失败: {e}")
            
            report_data = generate_report_data(
                selected_row, selected_row_idx, feedback_reader,
                data['jm_loader'], data['jm_detail_loader'],
                data['yd_loader'], resignation_analyzer,
                data['bp_loader'], data['jiangan_loader'],
                data['open_fb_loader'], data['jm_open_loader'],
                data['wc_loader']
            )
            
            org_id = str(selected_row[0]).strip() if selected_row[0] else ''
            org_path_safe = report_data['org_name'].replace('/', '-')
            org_path_safe = re.sub(r'[|\\:*?"<>]', '', org_path_safe)
            output_filename = f"{org_id}_{org_path_safe}_组织诊断报告.html"
            output_path = os.path.join(os.getcwd(), output_filename)
            
            generate_html_report(report_data, output_path)
            
            print(f"\n✅ 报告生成成功！")
            print(f"📁 文件位置: {output_path}")
            print(f"\n🌐 正在浏览器中打开...")
            
            import webbrowser
            webbrowser.open('file://' + output_path)
            
        except Exception as e:
            print(f"\n✗ 报告生成失败: {e}")
            import traceback
            traceback.print_exc()


def main():
    """主入口
    
    用法:
        python3 generate_html_report.py                  # 交互模式
        python3 generate_html_report.py --batch CDG      # 批量生成 CDG
        python3 generate_html_report.py --batch S3 CDG   # 批量生成 S3 和 CDG
        python3 generate_html_report.py --batch all      # 批量生成所有 BG
    """
    import argparse
    parser = argparse.ArgumentParser(description='组织诊断报告生成器')
    parser.add_argument('--batch', nargs='*', metavar='BG',
                        help='批量模式：指定BG名称(CDG/IEG/PCG/TEG/WXG/CSIG/S1/S2/S3/OFS)，或 all 生成全部')
    args = parser.parse_args()
    
    print("=" * 80)
    print("📊 组织诊断报告生成器")
    print("=" * 80)
    
    try:
        data = _load_all_data()
    except Exception as e:
        print(f"✗ 加载数据失败: {e}")
        return
    
    if args.batch is not None:
        # 批量模式
        if not args.batch or 'all' in [b.lower() for b in args.batch]:
            bg_list = list(BG_CONFIG.keys())
        else:
            bg_list = args.batch
        batch_generate(bg_list, data)
    else:
        # 交互模式
        interactive_generate(data)


if __name__ == '__main__':
    main()
