#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
敬满开放题BG汇总工具
从【敬满】/开放题文件夹中提取各BG的部门Excel文件，
为每个BG生成两个文件：
  - {BG}_敬满开放题汇总.xlsx（部门名称 + 详情）
  - {BG}_敬满开放题关键词汇总.xlsx（部门名称 + 关键词）
纯本地运行，不上传任何数据到云端。

用法（交互式）:
  python3 extract_jingman_open_answers.py
  然后输入BG序号/名称，或输入 all 批量处理所有BG
"""

import os
import re
import sys

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
SOURCE_DIR = os.path.join(SCRIPT_DIR, '开放题')
CELL_CHAR_LIMIT = 32767

FILE_PATTERN = re.compile(r'^(\d+)_2025_(.+?)_(.+)_敬满_整合版\.xlsx$')


def parse_filename(filename):
    m = FILE_PATTERN.match(filename)
    if not m:
        return None
    return m.group(1), m.group(2), m.group(3)


def list_all_bgs():
    bgs = set()
    for f in os.listdir(SOURCE_DIR):
        if not f.endswith('.xlsx') or f.startswith('~$'):
            continue
        info = parse_filename(f)
        if info:
            bgs.add(info[1])
    return sorted(bgs)


def get_bg_files(bg_name):
    results = []
    for f in os.listdir(SOURCE_DIR):
        if not f.endswith('.xlsx') or f.startswith('~$'):
            continue
        info = parse_filename(f)
        if info and info[1] == bg_name:
            results.append((f, info[2]))
    results.sort(key=lambda x: x[1])
    return results


def read_open_answers(filepath):
    from openpyxl import load_workbook
    wb = load_workbook(filepath, read_only=True, data_only=True)
    target_sheet = None
    for sn in wb.sheetnames:
        if '开放题回答' in sn:
            target_sheet = sn
            break
    if not target_sheet:
        wb.close()
        return [], [], [], []

    ws = wb[target_sheet]

    # 动态查找列：读取第1行表头，按关键字匹配
    header_row = None
    all_rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            header_row = row
        elif i >= 2:  # 从第3行开始是数据
            all_rows.append(row)

    wb.close()

    if not header_row:
        return [], [], [], []

    # 按表头关键字找列索引（初始值）
    col_positive = -1   # 正向观点（Q1相关，可能是选项列或文本列）
    col_negative = -1   # 负面观点（Q2相关）
    col_suggestion = -1 # 建议（Q3相关）
    col_keyword = -1    # 关键词（"请用三个关键词"）

    for i, h in enumerate(header_row):
        hs = str(h).strip() if h else ''
        if not hs:
            continue
        if '请用三个关键词' in hs:
            col_keyword = i
        elif ('做得好' in hs or '正向' in hs or 'Q1' in hs) and '员工填答' not in hs and col_positive == -1:
            col_positive = i
        elif ('做得更好' in hs or '改善' in hs or '负面' in hs or 'Q2' in hs) and '员工填答' not in hs and col_negative == -1:
            col_negative = i
        elif ('建议' in hs or 'Q3' in hs) and '关键词' not in hs and '员工填答' not in hs and col_suggestion == -1:
            col_suggestion = i

    # 某些表头在问题列留空，但员工填答在后一列（如 Q1 在第3列，正文在第4列）。
    # 简单启发：若下一列存在且本列表头非空、下一列表头为空，则在数据中比较本列与后一列文本长度，取更长的一列。
    def pick_text_col(col_idx, header_hint=''):
        if col_idx == -1:
            return -1
        alt = col_idx + 1
        if alt >= len(header_row):
            return col_idx
        # 规则1：如果当前列表头包含Q1/Q2/Q3且下一列表头为空，优先使用后一列（通常是“员工填答”列）
        if header_hint and not (header_row[alt] or '').strip():
            return alt
        # 规则2：否则比较前20行文本总长度，取更长的一列
        def col_len(c):
            total = 0
            for row in all_rows[:20]:
                if len(row) > c and row[c]:
                    total += len(str(row[c]).strip())
            return total
        cur_total = col_len(col_idx)
        alt_total = col_len(alt)
        if alt_total > cur_total:
            return alt
        return col_idx

    col_positive = pick_text_col(col_positive, 'Q1')
    col_negative = pick_text_col(col_negative, 'Q2')
    col_suggestion = pick_text_col(col_suggestion, 'Q3')
    col_keyword = pick_text_col(col_keyword)


    # 如果动态查找失败，回退到固定列号
    if col_positive == -1:
        col_positive = 3   # D列
    if col_negative == -1:
        col_negative = 5   # F列
    if col_suggestion == -1:
        col_suggestion = 7 # H列
    if col_keyword == -1:
        col_keyword = 8    # I列（默认）

    positive, negative, suggestion, keyword = [], [], [], []

    for row in all_rows:
        if len(row) > col_positive and row[col_positive] and str(row[col_positive]).strip():
            positive.append(str(row[col_positive]).strip())
        if len(row) > col_negative and row[col_negative] and str(row[col_negative]).strip():
            negative.append(str(row[col_negative]).strip())
        if len(row) > col_suggestion and row[col_suggestion] and str(row[col_suggestion]).strip():
            suggestion.append(str(row[col_suggestion]).strip())
        if len(row) > col_keyword and row[col_keyword] and str(row[col_keyword]).strip():
            keyword.append(str(row[col_keyword]).strip())

    return positive, negative, suggestion, keyword


def build_detail(positive, negative, suggestion):
    pos_text = '||'.join(positive) if positive else '无'
    neg_text = '||'.join(negative) if negative else '无'
    sug_text = '||'.join(suggestion) if suggestion else '无'
    return f"1）正向观点：{pos_text}\n2）负面观点：{neg_text}\n3）建议：{sug_text}"


def build_keywords(keyword_list):
    if not keyword_list:
        return '无'
    return '||'.join(keyword_list)


def split_text_to_cells(text, limit=CELL_CHAR_LIMIT):
    if len(text) <= limit:
        return [text]
    chunks = []
    remaining = text
    while remaining:
        if len(remaining) <= limit:
            chunks.append(remaining)
            break
        cut = remaining[:limit].rfind('\n')
        if cut == -1 or cut < limit // 2:
            cut = remaining[:limit].rfind('||')
        if cut == -1 or cut < limit // 2:
            cut = limit
        chunks.append(remaining[:cut])
        remaining = remaining[cut:].lstrip('\n')
    return chunks


def write_two_col_excel(filepath, sheet_title, header_b, rows):
    """写两列Excel：A=部门名称, B=内容。超长文本自动切割到相邻列。"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title

    hf = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')
    hfill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    ha = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin = Border(left=Side(style='thin'), right=Side(style='thin'),
                  top=Side(style='thin'), bottom=Side(style='thin'))
    df = Font(name='微软雅黑', size=10)
    da = Alignment(vertical='top', wrap_text=True)

    split_rows = []
    max_cols = 2
    for dept_name, content in rows:
        chunks = split_text_to_cells(content)
        split_rows.append((dept_name, chunks))
        if 1 + len(chunks) > max_cols:
            max_cols = 1 + len(chunks)

    # 表头
    for ci, text in enumerate(['部门名称', header_b], 1):
        cell = ws.cell(row=1, column=ci, value=text)
        cell.font, cell.fill, cell.alignment, cell.border = hf, hfill, ha, thin
    for ci in range(3, max_cols + 1):
        cell = ws.cell(row=1, column=ci, value='')
        cell.fill, cell.border = hfill, thin

    # 数据
    for ri, (dept_name, chunks) in enumerate(split_rows, 2):
        cell = ws.cell(row=ri, column=1, value=dept_name)
        cell.font, cell.alignment, cell.border = df, da, thin
        for ci, chunk in enumerate(chunks, 2):
            cell = ws.cell(row=ri, column=ci, value=chunk)
            cell.font, cell.alignment, cell.border = df, da, thin

    ws.column_dimensions['A'].width = 30
    for ci in range(2, max_cols + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 100
    ws.freeze_panes = 'A2'
    wb.save(filepath)


def process_bg(bg_name):
    """处理单个BG，生成汇总+关键词汇总两个文件。返回 (成功部门数, 失败部门数)"""
    bg_files = get_bg_files(bg_name)
    if not bg_files:
        print(f"  ⚠️  {bg_name} 无原始文件，跳过")
        return 0, 0

    print(f"\n📊 {bg_name} 共有 {len(bg_files)} 个部门，正在读取...")

    detail_rows = []   # [(dept_name, detail_text), ...]
    keyword_rows = []  # [(dept_name, keyword_text), ...]
    success, error = 0, 0

    for filename, dept_name in bg_files:
        filepath = os.path.join(SOURCE_DIR, filename)
        try:
            pos, neg, sug, kw_list = read_open_answers(filepath)
            detail_rows.append((dept_name, build_detail(pos, neg, sug)))
            keyword_rows.append((dept_name, build_keywords(kw_list)))
            print(f"  ✅ {dept_name}  —  正向{len(pos)} / 负面{len(neg)} / 建议{len(sug)} / 关键词{len(kw_list)}")
            success += 1
        except Exception as e:
            print(f"  ❌ {dept_name}  —  读取失败: {e}")
            detail_rows.append((dept_name, f"读取失败: {e}"))
            keyword_rows.append((dept_name, ""))
            error += 1

    print(f"  📊 完成：成功 {success}，失败 {error}")

    # 写详情文件
    detail_path = os.path.join(SCRIPT_DIR, f'{bg_name}_敬满开放题汇总.xlsx')
    write_two_col_excel(detail_path, f'{bg_name}敬满开放题汇总', '详情', detail_rows)
    print(f"  📝 {bg_name}_敬满开放题汇总.xlsx（{len(detail_rows)} 个部门）")

    # 写关键词文件
    kw_path = os.path.join(SCRIPT_DIR, f'{bg_name}_敬满开放题关键词汇总.xlsx')
    write_two_col_excel(kw_path, f'{bg_name}敬满开放题关键词汇总', '关键词', keyword_rows)
    print(f"  📝 {bg_name}_敬满开放题关键词汇总.xlsx（{len(keyword_rows)} 个部门）")

    return success, error


def main():
    print("\n" + "=" * 60)
    print("  📋 敬满开放题BG汇总工具（纯本地运行）")
    print("=" * 60)

    if not os.path.isdir(SOURCE_DIR):
        print(f"\n❌ 找不到数据源文件夹：{SOURCE_DIR}")
        return

    all_bgs = list_all_bgs()
    if not all_bgs:
        print("\n❌ 开放题文件夹中没有找到标准格式的文件")
        return

    print(f"\n📂 数据源：{SOURCE_DIR}")
    print(f"📋 共有 {len(all_bgs)} 个BG：")
    for i, bg in enumerate(all_bgs, 1):
        print(f"   {i}. {bg}")
    print(f"   输入 all 可批量处理所有BG")

    # 用户选择 BG
    while True:
        print()
        user_input = input("🔍 请输入BG名称/序号/all，输入 q 退出：").strip()

        if user_input.lower() == 'q':
            print("\n👋 已退出")
            return

        # 批量处理所有
        if user_input.lower() == 'all':
            selected_bgs = all_bgs
            print(f"\n✅ 将批量处理所有 {len(all_bgs)} 个BG")
            break

        # 按序号
        if user_input.isdigit():
            idx = int(user_input) - 1
            if 0 <= idx < len(all_bgs):
                selected_bgs = [all_bgs[idx]]
                print(f"\n✅ 已选择：{selected_bgs[0]}")
                break
            else:
                print("❌ 序号超出范围，请重试")
                continue

        # 模糊匹配
        matches = [bg for bg in all_bgs if user_input.upper() in bg.upper()]
        if len(matches) == 0:
            print("❌ 没有匹配的BG，请重试")
        elif len(matches) == 1:
            selected_bgs = [matches[0]]
            print(f"\n✅ 匹配到：{matches[0]}")
            break
        else:
            print(f"\n🔎 找到 {len(matches)} 个匹配结果：")
            for i, bg in enumerate(matches, 1):
                print(f"   {i}. {bg}")
            sub = input("   请输入序号确认：").strip()
            if sub.isdigit():
                idx = int(sub) - 1
                if 0 <= idx < len(matches):
                    selected_bgs = [matches[idx]]
                    print(f"\n✅ 已选择：{matches[idx]}")
                    break
            print("❌ 无效输入，请重试")

    # 处理
    total_s, total_e = 0, 0
    for bg in selected_bgs:
        s, e = process_bg(bg)
        total_s += s
        total_e += e

    # 汇总
    print("\n" + "=" * 60)
    print(f"  🎉 全部完成！成功 {total_s} 个部门，失败 {total_e} 个")
    print("=" * 60)
    print("\n📊 最终产物：")
    for bg in selected_bgs:
        d = '✅' if os.path.exists(os.path.join(SCRIPT_DIR, f'{bg}_敬满开放题汇总.xlsx')) else '❌'
        k = '✅' if os.path.exists(os.path.join(SCRIPT_DIR, f'{bg}_敬满开放题关键词汇总.xlsx')) else '❌'
        print(f"  {bg}: {d} 汇总  |  {k} 关键词汇总")


if __name__ == '__main__':
    main()
