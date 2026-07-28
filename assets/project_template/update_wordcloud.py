#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
更新 CSIG / WXG 报告的词云部分（仅替换词云 section，不改其他内容）
从更新后的关键词分析 xlsx 重新加载数据，
替换 HTML 报告中 <div class="wc-section"> ... </script> 部分。

用法:
  python3 update_wordcloud.py            # 更新 CSIG + WXG
  python3 update_wordcloud.py csig       # 只更新 CSIG
  python3 update_wordcloud.py wxg        # 只更新 WXG
"""
import sys, os, re, json

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
JINGMAN_DIR = os.path.join(SCRIPT_DIR, '【敬满】')

# BG 配置：(xlsx文件名, html目录, 文件名前缀, 后缀, org_path解析方式)
BG_CONFIG = {
    'csig': {
        'xlsx': os.path.join(JINGMAN_DIR, 'CSIG_敬满开放题关键词分析.xlsx'),
        'html_dir': os.path.join(SCRIPT_DIR, '报告', 'csig', 'html'),
        'file_prefix': '报告_CSIG',
        'file_suffix': '_20260409.html',
        'strip_prefix': '报告_',
    },
    'wxg': {
        'xlsx': os.path.join(JINGMAN_DIR, 'WXG_敬满开放题关键词分析.xlsx'),
        'html_dir': os.path.join(SCRIPT_DIR, '报告', 'wxg', 'html'),
        'file_prefix': '报告_WXG',
        'file_suffix': '_20260327_180000.html',
        'strip_prefix': '报告_',
    },
    'cdg': {
        'xlsx': os.path.join(JINGMAN_DIR, 'CDG_敬满开放题关键词分析.xlsx'),
        'html_dir': os.path.join(SCRIPT_DIR, '报告', 'cdg', 'html'),
        'file_prefix': '',
        'file_suffix': '_组织诊断报告.html',
        'strip_prefix': '',
    },
    's3': {
        'xlsx': os.path.join(JINGMAN_DIR, 'S3_敬满开放题关键词分析.xlsx'),
        'html_dir': os.path.join(SCRIPT_DIR, '报告', 's3', 'html'),
        'file_prefix': '报告_S3',
        'file_suffix': '_20260423.html',
        'strip_prefix': '报告_',
    },
    'ieg': {
        'xlsx': os.path.join(JINGMAN_DIR, 'IEG_敬满开放题关键词分析.xlsx'),
        'html_dir': os.path.join(SCRIPT_DIR, '报告', 'ieg', 'html'),
        'file_prefix': '',
        'file_suffix': '_组织诊断报告.html',
        'strip_prefix': '',
    },
}

_WC_FREQ_THRESHOLD = 20  # ≥2的词不足此数时降级取全部词

_SENTIMENT_MAP = {'正向': 'positive', '负向': 'negative', '中性': 'neutral', '中立': 'neutral'}

# 英文停用词 — 词云中过滤掉的无意义虚词/介词/连词
_WC_EN_STOPWORDS = {
    'a', 'an', 'the', 'to', 'and', 'or', 'of', 'in', 'on', 'at',
    'is', 'it', 'be', 'for', 'as', 'by', 'no', 'not', 'n', 'but',
    'with', 'that', 'this', 'are', 'was', 'were', 'has', 'have', 'had',
    'if', 'so', 'do', 'up', 'we', 'my', 'me', 'he', 'she', 'i', 'its',
    'can', 'will', 'just', 'should', 'would', 'could', 'may', 'might',
    'am', 'been', 'being', 'did', 'does', 'doing', 'each', 'few',
    'from', 'get', 'got', 'her', 'him', 'his', 'how', 'into',
    'more', 'most', 'much', 'must', 'nor', 'our', 'out', 'own',
    'same', 'some', 'such', 'than', 'them', 'then', 'too', 'very',
    'what', 'when', 'who', 'whom', 'why', 'you', 'your',
    'c',
}

def _wc_is_stopword(word):
    """判断是否为词云应过滤的停用词"""
    w = word.strip().lower()
    if w in _WC_EN_STOPWORDS:
        return True
    if re.fullmatch(r'\d+\.?', w):
        return True
    if re.fullmatch(r'[^a-zA-Z\u4e00-\u9fff\u3040-\u309f\u30a0-\u30ff]+', w):
        return True
    return False

# ── JS 模板 ──
_JS_TPL = '''
(function() {
    var wcData = PLACEHOLDER_JSON;
    function lerpColor(c1, c2, t) {
        return [
            Math.round(c1[0] + (c2[0] - c1[0]) * t),
            Math.round(c1[1] + (c2[1] - c1[1]) * t),
            Math.round(c1[2] + (c2[2] - c1[2]) * t),
        ];
    }
    function sentimentColor(sentiment, freq, maxF, minF) {
        var logMax = Math.log(maxF), logMin = Math.log(minF);
        var t = logMax === logMin ? 0.5 : (Math.log(freq) - logMin) / (logMax - logMin);
        var rgb;
        if (sentiment === 'positive') {
            rgb = lerpColor([183, 235, 143], [35, 120, 4], t);
        } else if (sentiment === 'negative') {
            rgb = lerpColor([255, 163, 158], [168, 7, 26], t);
        } else {
            rgb = lerpColor([255, 229, 143], [173, 104, 0], t);
        }
        return 'rgb(' + rgb[0] + ',' + rgb[1] + ',' + rgb[2] + ')';
    }
    var sentimentLabel = { positive: '正向', negative: '负向', neutral: '中性' };
    var maxFreq = Math.max.apply(null, wcData.map(function(w) { return w.freq; }));
    var minFreq = Math.min.apply(null, wcData.map(function(w) { return w.freq; }));
    function fontSize(freq) {
        var logMax = Math.log(maxFreq), logMin = Math.log(minFreq);
        var t = logMax === logMin ? 0.5 : (Math.log(freq) - logMin) / (logMax - logMin);
        return 13 + t * 44;
    }
    var box = document.getElementById('wcCloudBox');
    var W = box.clientWidth, H = box.clientHeight;
    window._wcRenderWidth = W;
    var cx = W / 2, cy = H / 2;
    var _mc = document.createElement('canvas').getContext('2d');
    wcData.sort(function(a, b) { return b.freq - a.freq; });
    var placed = [];
    function measure(text, size) {
        _mc.font = '700 ' + size + 'px "PingFang SC","Microsoft YaHei",sans-serif';
        return { w: _mc.measureText(text).width + 4, h: size * 1.25 + 2 };
    }
    function collides(r) {
        if (r.x < 2 || r.y < 2 || r.x + r.w > W - 2 || r.y + r.h > H - 2) return true;
        for (var i = 0; i < placed.length; i++) {
            var p = placed[i];
            if (!(r.x >= p.x + p.w || r.x + r.w <= p.x || r.y >= p.y + p.h || r.y + r.h <= p.y)) return true;
        }
        return false;
    }
    for (var wi = 0; wi < wcData.length; wi++) {
        var word = wcData[wi];
        var size = fontSize(word.freq);
        var m = measure(word.text, size);
        var ok = false;
        var a = Math.random() * Math.PI * 2;
        for (var i = 0; i < 12000 && !ok; i++) {
            var r = i * 0.04;
            var x = cx + r * Math.cos(a) - m.w / 2;
            var y = cy + r * Math.sin(a) - m.h / 2;
            a += 0.12;
            var rect = { x: x, y: y, w: m.w, h: m.h };
            if (!collides(rect)) {
                placed.push(rect);
                var el = document.createElement('span');
                el.className = 'wc-word';
                el.textContent = word.text;
                el.style.cssText = 'left:' + x + 'px;top:' + y + 'px;font-size:' + size + 'px;color:' + sentimentColor(word.sentiment, word.freq, maxFreq, minFreq) + ';';
                el.setAttribute('data-freq', word.freq);
                el.setAttribute('data-sentiment', sentimentLabel[word.sentiment]);
                box.appendChild(el);
                ok = true;
            }
        }
    }
    if (placed.length > 0) {
        var minY = Infinity, maxY = -Infinity;
        for (var pi = 0; pi < placed.length; pi++) {
            if (placed[pi].y < minY) minY = placed[pi].y;
            if (placed[pi].y + placed[pi].h > maxY) maxY = placed[pi].y + placed[pi].h;
        }
        var contentH = maxY - minY;
        var pad = 28;
        var newH = Math.max(contentH + pad * 2, 160);
        var offsetY = pad - minY + (newH - pad * 2 - contentH) / 2;
        var spans = box.querySelectorAll('.wc-word');
        for (var si = 0; si < spans.length; si++) {
            var curTop = parseFloat(spans[si].style.top);
            spans[si].style.top = (curTop + offsetY) + 'px';
        }
        box.style.height = newH + 'px';
        window._wcRenderHeight = newH;
    }
    var tip = document.getElementById('wcTooltip');
    box.addEventListener('mousemove', function(e) {
        var t = e.target.closest('.wc-word');
        if (t) {
            tip.textContent = t.textContent + '  x' + t.getAttribute('data-freq') + '  ' + t.getAttribute('data-sentiment');
            tip.style.display = 'block';
            tip.style.left = (e.clientX + 12) + 'px';
            tip.style.top  = (e.clientY - 30) + 'px';
        } else {
            tip.style.display = 'none';
        }
    });
    box.addEventListener('mouseleave', function() { tip.style.display = 'none'; });
})();
'''


def load_keyword_data(xlsx_path):
    """从关键词分析xlsx加载数据。
    新格式：A=部门名称, B=关键词（词性 频次），用；分隔
    返回 {dept_name: [(word, sentiment, freq), ...]}
    """
    from openpyxl import load_workbook
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    result = {}
    kw_re = re.compile(r'^(.+?)\s+(正向|负向|中性|中立)\s+(\d+)$')

    for row in ws.iter_rows(min_row=2, values_only=True):
        dept = str(row[0]).strip() if row[0] else ''
        raw = str(row[1]).strip() if len(row) > 1 and row[1] else ''
        if not dept or not raw or raw in ('无关键词', '无高频关键词', '无', 'nan'):
            continue
        entries = []
        for item in re.split(r'[；;]', raw):
            item = item.strip()
            m = kw_re.match(item)
            if m:
                word = m.group(1).strip()
                sentiment = _SENTIMENT_MAP.get(m.group(2), 'neutral')
                freq = int(m.group(3))
                if not _wc_is_stopword(word):
                    entries.append((word, sentiment, freq))
        if entries:
            result[dept] = entries
    wb.close()
    return result


def get_org_path_from_filename(filename, config):
    """从文件名解析出部门路径，用于匹配词云数据"""
    name = filename
    if name.startswith(config['strip_prefix']):
        name = name[len(config['strip_prefix']):]
    if name.endswith(config['file_suffix']):
        name = name[:-len(config['file_suffix'])]
    return name


def match_dept(dept_data, org_name):
    """模糊匹配部门名"""
    parts = [p.strip() for p in org_name.replace('\\', '/').split('_') if p.strip()]
    if not parts:
        return None
    dept_name = parts[-1]

    # 对于 "20806_CDG企业发展事业群-支付基础平台与金融应用线-Tango项目中心" 这种格式
    # 用 '-' 再拆分最后一段，取真正的部门名
    if '-' in dept_name:
        dash_parts = [p.strip() for p in dept_name.split('-') if p.strip()]
        dept_name_short = dash_parts[-1] if dash_parts else dept_name
        # 尝试最后两级（线-部门）组合
        if len(dash_parts) >= 2:
            combo_dash = dash_parts[-2] + '_' + dash_parts[-1]
            if combo_dash in dept_data:
                return combo_dash
    else:
        dept_name_short = dept_name

    # 精确匹配
    if dept_name_short in dept_data:
        return dept_name_short
    if dept_name in dept_data:
        return dept_name
    # 尝试最后两级组合（下划线分割）
    if len(parts) >= 2:
        combo = parts[-2] + '_' + parts[-1]
        if combo in dept_data:
            return combo
    # 模糊包含
    for key in dept_data:
        if dept_name_short == key or dept_name_short.endswith(key) or key.endswith(dept_name_short):
            return key
    for key in dept_data:
        if dept_name_short in key or key in dept_name_short:
            return key
    for key in dept_data:
        if dept_name in key or key in dept_name:
            return key
    if len(parts) >= 2:
        combo = parts[-2] + '_' + parts[-1]
        for key in dept_data:
            if combo in key or key in combo:
                return key
    return None


def gen_wc_section(wc_json_str, wc_dept):
    wc_dept_safe = wc_dept.replace('"', '&quot;').replace("'", '&#39;')
    html_part = (
        '\n        <div class="wc-section">\n'
        '            <div class="wc-section-title">' + wc_dept_safe + ' \u00b7 \u656c\u6ee1\u5f00\u653e\u9898\u5173\u952e\u8bcd\u4e91</div>\n'
        '            <div class="wc-section-subtitle">\u5b57\u53f7 = \u8bcd\u9891 &nbsp;&middot;&nbsp; \u989c\u8272\u6df1\u6d45 = \u8bcd\u9891\u5f3a\u5ea6 &nbsp;&middot;&nbsp; \u8272\u7cfb = \u60c5\u611f\u503e\u5411</div>\n'
        '            <div class="wc-cloud-container" id="wcCloudBox"></div>\n'
        '            <div id="wcTooltip" class="wc-tooltip"></div>\n'
        '            <div class="wc-legend">\n'
        '                <div class="wc-legend-item"><div class="wc-legend-bar positive"></div>\u6b63\u5411\u8bcd</div>\n'
        '                <div class="wc-legend-item"><div class="wc-legend-bar neutral"></div>\u4e2d\u6027\u8bcd</div>\n'
        '                <div class="wc-legend-item"><div class="wc-legend-bar negative"></div>\u8d1f\u5411\u8bcd</div>\n'
        '            </div>\n'
        '        </div>\n'
    )
    js_code = _JS_TPL.replace('PLACEHOLDER_JSON', wc_json_str)
    return html_part + '<script>' + js_code + '        </script>\n'


def replace_wc_section(html, new_wc):
    footer_pos = html.find('<div class="footer">')
    
    # 1. 查找已有的词云 section
    start = html.find('<div class="wc-section">')
    if start >= 0:
        script_start = html.find('<script>', start)
        if script_start >= 0:
            end_tag = html.find('</script>', script_start)
            if end_tag >= 0:
                end = end_tag + len('</script>')
                # 先删除旧词云
                html = html[:start] + html[end:]
                # 重新定位 footer（删除后位置可能变了）
                footer_pos = html.find('<div class="footer">')
    
    # 2. 插入到 footer 之前
    if footer_pos > 0:
        return html[:footer_pos] + new_wc + '\n        ' + html[footer_pos:], None
    
    # 3. 没有 footer -> 在 </body> 之前
    body_end = html.rfind('</body>')
    if body_end > 0:
        return html[:body_end] + new_wc + '\n' + html[body_end:], None
    
    return None, 'no insertion point found'


def update_bg(bg_key):
    config = BG_CONFIG[bg_key]
    print('\n' + '=' * 50)
    print('  ' + bg_key.upper() + ' \u8bcd\u4e91\u66f4\u65b0')
    print('=' * 50)

    if not os.path.exists(config['xlsx']):
        print('  \u274c \u5173\u952e\u8bcd\u5206\u6790\u6587\u4ef6\u4e0d\u5b58\u5728: ' + config['xlsx'])
        return

    dept_data = load_keyword_data(config['xlsx'])
    print('  \u2713 \u52a0\u8f7d\u5173\u952e\u8bcd\u6570\u636e: ' + str(len(dept_data)) + ' \u4e2a\u90e8\u95e8')

    html_dir = config['html_dir']
    if not os.path.isdir(html_dir):
        print('  \u274c HTML\u76ee\u5f55\u4e0d\u5b58\u5728: ' + html_dir)
        return

    # 找需要更新的 HTML（含 S1/S2/S3 职能系统的 CSIG/WXG 报告也要处理）
    files = sorted([f for f in os.listdir(html_dir) if f.endswith('.html')])
    print('  \u627e\u5230 ' + str(len(files)) + ' \u4e2a HTML \u6587\u4ef6')

    updated = 0
    skipped = 0

    for filename in files:
        org_name = get_org_path_from_filename(filename, config)
        dept_key = match_dept(dept_data, org_name)
        if not dept_key:
            skipped += 1
            continue

        entries = dept_data[dept_key]
        # 词频阈值降级逻辑：≥2的词不足20个时取全部词
        high_freq_entries = [(w, s, f) for w, s, f in entries if f >= 2]
        if len(high_freq_entries) >= _WC_FREQ_THRESHOLD:
            use_entries = high_freq_entries
        else:
            use_entries = entries
        top_n = min(80, len(use_entries))
        words_data = [{'text': w, 'freq': f, 'sentiment': s} for w, s, f in use_entries[:top_n]]
        wc_json = json.dumps(words_data, ensure_ascii=False)
        new_wc = gen_wc_section(wc_json, dept_key)

        filepath = os.path.join(html_dir, filename)
        with open(filepath, 'r', encoding='utf-8') as f:
            html = f.read()

        new_html, err = replace_wc_section(html, new_wc)
        if err:
            skipped += 1
            continue

        with open(filepath, 'w', encoding='utf-8') as f:
            f.write(new_html)

        print('  \u2713 ' + dept_key + '  (' + filename[:40] + '...)')
        updated += 1

    print('\n  \u5b8c\u6210\uff1a\u66f4\u65b0 ' + str(updated) + ' \u4e2a\uff0c\u8df3\u8fc7 ' + str(skipped) + ' \u4e2a')


def main():
    args = [a.lower() for a in sys.argv[1:]]
    if not args:
        args = ['csig', 'wxg']

    print('\n' + '=' * 60)
    print('  \U0001f504 \u66f4\u65b0\u62a5\u544a\u8bcd\u4e91\u90e8\u5206\uff08\u4ec5\u8bcd\u4e91\uff0c\u5176\u4f59\u4e0d\u52a8\uff09')
    print('=' * 60)

    for bg in args:
        if bg in BG_CONFIG:
            update_bg(bg)
        else:
            print('\n  \u26a0 \u672a\u77e5BG: ' + bg + '\uff0c\u652f\u6301: ' + ', '.join(BG_CONFIG.keys()))

    print('\n\u5168\u90e8\u5b8c\u6210\uff01\n')


if __name__ == '__main__':
    main()
