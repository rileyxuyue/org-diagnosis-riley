#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
全面反馈开放题提取工具
根据BG筛选组织诊断数据，匹配全面反馈的闪光点、更多期待和BP观察，生成Excel
"""

import os
import re
import sys
import glob
import zipfile
import xml.etree.ElementTree as ET

# ─── 配置 ────────────────────────────────────────────────
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PARENT_DIR = os.path.dirname(SCRIPT_DIR)

ORG_DIAG_FILE = os.path.join(PARENT_DIR, '【组织诊断结果】', '组织诊断2025全年最终版.xlsx')
FEEDBACK_FILE = os.path.join(SCRIPT_DIR, '全面反馈25H2v2.xlsx')
OUTPUT_FILE = os.path.join(SCRIPT_DIR, '全面反馈开放题.xlsx')

# Excel 单元格字符上限
CELL_CHAR_LIMIT = 32767

NS = {'s': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}


# ─── xlsx 读取（纯 zipfile 方式，避免 openpyxl 兼容问题） ──
def read_xlsx_sheet(filepath, sheet_index=0):
    """读取 xlsx 的指定 sheet，返回 (headers, rows)"""
    z = zipfile.ZipFile(filepath)

    # 1. 解析 shared strings
    ss = []
    if 'xl/sharedStrings.xml' in z.namelist():
        tree = ET.parse(z.open('xl/sharedStrings.xml'))
        for si in tree.findall('.//s:si', NS):
            text = ''.join(t.text or '' for t in si.findall('.//s:t', NS))
            ss.append(text)

    # 2. 找到对应 sheet 文件
    sheet_file = f'xl/worksheets/sheet{sheet_index + 1}.xml'
    tree = ET.parse(z.open(sheet_file))
    xml_rows = tree.findall('.//s:sheetData/s:row', NS)

    # 3. 解析列号
    def col_index(ref):
        """将 'AB3' -> 列索引 (A=0, B=1, ..., Z=25, AA=26, ...)"""
        letters = re.match(r'([A-Z]+)', ref).group(1)
        idx = 0
        for ch in letters:
            idx = idx * 26 + (ord(ch) - ord('A') + 1)
        return idx - 1

    # 4. 读取所有行
    all_rows = []
    max_col = 0
    for row_el in xml_rows:
        cells = row_el.findall('s:c', NS)
        row_dict = {}
        for c in cells:
            ref = c.get('r')
            ci = col_index(ref)
            max_col = max(max_col, ci)
            t = c.get('t')
            v_el = c.find('s:v', NS)
            v = v_el.text if v_el is not None else ''
            if t == 's' and v:
                v = ss[int(v)]
            row_dict[ci] = v
        all_rows.append(row_dict)

    z.close()

    # 5. 转为列表格式
    if not all_rows:
        return [], []
    headers = [all_rows[0].get(i, '') for i in range(max_col + 1)]
    rows = []
    for row_dict in all_rows[1:]:
        row = [row_dict.get(i, '') for i in range(max_col + 1)]
        rows.append(row)

    return headers, rows


def find_col(headers, name):
    """根据列名查找索引"""
    for i, h in enumerate(headers):
        if h.strip() == name.strip():
            return i
    return -1


# ─── 人名解析 ──────────────────────────────────────────────
def parse_team_members(team_str):
    """
    解析 '部门负责人-1管理者团队' 列
    格式：zhangsan(张三)：绿灯；lisi(李四)：绿灯
    返回：['zhangsan(张三)', 'lisi(李四)']
    """
    if not team_str or not team_str.strip():
        return []
    members = []
    # 按中文分号分割
    parts = re.split(r'[；;]', team_str)
    for part in parts:
        part = part.strip()
        if not part:
            continue
        # 提取 englishname(中文名) 部分
        match = re.match(r'(\w+\([^)]+\))', part)
        if match:
            members.append(match.group(1))
    return members


def extract_english_name(name_str):
    """从 'zhangsan(张三)' 提取英文名 'zhangsan' 用于匹配"""
    match = re.match(r'(\w+)\(', name_str)
    return match.group(1).lower() if match else name_str.lower().strip()


# ─── BP观察文件查找 ──────────────────────────────────────
def find_bp_file(bg_name):
    """
    在【全面反馈】文件夹中查找对应BG的BP观察文件
    文件名格式：XXX部门负责人-1BP观察.xlsx
    其中 XXX 可能是BG缩写（如WXG）或完整名称
    """
    # 提取BG缩写（取中文前的英文/数字部分）
    bg_short = re.match(r'([A-Za-z0-9]+)', bg_name)
    bg_short = bg_short.group(1) if bg_short else ''

    # 搜索匹配的文件
    pattern = os.path.join(SCRIPT_DIR, '*BP观察*.xlsx')
    candidates = [f for f in glob.glob(pattern) if not os.path.basename(f).startswith('~$')]

    for f in candidates:
        basename = os.path.basename(f)
        # 精确匹配：文件名以 "BG缩写" 开头，后面紧跟非字母数字字符（如中文或符号）
        if bg_short and re.match(re.escape(bg_short) + r'(?![A-Za-z0-9])', basename):
            return f
        # 也匹配完整BG名
        if bg_name in basename:
            return f

    return None


def load_bp_observations(bp_file):
    """加载BP观察数据，返回 {英文名小写: bp观察内容}"""
    if not bp_file or not os.path.exists(bp_file):
        return {}

    headers, rows = read_xlsx_sheet(bp_file, sheet_index=0)
    col_name = find_col(headers, '中英文名')
    col_bp = find_col(headers, 'BP观察')

    if col_name == -1 or col_bp == -1:
        print(f"  ⚠️  BP观察文件列名不匹配，跳过（需要'中英文名'和'BP观察'列）")
        return {}

    bp_map = {}
    for row in rows:
        name = row[col_name].strip() if col_name < len(row) else ''
        bp = row[col_bp] if col_bp < len(row) else ''
        if name:
            en = extract_english_name(name)
            bp_map[en] = bp
    return bp_map


def split_text_to_cells(text, limit=CELL_CHAR_LIMIT):
    """
    将超长文本按字符上限切割成多段
    尽量在换行处切割，避免截断内容
    返回列表
    """
    if len(text) <= limit:
        return [text]

    chunks = []
    remaining = text
    while remaining:
        if len(remaining) <= limit:
            chunks.append(remaining)
            break
        # 在限制内找最后一个双换行（人与人之间的分隔）
        cut = remaining[:limit].rfind('\n\n')
        if cut == -1 or cut < limit // 2:
            # 找不到合适的双换行，找普通换行
            cut = remaining[:limit].rfind('\n')
        if cut == -1 or cut < limit // 2:
            # 实在没有好的切割点，硬切
            cut = limit
        chunks.append(remaining[:cut])
        remaining = remaining[cut:].lstrip('\n')
    return chunks


# ─── 主流程 ──────────────────────────────────────────────
def main():
    print("\n📂 正在读取数据文件...\n")

    # 检查文件存在
    if not os.path.exists(ORG_DIAG_FILE):
        print(f"❌ 找不到文件：{ORG_DIAG_FILE}")
        return
    if not os.path.exists(FEEDBACK_FILE):
        print(f"❌ 找不到文件：{FEEDBACK_FILE}")
        return

    # 1. 读取组织诊断 - 亮灯明细 (sheet index 1)
    print("  📊 读取组织诊断数据...")
    org_headers, org_rows = read_xlsx_sheet(ORG_DIAG_FILE, sheet_index=1)

    col_bg = find_col(org_headers, 'BG')
    col_path = find_col(org_headers, '组织全路径')
    col_leader = find_col(org_headers, '组织负责人')
    col_team = find_col(org_headers, '部门负责人-1管理者团队')

    if any(c == -1 for c in [col_bg, col_path, col_leader, col_team]):
        print(f"❌ 组织诊断表中找不到必要的列！")
        print(f"   找到的列：{org_headers[:16]}")
        return

    # 2. 读取全面反馈 (sheet index 0)
    print("  📊 读取全面反馈数据...")
    fb_headers, fb_rows = read_xlsx_sheet(FEEDBACK_FILE, sheet_index=0)

    col_fb_name = find_col(fb_headers, '被评估人中英文名')
    col_fb_shine = find_col(fb_headers, '闪光点')
    col_fb_expect = find_col(fb_headers, '更多期待')

    if any(c == -1 for c in [col_fb_name, col_fb_shine, col_fb_expect]):
        print(f"❌ 全面反馈表中找不到必要的列！")
        return

    # 建立全面反馈查找字典 (英文名小写 -> {name, shine, expect})
    feedback_map = {}
    for row in fb_rows:
        name_full = row[col_fb_name].strip() if col_fb_name < len(row) else ''
        if not name_full:
            continue
        en_name = extract_english_name(name_full)
        shine = row[col_fb_shine] if col_fb_shine < len(row) else ''
        expect = row[col_fb_expect] if col_fb_expect < len(row) else ''
        feedback_map[en_name] = {
            'name': name_full,
            'shine': shine,
            'expect': expect
        }

    print(f"  ✅ 组织诊断：{len(org_rows)} 条记录")
    print(f"  ✅ 全面反馈：{len(feedback_map)} 人\n")

    # 3. 获取所有不重复的 BG
    all_bgs = sorted(set(row[col_bg].strip() for row in org_rows if row[col_bg].strip()))
    print(f"📋 共有 {len(all_bgs)} 个BG：")
    for i, bg in enumerate(all_bgs, 1):
        print(f"   {i}. {bg}")

    # 4. 让用户选择 BG（支持模糊搜索）
    while True:
        print()
        user_input = input("🔍 请输入BG名称（支持模糊搜索）或序号，输入 q 退出：").strip()

        if user_input.lower() == 'q':
            print("\n👋 已退出")
            return

        # 尝试按序号
        if user_input.isdigit():
            idx = int(user_input) - 1
            if 0 <= idx < len(all_bgs):
                selected_bg = all_bgs[idx]
                print(f"\n✅ 已选择：{selected_bg}")
                break
            else:
                print("❌ 序号超出范围，请重试")
                continue

        # 模糊匹配
        matches = [bg for bg in all_bgs if user_input.lower() in bg.lower()]

        if len(matches) == 0:
            print("❌ 没有匹配的BG，请重试")
            continue
        elif len(matches) == 1:
            selected_bg = matches[0]
            print(f"\n✅ 匹配到：{selected_bg}")
            break
        else:
            print(f"\n🔎 找到 {len(matches)} 个匹配结果：")
            for i, bg in enumerate(matches, 1):
                print(f"   {i}. {bg}")
            sub = input("   请输入序号确认：").strip()
            if sub.isdigit():
                idx = int(sub) - 1
                if 0 <= idx < len(matches):
                    selected_bg = matches[idx]
                    print(f"\n✅ 已选择：{selected_bg}")
                    break
            print("❌ 无效输入，请重试")

def process_bg(selected_bg, org_rows, col_bg, col_path, col_leader, col_team, feedback_map):
    """处理单个BG，返回 (output_rows, matched, missing)"""
    bg_depts = [row for row in org_rows if row[col_bg].strip() == selected_bg]
    print(f"\n📊 {selected_bg} 共有 {len(bg_depts)} 个部门，正在匹配全面反馈...")

    # 加载BP观察数据
    bp_file = find_bp_file(selected_bg)
    bp_map = {}
    if bp_file:
        print(f"  📋 找到BP观察文件：{os.path.basename(bp_file)}")
        bp_map = load_bp_observations(bp_file)
        print(f"  ✅ BP观察：{len(bp_map)} 人")
    else:
        print(f"  ℹ️  未找到 {selected_bg} 的BP观察文件，将跳过BP观察内容")

    output_rows = []
    total_matched = 0
    total_missing = 0

    SEPARATOR = '\n————————————————————\n'

    def format_person(role, name, shine, expect, bp_text):
        """按新格式组装单人文本"""
        lines = []
        lines.append(f'【管理者：{name}（{role}）】')
        lines.append('<闪光点>')
        lines.append(shine if shine else '无')
        lines.append('</闪光点>')
        lines.append('<更多期待>')
        lines.append(expect if expect else '无')
        lines.append('</更多期待>')
        lines.append('<bp观察>')
        lines.append(bp_text if bp_text else '无')
        lines.append('</bp观察>')
        return '\n'.join(lines)

    for dept_row in bg_depts:
        org_path = dept_row[col_path].strip()
        leader_name = dept_row[col_leader].strip()
        team_str = dept_row[col_team].strip() if col_team < len(dept_row) else ''

        feedback_texts = []

        # 组织负责人
        if leader_name:
            en = extract_english_name(leader_name)
            fb = feedback_map.get(en)
            bp = bp_map.get(en, '')
            display_name = fb['name'] if fb else leader_name
            shine = fb['shine'] if fb and fb['shine'] else '无'
            expect = fb['expect'] if fb and fb['expect'] else '无'
            bp_text = bp if bp else '无'
            feedback_texts.append(format_person('组织负责人', display_name, shine, expect, bp_text))
            if fb:
                total_matched += 1
            else:
                total_missing += 1

        # 部门负责人-1 管理团队
        team_members = parse_team_members(team_str)
        for member in team_members:
            en = extract_english_name(member)
            fb = feedback_map.get(en)
            bp = bp_map.get(en, '')
            display_name = fb['name'] if fb else member
            shine = fb['shine'] if fb and fb['shine'] else '无'
            expect = fb['expect'] if fb and fb['expect'] else '无'
            bp_text = bp if bp else '无'
            feedback_texts.append(format_person('部门负责人-1', display_name, shine, expect, bp_text))
            if fb:
                total_matched += 1
            else:
                total_missing += 1

        # 组装：部门标题 + 各管理者
        header = f'【部门】{org_path}'
        combined = header + '\n\n' + SEPARATOR.join(feedback_texts)
        output_rows.append((org_path, combined))

    print(f"  ✅ 匹配成功：{total_matched} 人")
    print(f"  ⚠️  未找到反馈：{total_missing} 人")
    return output_rows, total_matched, total_missing


def write_output(output_rows, output_file, bg_name):
    """写入Excel输出"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        os.system(f'{sys.executable} -m pip install openpyxl')
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = '全面反馈开放题'

    header_font = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    max_cols = 2
    split_info = []
    overflow_count = 0

    for org_path, combined in output_rows:
        chunks = split_text_to_cells(combined)
        needed = 1 + len(chunks)
        max_cols = max(max_cols, needed)
        split_info.append((org_path, chunks))
        if len(chunks) > 1:
            overflow_count += 1

    if overflow_count > 0:
        print(f"  ⚠️  {overflow_count} 个部门的反馈超出单元格字符上限，已自动切割到多列")

    headers_text = ['组织全路径', '全面反馈']
    for i in range(3, max_cols + 1):
        headers_text.append(f'全面反馈（续{i-1}）')

    for ci, text in enumerate(headers_text, 1):
        cell = ws.cell(row=1, column=ci, value=text)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    data_font = Font(name='微软雅黑', size=10)
    data_align = Alignment(vertical='top', wrap_text=True)

    for i, (path, chunks) in enumerate(split_info, 2):
        cell1 = ws.cell(row=i, column=1, value=path)
        cell1.font = data_font
        cell1.alignment = Alignment(vertical='top', wrap_text=True)
        cell1.border = thin_border

        for ci, chunk in enumerate(chunks):
            cell = ws.cell(row=i, column=2 + ci, value=chunk)
            cell.font = data_font
            cell.alignment = data_align
            cell.border = thin_border

    ws.column_dimensions['A'].width = 45
    for ci in range(2, max_cols + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 100

    ws.freeze_panes = 'A2'
    wb.save(output_file)
    print(f"  📝 已保存：{os.path.basename(output_file)}（{len(output_rows)} 个部门）")


# ─── 主流程 ──────────────────────────────────────────────
def main():
    print("\n📂 正在读取数据文件...\n")

    if not os.path.exists(ORG_DIAG_FILE):
        print(f"❌ 找不到文件：{ORG_DIAG_FILE}")
        return
    if not os.path.exists(FEEDBACK_FILE):
        print(f"❌ 找不到文件：{FEEDBACK_FILE}")
        return

    # 1. 读取组织诊断 - 亮灯明细 (sheet index 1)
    print("  📊 读取组织诊断数据...")
    org_headers, org_rows = read_xlsx_sheet(ORG_DIAG_FILE, sheet_index=1)

    col_bg = find_col(org_headers, 'BG')
    col_path = find_col(org_headers, '组织全路径')
    col_leader = find_col(org_headers, '组织负责人')
    col_team = find_col(org_headers, '部门负责人-1管理者团队')

    if any(c == -1 for c in [col_bg, col_path, col_leader, col_team]):
        print(f"❌ 组织诊断表中找不到必要的列！")
        print(f"   找到的列：{org_headers[:16]}")
        return

    # 2. 读取全面反馈 (sheet index 0)
    print("  📊 读取全面反馈数据...")
    fb_headers, fb_rows = read_xlsx_sheet(FEEDBACK_FILE, sheet_index=0)

    col_fb_name = find_col(fb_headers, '被评估人中英文名')
    col_fb_shine = find_col(fb_headers, '闪光点')
    col_fb_expect = find_col(fb_headers, '更多期待')

    if any(c == -1 for c in [col_fb_name, col_fb_shine, col_fb_expect]):
        print(f"❌ 全面反馈表中找不到必要的列！")
        return

    feedback_map = {}
    for row in fb_rows:
        name_full = row[col_fb_name].strip() if col_fb_name < len(row) else ''
        if not name_full:
            continue
        en_name = extract_english_name(name_full)
        shine = row[col_fb_shine] if col_fb_shine < len(row) else ''
        expect = row[col_fb_expect] if col_fb_expect < len(row) else ''
        feedback_map[en_name] = {
            'name': name_full,
            'shine': shine,
            'expect': expect
        }

    print(f"  ✅ 组织诊断：{len(org_rows)} 条记录")
    print(f"  ✅ 全面反馈：{len(feedback_map)} 人\n")

    # 3. 获取所有不重复的 BG
    all_bgs = sorted(set(row[col_bg].strip() for row in org_rows if row[col_bg].strip()))

    # 4. 判断运行模式：命令行参数 or 交互
    if len(sys.argv) > 1:
        arg = sys.argv[1].strip().lower()
        if arg == 'all':
            selected_bgs = all_bgs
        else:
            # 模糊匹配
            matches = [bg for bg in all_bgs if arg.upper() in bg.upper()]
            if not matches:
                print(f"❌ 没有匹配的BG: {arg}")
                return
            selected_bgs = matches

        print(f"📋 批量处理 {len(selected_bgs)} 个BG")
        for bg in selected_bgs:
            output_file = os.path.join(SCRIPT_DIR, f'全面反馈开放题{re.match(r"([A-Za-z0-9]+)", bg).group(1) if re.match(r"[A-Za-z0-9]+", bg) else bg}.xlsx')
            rows, matched, missing = process_bg(bg, org_rows, col_bg, col_path, col_leader, col_team, feedback_map)
            write_output(rows, output_file, bg)

        print(f"\n🎉 全部完成！")
        return

    # 交互模式
    print(f"📋 共有 {len(all_bgs)} 个BG：")
    for i, bg in enumerate(all_bgs, 1):
        print(f"   {i}. {bg}")

    while True:
        print()
        user_input = input("🔍 请输入BG名称（支持模糊搜索）或序号，输入 all 批量处理，输入 q 退出：").strip()

        if user_input.lower() == 'q':
            print("\n👋 已退出")
            return

        if user_input.lower() == 'all':
            selected_bgs = all_bgs
            print(f"\n✅ 将批量处理所有 {len(all_bgs)} 个BG")
            break

        if user_input.isdigit():
            idx = int(user_input) - 1
            if 0 <= idx < len(all_bgs):
                selected_bgs = [all_bgs[idx]]
                print(f"\n✅ 已选择：{selected_bgs[0]}")
                break
            else:
                print("❌ 序号超出范围，请重试")
                continue

        matches = [bg for bg in all_bgs if user_input.lower() in bg.lower()]
        if len(matches) == 0:
            print("❌ 没有匹配的BG，请重试")
            continue
        elif len(matches) == 1:
            selected_bgs = [matches[0]]
            print(f"\n✅ 匹配到：{matches[0]}")
            break
        else:
            print(f"\n🔎 找到 {len(matches)} 个匹配结果：")
            for i, bg in enumerate(matches, 1):
                print(f"   {i}. {bg}")
            sub = input("   请输入序号确认：").strip()
            if sub.isdigit():
                idx = int(sub) - 1
                if 0 <= idx < len(matches):
                    selected_bgs = [matches[idx]]
                    print(f"\n✅ 已选择：{matches[idx]}")
                    break
            print("❌ 无效输入，请重试")

    for bg in selected_bgs:
        output_file = os.path.join(SCRIPT_DIR, f'全面反馈开放题{re.match(r"([A-Za-z0-9]+)", bg).group(1) if re.match(r"[A-Za-z0-9]+", bg) else bg}.xlsx')
        rows, matched, missing = process_bg(bg, org_rows, col_bg, col_path, col_leader, col_team, feedback_map)
        write_output(rows, output_file, bg)

    print(f"\n🎉 全部完成！")


if __name__ == '__main__':
    main()
