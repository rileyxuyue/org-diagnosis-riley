#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
BP观察模板生成工具
从组织诊断表中按BG提取「组织负责人」和「部门负责人-1管理者团队」列的所有人名，
去重后为每个BG（除WXG）生成一个 Excel 模板文件。
全程本地处理，不上传任何数据。
"""

import os
import re
import sys
import zipfile
import xml.etree.ElementTree as ET

# ─── 配置 ────────────────────────────────────────────────
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PARENT_DIR = os.path.dirname(SCRIPT_DIR)

ORG_DIAG_FILE = os.path.join(PARENT_DIR, '【组织诊断结果】', '组织诊断2025全年最终版.xlsx')
OUTPUT_DIR = SCRIPT_DIR  # 输出到【全面反馈】文件夹

NS = {'s': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}


# ─── xlsx 读取（纯 zipfile 方式，兼容 sharedStrings 和 inlineStr） ──
def read_xlsx_sheet(filepath, sheet_index=0):
    """读取 xlsx 的指定 sheet，返回 (headers, rows)"""
    z = zipfile.ZipFile(filepath)

    # 1. 解析 shared strings
    ss = []
    if 'xl/sharedStrings.xml' in z.namelist():
        tree = ET.parse(z.open('xl/sharedStrings.xml'))
        for si in tree.findall('.//s:si', NS):
            text = ''.join(t.text or '' for t in si.findall('.//s:t', NS))
            ss.append(text)

    # 2. 找到对应 sheet 文件
    sheet_file = f'xl/worksheets/sheet{sheet_index + 1}.xml'
    tree = ET.parse(z.open(sheet_file))
    xml_rows = tree.findall('.//s:sheetData/s:row', NS)

    # 3. 解析列号
    def col_index(ref):
        letters = re.match(r'([A-Z]+)', ref).group(1)
        idx = 0
        for ch in letters:
            idx = idx * 26 + (ord(ch) - ord('A') + 1)
        return idx - 1

    # 4. 读取所有行
    all_rows = []
    max_col = 0
    for row_el in xml_rows:
        cells = row_el.findall('s:c', NS)
        row_dict = {}
        for c in cells:
            ref = c.get('r')
            ci = col_index(ref)
            max_col = max(max_col, ci)
            t = c.get('t')

            # 兼容 inlineStr 格式
            if t == 'inlineStr':
                is_el = c.find('s:is', NS)
                if is_el is not None:
                    row_dict[ci] = ''.join(
                        tt.text or '' for tt in is_el.findall('.//s:t', NS)
                    )
                else:
                    row_dict[ci] = ''
                continue

            v_el = c.find('s:v', NS)
            v = v_el.text if v_el is not None else ''
            if t == 's' and v:
                v = ss[int(v)]
            row_dict[ci] = v
        all_rows.append(row_dict)

    z.close()

    # 5. 转为列表格式
    if not all_rows:
        return [], []
    headers = [all_rows[0].get(i, '') for i in range(max_col + 1)]
    rows = []
    for row_dict in all_rows[1:]:
        row = [row_dict.get(i, '') for i in range(max_col + 1)]
        rows.append(row)

    return headers, rows


def find_col(headers, name):
    """根据列名查找索引"""
    for i, h in enumerate(headers):
        if h.strip() == name.strip():
            return i
    return -1


# ─── 人名解析 ──────────────────────────────────────────────
def parse_person_name(name_str):
    """
    从 'englishname(中文名)' 或 'englishname(中文名)：灯色' 中提取人名部分
    兼容括号不闭合的情况，如 'zhangsan(张三'
    返回 'englishname(中文名)' 格式的字符串，若不闭合则补全括号
    """
    name_str = name_str.strip()
    if not name_str:
        return None
    # 匹配 englishname(中文名) 或 englishname(中文名
    match = re.match(r'(\w+)\(([^)]*)\)?', name_str)
    if match:
        en = match.group(1)
        cn = match.group(2)
        return f'{en}({cn})'
    return None


def parse_team_members(team_str):
    """
    解析「部门负责人-1管理者团队」列
    格式：zhangsan(张三)：绿灯；lisi(李四)：绿灯
    返回人名列表：['zhangsan(张三)', 'lisi(李四)']
    """
    if not team_str or not team_str.strip():
        return []
    members = []
    # 按中文分号或英文分号分割
    parts = re.split(r'[；;]', team_str)
    for part in parts:
        part = part.strip()
        if not part:
            continue
        name = parse_person_name(part)
        if name:
            members.append(name)
    return members


def parse_leader_name(leader_str):
    """
    解析「组织负责人」列
    格式：zhangsan(张三) 或多人用分号分隔
    返回人名列表
    """
    if not leader_str or not leader_str.strip():
        return []
    leaders = []
    parts = re.split(r'[；;]', leader_str)
    for part in parts:
        name = parse_person_name(part)
        if name:
            leaders.append(name)
    return leaders


def extract_bg_short(bg_full):
    """
    从BG全名提取缩写
    如 'CDG企业发展事业群' -> 'CDG'
       'Overseas Functional System' -> 'OFS'
    """
    bg_full = bg_full.strip()
    # 先尝试提取开头的英文+数字部分
    match = re.match(r'([A-Za-z0-9]+)', bg_full)
    if match:
        short = match.group(1)
        # 'Overseas' 开头的特殊处理
        if short == 'Overseas':
            return 'OFS'
        return short
    return bg_full


# ─── 主流程 ──────────────────────────────────────────────
def main():
    print("\n📂 BP观察模板生成工具\n")

    # 检查文件
    if not os.path.exists(ORG_DIAG_FILE):
        print(f"❌ 找不到文件：{ORG_DIAG_FILE}")
        return

    # 1. 读取组织诊断 - 亮灯明细 (sheet index 1)
    print("  📊 读取组织诊断数据（亮灯明细）...")
    org_headers, org_rows = read_xlsx_sheet(ORG_DIAG_FILE, sheet_index=1)

    col_bg = find_col(org_headers, 'BG')
    col_leader = find_col(org_headers, '组织负责人')
    col_team = find_col(org_headers, '部门负责人-1管理者团队')

    if any(c == -1 for c in [col_bg, col_leader, col_team]):
        print(f"❌ 组织诊断表中找不到必要的列！")
        print(f"   BG={col_bg}, 组织负责人={col_leader}, 管理者团队={col_team}")
        print(f"   实际列名：{org_headers[:12]}")
        return

    print(f"  ✅ 读取到 {len(org_rows)} 条记录")

    # 2. 按 BG 分组，收集人名
    bg_names = {}  # {bg_full_name: set_of_names}

    for row in org_rows:
        bg = row[col_bg].strip() if col_bg < len(row) else ''
        if not bg:
            continue

        if bg not in bg_names:
            bg_names[bg] = set()

        # 提取组织负责人
        leader_str = row[col_leader].strip() if col_leader < len(row) else ''
        for name in parse_leader_name(leader_str):
            bg_names[bg].add(name)

        # 提取管理者团队
        team_str = row[col_team].strip() if col_team < len(row) else ''
        for name in parse_team_members(team_str):
            bg_names[bg].add(name)

    print(f"\n📋 共 {len(bg_names)} 个BG：")
    for bg in sorted(bg_names.keys()):
        short = extract_bg_short(bg)
        print(f"   {short}（{bg}）：{len(bg_names[bg])} 人")

    # 3. 排除 WXG，生成 Excel
    print(f"\n📝 开始生成Excel（排除WXG）...\n")

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    except ImportError:
        print("  ⚙️  安装 openpyxl...")
        os.system(f'{sys.executable} -m pip install openpyxl')
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    generated = 0
    for bg_full in sorted(bg_names.keys()):
        bg_short = extract_bg_short(bg_full)

        # 排除 WXG
        if bg_short.upper() == 'WXG':
            print(f"  ⏭️  跳过 {bg_short}（{bg_full}）- 已有成品")
            continue

        names = sorted(bg_names[bg_full])  # 按英文名排序，保证稳定
        if not names:
            print(f"  ⚠️  {bg_short}（{bg_full}）没有人名数据，跳过")
            continue

        # 创建 Excel
        wb = Workbook()
        ws = wb.active
        ws.title = '部门负责人-1BP观察'

        # 表头样式
        header_font = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_align = Alignment(horizontal='center', vertical='center')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # 写表头
        for ci, text in enumerate(['中英文名', 'BP观察'], 1):
            cell = ws.cell(row=1, column=ci, value=text)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        # 数据样式
        data_font = Font(name='微软雅黑', size=10)
        data_align = Alignment(vertical='center')

        # 写数据
        for ri, name in enumerate(names, 2):
            cell_name = ws.cell(row=ri, column=1, value=name)
            cell_name.font = data_font
            cell_name.alignment = data_align
            cell_name.border = thin_border

            cell_bp = ws.cell(row=ri, column=2, value='')
            cell_bp.font = data_font
            cell_bp.alignment = data_align
            cell_bp.border = thin_border

        # 列宽
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 60

        # 冻结首行
        ws.freeze_panes = 'A2'

        # 保存
        output_name = f'{bg_short}部门负责人-1BP观察.xlsx'
        output_path = os.path.join(OUTPUT_DIR, output_name)
        wb.save(output_path)

        print(f"  ✅ {output_name} — {len(names)} 人")
        generated += 1

    print(f"\n🎉 完成！共生成 {generated} 个文件，保存在：")
    print(f"   {OUTPUT_DIR}/")


if __name__ == '__main__':
    main()
